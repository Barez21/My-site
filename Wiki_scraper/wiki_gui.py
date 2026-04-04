#!/usr/bin/env python3
"""Wiki Scraper GUI v8 — spusť a otevři http://localhost:7842"""
import sys, json, time, subprocess, zipfile, io, re, threading, logging, datetime
from pathlib import Path
from logging.handlers import RotatingFileHandler
from flask import Flask, Response, request, jsonify, send_file, session, redirect, url_for
import functools, hashlib, os
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)

# ProxyFix — správně čte X-Forwarded-* headers od Traefiku
# a nastavuje SCRIPT_NAME podle SCRIPT_NAME env proměnné
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.middleware.dispatcher import DispatcherMiddleware

def _make_app():
    import os
    prefix = os.environ.get("SCRIPT_NAME", "")
    if prefix:
        # Namontovat app na prefix
        from werkzeug.wrappers import Response as _R
        def _empty(env, start): return _R("", status=404)(env, start)
        _wrapped = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
        app.wsgi_app = DispatcherMiddleware(_empty, {prefix: _wrapped})
    else:
        app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
    return app
app.secret_key = os.environ.get("SECRET_KEY", "wikiscraper2025-change-me")
app.permanent_session_lifetime = datetime.timedelta(days=30)

# ─── AUTH ─────────────────────────────────────────────────────────────────────
# Uživatelé: načtou se z env proměnné WIKI_USERS nebo z USERS dict níže.
# Formát env: "user1:heslo1,user2:heslo2"
# Hesla zadávej v plaintextu — aplikace je při startu automaticky zahashuje do paměti.

USERS_FILE = Path("users.json")
_users_lock = threading.Lock()

def _load_users() -> dict:
    """Načte uživatele ze souboru users.json."""
    if USERS_FILE.exists():
        try:
            with open(USERS_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # Při prvním spuštění — vytvořit výchozího admina s heslem "admin"
    # Aplikace ho hned při startu zobrazí výzvu ke změně
    default = {"admin": generate_password_hash("admin")}
    _save_users(default)
    return default

def _save_users(users: dict):
    """Uloží uživatele do users.json (hesla jako hashe)."""
    with _users_lock:
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, ensure_ascii=False, indent=2)

USERS = _load_users()

def _check_password(username: str, password: str) -> bool:
    expected = USERS.get(username)
    if not expected:
        return False
    return check_password_hash(expected, password)

def _is_admin(username: str) -> bool:
    """Admin = první uživatel v users.json (ten kdo jako první nastavil heslo)."""
    if not USERS: return False
    return username == next(iter(USERS))


def _redirect(path):
    """Redirect respektující SCRIPT_NAME prefix."""
    prefix = request.environ.get("SCRIPT_NAME", "")
    return redirect(prefix + path)

def login_required(f):
    """Decorator — přesměruje na /login pokud není přihlášen."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login_page", next=request.url))
        return f(*args, **kwargs)
    return wrapper
from datetime import timedelta
app.permanent_session_lifetime = timedelta(days=30)

# ─── LOGOVÁNÍ ────────────────────────────────────────────────────────────────
LOG_DIR = Path("wiki_logs")
LOG_DIR.mkdir(exist_ok=True)

class _JsonFormatter(logging.Formatter):
    """Formátuje log záznamy jako JSON řádky."""
    def format(self, record):
        obj = {
            "ts":      datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00","Z"),
            "level":   record.levelname,
            "logger":  record.name,
            "msg":     record.getMessage(),
        }
        if record.exc_info:
            obj["exc"] = self.formatException(record.exc_info)
        return json.dumps(obj, ensure_ascii=False)

def _make_logger(name: str, filename: str, level=logging.DEBUG) -> logging.Logger:
    logger = logging.getLogger(name)
    logger.setLevel(level)
    if not logger.handlers:
        fh = RotatingFileHandler(
            LOG_DIR / filename, maxBytes=5*1024*1024, backupCount=3, encoding="utf-8"
        )
        fh.setFormatter(_JsonFormatter())
        fh.setLevel(level)
        logger.addHandler(fh)
        # Konzole — INFO a výše
        ch = logging.StreamHandler(sys.stdout)
        ch.setLevel(logging.INFO)
        ch.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S"))
        logger.addHandler(ch)
    return logger

app_log   = _make_logger("wikiscraper.app",   "app.jsonl",   logging.DEBUG)
audit_log = _make_logger("wikiscraper.audit", "audit.jsonl", logging.INFO)

def log_audit(action: str, detail: str = "", session: str = ""):
    """Zapíše do separátního audit logu — stahování, exporty, smazání."""
    audit_log.info(json.dumps({
        "action": action, "detail": detail,
        "session": session,
        "ip": request.remote_addr if request else ""
    }, ensure_ascii=False))

# Per-session job state — klíč je session_prefix (UUID)
_jobs = {}  # {session_prefix: {"running": bool, "proc": proc}}
_jobs_lock = threading.Lock()

def get_job(session_prefix):
    with _jobs_lock:
        if session_prefix not in _jobs:
            _jobs[session_prefix] = {"running": False, "proc": None}
        return _jobs[session_prefix]

OUTPUT_DIR = Path("wiki_output")
OUTPUT_DIR.mkdir(exist_ok=True)
FILE_TTL = 60 * 60  # 60 minut v sekundách

def cleanup_worker():
    """Maže soubory starší než FILE_TTL — běží v pozadí."""
    while True:
        try:
            now = time.time()
            for f in OUTPUT_DIR.iterdir():
                try:
                    if now - f.stat().st_mtime > FILE_TTL:
                        f.unlink()
                except Exception:
                    pass
        except Exception:
            pass
        time.sleep(300)  # každých 5 minut

_cleanup_thread = threading.Thread(target=cleanup_worker, daemon=True)
_cleanup_thread.start()

HTML = r"""<!DOCTYPE html>
<html lang="cs">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="theme-color" content="#0f1117" id="metaThemeColor">
<title>Wiki Scraper</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/canvas-confetti@1.9.2/dist/confetti.browser.min.js"></script>
<style>
/* ── TOKENS ──────────────────────────────────────────────────────────────── */
:root {
  color-scheme: dark;
  --bg:        #0f1117;
  --surface:   #171b24;
  --surface2:  #1e2330;
  --border:    #2a2f3d;
  --border2:   #363c4f;
  --text:      #e2e8f0;
  --text2:     #8892a4;
  --text3:     #4a5568;
  --accent:    #6ee7b7;
  --accent2:   #60a5fa;
  --warn:      #fbbf24;
  --danger:    #f87171;
  --accent-tint:  rgba(110,231,183,.08);
  --accent2-tint: rgba(96,165,250,.08);
  --danger-tint:  rgba(248,113,113,.08);
  --chart-grid:   #2a2f3d;
  --chart-tick:   #4a5568;
  --shadow:    0 2px 8px rgba(0,0,0,.35);
  --shadow-lg: 0 8px 32px rgba(0,0,0,.5);
  --mono: 'IBM Plex Mono', monospace;
  --sans: 'IBM Plex Sans', sans-serif;
  --radius: 8px;
  --radius-sm: 5px;
}

/* ── LIGHT MODE ───────────────────────────────────────────────────────────── */
html.light {
  color-scheme: light;
  --bg:        #f7f6f3;   /* warm off-white */
  --surface:   #ffffff;
  --surface2:  #eeecea;   /* viditelné oddělení od surface */
  --border:    #dddbd8;   /* warm grey */
  --border2:   #c4c2bf;
  --text:      #111110;   /* near-black warm */
  --text2:     #4a4945;   /* readable secondary */
  --text3:     #8f8d8a;   /* tertiary — stale čitelné */
  --accent:    #0a6640;   /* tmavší emerald = lepší kontrast na bílé */
  --accent2:   #1a4ecf;   /* solid blue */
  --warn:      #924d0a;
  --danger:    #a31515;
  --accent-tint:  rgba(10,102,64,.09);
  --accent2-tint: rgba(26,78,207,.09);
  --danger-tint:  rgba(163,21,21,.08);
  --chart-grid:   #e5e3e0;
  --chart-tick:   #4a4945;
  --shadow:    0 2px 8px rgba(0,0,0,.08);
  --shadow-lg: 0 8px 32px rgba(0,0,0,.14);
}

/* ── RESET & BASE ─────────────────────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html, body { height: 100%; overflow: hidden; }
body {
  background: var(--bg); color: var(--text);
  font-family: var(--sans); font-size: 14px; line-height: 1.5;
  touch-action: manipulation;
  -webkit-tap-highlight-color: transparent;
}

/* ── REDUCED MOTION ───────────────────────────────────────────────────────── */

/* ════════════════════════════════════════════════════════════════════════════
   MOBILNÍ RESPONZIVITA — breakpoint 768px
   Filosofie: sidebar se stane spodní navigací (bottom drawer), hlavní obsah
   zabírá celou výšku, taby se scrollují horizontálně.
   ════════════════════════════════════════════════════════════════════════════ */

@media (max-width: 768px) {

  /* ── Základní layout ─────────────────────────────────────────────────────── */
  .app {
    grid-template-rows: 48px auto auto 1fr 56px; /* přidat spodní nav lištu */
    height: 100dvh; /* dynamic viewport height — správně na iOS Safari */
  }
  .body {
    grid-template-columns: 1fr; /* sidebar schovat, main na celou šířku */
    grid-column: 1 / -1;
    position: relative;
  }

  /* ── Header ─────────────────────────────────────────────────────────────── */
  .header { padding: 0 12px; gap: 8px; }
  .header-logo { font-size: 15px; }
  .header-badge { display: none; } /* skrýt verzi badge — šetřit místo */
  #sessionBadge { display: none; }
  .stepper { display: none; } /* stepper na mobilu zabírá moc místa */
  .logout-btn { padding: 3px 8px; font-size: 10px; }

  /* ── Progress sekce ─────────────────────────────────────────────────────── */
  .progress-section { padding: 6px 12px; }
  .progress-name { font-size: 9px; min-width: 80px; }
  .current-task { padding: 4px 6px; min-height: 24px; }
  .task-phase { font-size: 9px; padding: 2px 5px; }
  .task-name { font-size: 10px; }

  /* ── Stats bar ──────────────────────────────────────────────────────────── */
  .stats-bar { gap: 8px; padding: 4px 12px; overflow-x: auto; }
  .stat-val { font-size: 16px; }
  .stat-lbl { font-size: 9px; }

  /* ── Tabs ───────────────────────────────────────────────────────────────── */
  .tabs { overflow-x: auto; -webkit-overflow-scrolling: touch; gap: 0;
          scrollbar-width: none; padding: 0 4px; }
  .tabs::-webkit-scrollbar { display: none; }
  .tab-btn { white-space: nowrap; padding: 8px 12px; font-size: 11px; flex-shrink: 0; }

  /* ── Main content ───────────────────────────────────────────────────────── */
  .main { height: calc(100dvh - 48px - var(--progress-h, 80px)); }
  .tab-panel { overflow-y: auto; -webkit-overflow-scrolling: touch; }

  /* ── Data tabulka ───────────────────────────────────────────────────────── */
  .dt-toolbar { flex-wrap: wrap; gap: 4px; padding: 8px; }
  .dt-toolbar input { min-width: 0; }
  /* Skrýt méně důležité sloupce na malém displeji */
  .col-c3, .col-c4 { display: none; } /* kategorie, délka */
  .td-title { max-width: none; }
  .td-text { max-width: 160px; }
  thead th { padding: 6px 8px; font-size: 9px; }
  tbody td { padding: 5px 8px; font-size: 10px; }

  /* ── Sidebar — přesunut do bottom sheet ─────────────────────────────────── */
  aside.sidebar {
    position: fixed;
    bottom: 56px; /* nad spodní nav lištou */
    left: 0; right: 0;
    height: auto;
    max-height: 80dvh;
    border-radius: 16px 16px 0 0;
    border-top: 1px solid var(--border2);
    box-shadow: 0 -8px 32px rgba(0,0,0,.4);
    z-index: 200;
    transform: translateY(100%);
    transition: transform .3s cubic-bezier(.4,0,.2,1);
    overflow: hidden;
    /* Schovat resize handle na mobilu */
  }
  aside.sidebar.mob-open {
    transform: translateY(0);
  }
  .sidebar-scroll { padding: 12px; max-height: calc(80dvh - 56px); }
  .resize-handle-x { display: none; }

  /* ── Spodní nav lišta ───────────────────────────────────────────────────── */
  #mobileNav {
    display: flex !important;
    position: fixed;
    bottom: 0; left: 0; right: 0;
    height: 56px;
    background: var(--surface);
    border-top: 1px solid var(--border);
    z-index: 250;
    align-items: stretch;
  }
  .mob-nav-btn {
    flex: 1;
    display: flex; flex-direction: column;
    align-items: center; justify-content: center;
    gap: 3px;
    font-size: 9px; color: var(--text3);
    background: transparent; border: none;
    cursor: pointer;
    -webkit-tap-highlight-color: transparent;
    transition: color .15s;
  }
  .mob-nav-btn.active { color: var(--accent); }
  .mob-nav-btn .nav-icon { font-size: 20px; line-height: 1; }

  /* ── Sidebar overlay backdrop ───────────────────────────────────────────── */
  #mobSidebarBackdrop {
    display: none;
    position: fixed; inset: 0;
    background: rgba(0,0,0,.5);
    z-index: 190;
  }
  #mobSidebarBackdrop.on { display: block; }

  /* ── Review dialog ──────────────────────────────────────────────────────── */
  .review-modal-inner {
    width: 100vw !important;
    max-width: 100vw !important;
    height: 100dvh !important;
    max-height: 100dvh !important;
    border-radius: 0 !important;
  }
  .review-cols { flex-direction: column; gap: 8px; }
  .review-col { min-width: 0 !important; width: 100% !important; max-height: 280px; }

  /* ── Export box ─────────────────────────────────────────────────────────── */
  .export-details-body { gap: 6px; }
  .export-btn { font-size: 11px; padding: 8px 10px; }

  /* ── Overview karty ─────────────────────────────────────────────────────── */
  .sg-card { padding: 10px 8px; }
  .sg-val { font-size: 20px; }
  .sg-lbl { font-size: 9px; }

  /* ── Modals — full screen na mobilu ────────────────────────────────────── */
  .modal-inner, .ib-popup {
    width: 96vw !important;
    max-width: 96vw !important;
    max-height: 85dvh !important;
  }
  .tbl-popup { width: 96vw; left: 2vw !important; }

  /* ── Console ────────────────────────────────────────────────────────────── */
  .console-body.open { max-height: 120px; }

  /* ── Tlačítka sidebar footer ────────────────────────────────────────────── */
  .sidebar-footer { padding: 8px 12px; }
  .btn { font-size: 12px; padding: 9px 12px; }
}

/* Extra malé displeje — 380px a méně */
@media (max-width: 380px) {
  .tab-btn { padding: 7px 9px; font-size: 10px; }
  .col-c5, .col-c6 { display: none; } /* skrýt quality a infobox count */
}

/* Landscape mobile — horizontální orientace */
@media (max-width: 768px) and (orientation: landscape) {
  .app { grid-template-rows: 40px auto 1fr 48px; }
  .header { height: 40px; }
  aside.sidebar { max-height: 90dvh; bottom: 48px; }
  #mobileNav { height: 48px; }
  .mob-nav-btn .nav-icon { font-size: 16px; }
}

@media (prefers-reduced-motion: reduce) {
  *, *::before, *::after {
    animation-duration: .01ms !important;
    animation-iteration-count: 1 !important;
    transition-duration: .01ms !important;
  }
}

/* ── FOCUS RING ───────────────────────────────────────────────────────────── */
:focus-visible {
  outline: 2px solid var(--accent2);
  outline-offset: 2px;
  border-radius: 3px;
}

/* ── THEME BUTTON ─────────────────────────────────────────────────────────── */
.theme-btn {
  background: transparent; border: 1px solid var(--border);
  border-radius: var(--radius-sm); padding: 4px 8px;
  cursor: pointer; font-size: 14px; color: var(--text2);
  transition: border-color .15s, color .15s; line-height: 1;
}
.theme-btn:hover { border-color: var(--border2); color: var(--text); }

/* ── STEPPER ──────────────────────────────────────────────────────────────── */
.stepper {
  display: flex; align-items: center; gap: 0;
  padding: 0 20px; flex: 1; justify-content: center;
}
.step {
  display: flex; align-items: center; gap: 6px;
  font-family: var(--mono); font-size: 11px; color: var(--text3);
  padding: 4px 10px; border-radius: 20px; transition: color .2s, background .2s;
}
.step.done  { color: var(--accent); }
.step.active{ color: var(--text); background: var(--surface2); }
.step-dot {
  width: 18px; height: 18px; border-radius: 50%; flex-shrink: 0;
  border: 1.5px solid var(--text3);
  display: flex; align-items: center; justify-content: center;
  font-size: 9px; transition: background .2s, border-color .2s, color .2s;
}
.step.done   .step-dot { background: var(--accent);  border-color: var(--accent);  color: var(--bg); }
.step.active .step-dot { background: var(--surface2); border-color: var(--text);   color: var(--text); }
.step-line { width: 24px; height: 1px; background: var(--border); flex-shrink: 0; transition: background .3s; }
.step-line.done { background: var(--accent); }

/* ── TOOLTIPS — JS řízený, viz initTooltips() ─────────────────────────────── */
[data-tip] { position: relative; }
#globalTooltip {
  position: fixed; z-index: 9999; pointer-events: none;
  background: var(--surface2); border: 1px solid var(--border2);
  color: var(--text2); font-family: var(--mono); font-size: 10px; line-height: 1.6;
  padding: 7px 11px; border-radius: 5px; white-space: pre-wrap;
  box-shadow: 0 4px 16px rgba(0,0,0,.4);
  max-width: 260px; opacity: 0; transition: opacity .12s;
  visibility: hidden;
}
#globalTooltip.visible { opacity: 1; visibility: visible; }

/* ── DRAG-DROP HIGHLIGHT ──────────────────────────────────────────────────── */
#url.drag-over {
  border-color: var(--accent2);
  background: var(--accent2-tint);
  box-shadow: 0 0 0 2px var(--accent2-tint);
}





/* ── COLLAPSIBLE SECTIONS ─────────────────────────────────────────────────── */
.collapsible-section { }
.sec-header {
  display: flex; align-items: center; justify-content: space-between;
  cursor: pointer; user-select: none;
}
.sec-header:hover { color: var(--text); }
.sec-caret {
  font-size: 10px; color: var(--text3); transition: transform .2s;
  flex-shrink: 0; margin-left: 4px;
}
.collapsible-section.collapsed .sec-caret { transform: rotate(-90deg); }
.sec-body {
  overflow: hidden;
  max-height: 2000px;
  transition: max-height .25s ease, opacity .2s ease;
  opacity: 1;
}
.collapsible-section.collapsed .sec-body {
  max-height: 0;
  opacity: 0;
}
/* Review sloupce maji flex layout */
.review-col .sec-body { max-height:none !important; flex:1; overflow:hidden; transition:opacity .2s; }
.review-col.collapsed .sec-body { opacity:0; flex:0 !important; max-height:0 !important; pointer-events:none; }
/* Viz karty */
.viz-card { display: flex; flex-direction: column; }
.viz-card-title {
  display: flex; align-items: center;
  cursor: pointer; user-select: none;
}
.viz-card-title:hover { color: var(--text); }
.viz-card-body { overflow: hidden; max-height: 2000px; transition: max-height .25s ease, opacity .2s ease; opacity: 1; }
.viz-card.collapsed .viz-card-body { max-height: 0; opacity: 0; }
.viz-card.collapsed .viz-card-title .sec-caret { transform: rotate(-90deg); }


/* ── PAUSE / STOP GROUP ──────────────────────────────────────────────────── */
#btnStopGroup { display: none; flex-direction: column; gap: 6px; }
#btnStopGroup.on { display: flex !important; }


/* ── IMAGES & TABLES ─────────────────────────────────────────────────────── */
.img-strip {
  display:flex; gap:4px; flex-wrap:wrap; margin-top:4px;
}
.img-thumb {
  width:40px; height:30px; object-fit:cover; border-radius:3px;
  border:1px solid var(--border); cursor:pointer;
  transition:transform .15s, box-shadow .15s;
}
.img-thumb:hover { transform:scale(1.05); box-shadow:0 2px 8px rgba(0,0,0,.3); }
.tbl-badge {
  display:inline-flex; align-items:center; gap:3px;
  font-family:var(--mono); font-size:9px;
  padding:1px 5px; border-radius:10px; cursor:pointer;
  background:rgba(96,165,250,.1); color:var(--accent2);
  border:1px solid rgba(96,165,250,.2);
  margin-left:5px; vertical-align:middle;
  transition:background .15s;
}
.tbl-badge:hover { background:rgba(96,165,250,.2); }

/* Lightbox pro plný obrázek */
.img-lightbox {
  display:none; position:fixed; inset:0; background:rgba(0,0,0,.85);
  z-index:800; align-items:center; justify-content:center; flex-direction:column;
  gap:12px;
}
.img-lightbox.on { display:flex; }
.img-lightbox img { max-width:90vw; max-height:80vh; object-fit:contain; border-radius:6px; }
.img-lightbox-cap { font-family:var(--mono); font-size:11px; color:#ddd; max-width:600px; text-align:center; }
.img-lightbox-meta { font-family:var(--mono); font-size:10px; color:#aaa; }

/* Table popup */
.tbl-popup {
  position:fixed; z-index:500; background:var(--surface);
  border:1px solid var(--border2); border-radius:var(--radius);
  max-width:680px; max-height:60vh; overflow:auto;
  box-shadow:0 8px 32px rgba(0,0,0,.5); padding:12px;
  display:none;
}
.tbl-popup.on { display:block; }
.tbl-popup table { width:100%; border-collapse:collapse; font-family:var(--mono); font-size:11px; }
.tbl-popup th { background:var(--surface2); padding:5px 8px; border:1px solid var(--border); text-align:left; font-weight:600; color:var(--text2); }
.tbl-popup td { padding:4px 8px; border:1px solid var(--border); color:var(--text2); vertical-align:top; }
.tbl-popup tr:nth-child(even) td { background:rgba(255,255,255,.02); }
.tbl-caption { font-size:12px; font-weight:600; color:var(--text); margin-bottom:8px; }
.tbl-nav { display:flex; gap:6px; margin-bottom:8px; flex-wrap:wrap; }

/* ── EXPORT BOX UI ───────────────────────────────────────────────────────── */
.export-group { margin-bottom:8px; }
.export-group-label {
  font-family:var(--mono); font-size:9px; text-transform:uppercase;
  letter-spacing:.07em; color:var(--text3); margin-bottom:4px;
}
.export-btn { width:100%; justify-content:flex-start; margin-bottom:4px; font-size:12px; }
.export-details {
  margin-bottom:6px; border:1px solid var(--border);
  border-radius:var(--radius-sm); overflow:hidden;
}
.export-summary {
  font-family:var(--mono); font-size:11px; color:var(--text2);
  padding:7px 10px; cursor:pointer; list-style:none;
  background:var(--surface); user-select:none;
  display:flex; align-items:center; gap:6px;
  transition:background .1s;
}
.export-summary:hover { background:var(--surface2); }
.export-summary::-webkit-details-marker { display:none; }
.export-details[open] .export-summary { border-bottom:1px solid var(--border); }
.export-details-body { padding:8px; display:flex; flex-direction:column; gap:4px; background:var(--bg); }

/* ── CATEGORY TREE MODAL ─────────────────────────────────────────────────── */
.tree-backdrop {
  display:none; position:fixed; inset:0; background:rgba(0,0,0,.7);
  z-index:350; align-items:center; justify-content:center;
}
.tree-backdrop.on { display:flex; }
.tree-modal {
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:700px; max-width:96vw; max-height:88vh;
  display:flex; flex-direction:column; box-shadow:0 16px 48px rgba(0,0,0,.6);
}
.tree-header { padding:16px 20px; border-bottom:1px solid var(--border); flex-shrink:0; }
.tree-title { font-size:15px; font-weight:600; margin-bottom:2px; }
.tree-sub { font-family:var(--mono); font-size:11px; color:var(--text3); }
.tree-toolbar {
  padding:8px 16px; border-bottom:1px solid var(--border);
  display:flex; gap:8px; align-items:center; flex-shrink:0; background:var(--bg);
  flex-wrap:wrap;
}
.tree-body { flex:1; overflow-y:auto; padding:8px 0; }
.tree-body::-webkit-scrollbar { width:4px; }
.tree-body::-webkit-scrollbar-thumb { background:var(--border2); }
.tree-node {
  display:flex; align-items:center; gap:0;
  padding:3px 0; font-size:12px; user-select:none;
}
.tree-indent { flex-shrink:0; }
.tree-toggle {
  width:20px; height:20px; display:flex; align-items:center; justify-content:center;
  cursor:pointer; color:var(--text3); font-size:10px; flex-shrink:0;
  border-radius:3px; transition:background .1s;
}
.tree-toggle:hover { background:var(--surface2); color:var(--text); }
.tree-toggle.leaf { cursor:default; color:transparent; }
.tree-cb-wrap {
  width:24px; height:24px; display:flex; align-items:center; justify-content:center;
  cursor:pointer; flex-shrink:0;
}
.tree-cb {
  width:13px; height:13px; border-radius:3px; border:1.5px solid var(--border2);
  display:flex; align-items:center; justify-content:center;
  font-size:8px; transition:all .12s;
}
.tree-cb.checked { background:var(--accent); border-color:var(--accent); color:var(--bg); }
.tree-cb.partial { background:rgba(110,231,183,.3); border-color:var(--accent); }
.tree-label {
  flex:1; padding:2px 6px; cursor:pointer; border-radius:3px;
  color:var(--text2); transition:background .1s; white-space:nowrap;
  overflow:hidden; text-overflow:ellipsis;
}
.tree-label:hover { background:var(--surface2); color:var(--text); }
.tree-count {
  font-family:var(--mono); font-size:10px; color:var(--text3);
  padding:1px 6px; border-radius:10px; background:var(--surface2);
  flex-shrink:0; margin-right:8px;
}
.tree-node.selected > .tree-label { color:var(--accent); }
.tree-footer {
  padding:12px 16px; border-top:1px solid var(--border);
  display:flex; gap:8px; align-items:center; flex-shrink:0; flex-wrap:wrap;
}
.tree-summary { font-family:var(--mono); font-size:11px; color:var(--text2); flex:1; }

.logout-btn {
  margin-left: auto;
  font-family: var(--mono); font-size: 10px; color: var(--text3);
  text-decoration: none; padding: 3px 10px;
  border: 1px solid var(--border); border-radius: var(--radius-sm);
  transition: color .15s, border-color .15s;
}
.logout-btn:hover { color: var(--danger); border-color: var(--danger); }

/* ── RESIZE HANDLES ───────────────────────────────────────────────────────── */
.resize-handle-x {
  position: absolute; right: 0; top: 0; bottom: 0;
  width: 5px; cursor: col-resize; z-index: 100;
  background: transparent; transition: background .15s;
}
.resize-handle-x:hover, .resize-handle-x.dragging {
  background: var(--accent2);
  opacity: .5;
}
.sidebar { position: relative; }

.resize-handle-y {
  height: 6px; cursor: row-resize; flex-shrink: 0;
  background: var(--border); transition: background .15s;
  position: relative; z-index: 10;
  border-top: 1px solid var(--border2);
}
.resize-handle-y:hover, .resize-handle-y.dragging {
  background: var(--accent2);
  opacity: .6;
}

/* Tabulka — column resize */
table { table-layout: fixed; width: 100%; }
th { position: relative; overflow: hidden; }
.th-resize {
  position: absolute; right: 0; top: 0; bottom: 0;
  width: 4px; cursor: col-resize; background: transparent;
  transition: background .12s;
}
.th-resize:hover, .th-resize.dragging { background: var(--accent2); opacity: .6; }

/* Výchozí šířky sloupců tabulky */
.col-c0 { width: 38px; }
.col-c1 { width: 26%; }
.col-c2 { width: 28%; }
.col-c3 { width: 18%; }
.col-c4 { width: 52px; }
.col-c5 { width: 42px; }
.col-c6 { width: 72px; }
.col-c7 { width: 30px; }

/* Disable text select during drag */
body.resizing { user-select: none; cursor: col-resize !important; }
body.resizing-y { user-select: none; cursor: row-resize !important; }

/* ── COPY BUTTON ──────────────────────────────────────────────────────────── */
.copy-btn {
  font-size: 11px; padding: 2px 5px; cursor: pointer;
  color: var(--text3); background: transparent; border: none;
  transition: color .15s; border-radius: 3px;
}
.copy-btn:hover { color: var(--accent2); }
.copy-btn.copied { color: var(--accent); }

/* ── RESULT FILTERS ───────────────────────────────────────────────────────── */
.result-filters {
  display: flex; gap: 6px; flex-wrap: wrap; align-items: center;
}
.rf-chip {
  font-family: var(--mono); font-size: 10px;
  padding: 3px 9px; border-radius: 20px;
  border: 1px solid var(--border); background: var(--bg);
  color: var(--text3); cursor: pointer; transition: border-color .15s, color .15s, background .15s;
}
.rf-chip:hover { border-color: var(--border2); color: var(--text2); }
.rf-chip.on { border-color: var(--accent2); color: var(--accent2); background: var(--accent2-tint); }

/* ── TEST URL ─────────────────────────────────────────────────────────────── */
.test-result {
  font-family: var(--mono); font-size: 11px; margin-top: 4px;
  padding: 6px 9px; border-radius: 5px; display: none;
  background: var(--surface2); line-height: 1.5;
}
.test-result.on { display: block; }
.test-result.ok  { border-left: 3px solid var(--accent); }
.test-result.err { border-left: 3px solid var(--danger); }

/* ── BLACKLIST ────────────────────────────────────────────────────────────── */
.blacklist-wrap { position: relative; }
.bl-count {
  font-family: var(--mono); font-size: 10px; color: var(--warn);
  margin-top: 3px; display: none;
}
.bl-count.on { display: block; }

/* ── REVIEW ANIMATION ────────────────────────────────────────────────────── */
.review-reveal {
  display: none; flex-direction: column;
  align-items: center; justify-content: center;
  padding: 40px 20px; gap: 12px;
}
.review-reveal.on { display: flex; }
.rv-number {
  font-size: 72px; font-weight: 800; color: var(--accent);
  font-family: var(--sans); line-height: 1;
  animation: countUp .6s cubic-bezier(.22,.61,.36,1) both;
}
@keyframes countUp {
  from { transform: scale(.5); opacity: 0; }
  to   { transform: scale(1);  opacity: 1; }
}
.rv-label { font-family: var(--mono); font-size: 13px; color: var(--text2); }
.rv-sub   { font-family: var(--mono); font-size: 11px; color: var(--text3); }
.rv-btn   { margin-top: 8px; }

/* ── STATS GRID ───────────────────────────────────────────────────────────── */
.stats-grid {
  display: grid; grid-template-columns: repeat(auto-fill, minmax(120px,1fr));
  gap: 8px; margin-bottom: 12px;
}
.sg-card {
  background: var(--surface2); border: 1px solid var(--border);
  border-radius: var(--radius-sm); padding: 10px 12px;
}
.sg-val { font-size: 22px; font-weight: 700; font-family: var(--mono); color: var(--accent); }
.sg-lbl { font-size: 10px; font-family: var(--mono); color: var(--text3); text-transform: uppercase; letter-spacing: .05em; margin-top: 2px; }
.intro-histogram { margin-top: 8px; }
.hist-bar-wrap { display: flex; align-items: center; gap: 6px; margin-bottom: 3px; }
.hist-label { font-family: var(--mono); font-size: 10px; color: var(--text3); min-width: 60px; }
.hist-bar { height: 8px; border-radius: 2px; background: var(--accent2); min-width: 2px; transition: width .4s; }
.hist-count { font-family: var(--mono); font-size: 10px; color: var(--text3); }

/* ── PROJECT TABS ─────────────────────────────────────────────────────────── */
/* ── COLUMN PICKER MODAL ─────────────────────────────────────────────────── */
/* ── INFO MODAL ───────────────────────────────────────────────────────────── */
.info-backdrop {
  display:none; position:fixed; inset:0; background:rgba(0,0,0,.7);
  z-index:400; align-items:center; justify-content:center;
}
.info-backdrop.on { display:flex; }
.info-modal {
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:640px; max-width:95vw; max-height:85vh;
  display:flex; flex-direction:column; box-shadow:0 16px 48px rgba(0,0,0,.6);
}
.info-tabs { display:flex; border-bottom:1px solid var(--border); flex-shrink:0; }
.info-tab {
  padding:10px 20px; font-family:var(--mono); font-size:12px;
  color:var(--text3); cursor:pointer; border-bottom:2px solid transparent;
  transition: color .15s, border-color .15s;
}
.info-tab.active { color:var(--text); border-bottom-color:var(--accent2); }
.info-body { flex:1; overflow-y:auto; padding:24px; }
.info-body::-webkit-scrollbar { width:3px; }
.info-body::-webkit-scrollbar-thumb { background:var(--border2); }
.info-body h2 { font-size:15px; color:var(--accent); margin-bottom:4px; }
.info-body h3 { font-size:12px; font-weight:600; margin:16px 0 6px; color:var(--text2); text-transform:uppercase; letter-spacing:.06em; }
.info-body p  { font-size:13px; color:var(--text2); line-height:1.7; margin-bottom:10px; }
.info-body kbd {
  background:var(--surface2); border:1px solid var(--border2);
  border-radius:3px; padding:1px 5px; font-family:var(--mono); font-size:11px;
}
.info-body .cl-entry { margin-bottom:14px; padding-bottom:14px; border-bottom:1px solid var(--border); }
.info-body .cl-version { font-family:var(--mono); font-size:11px; color:var(--accent2); font-weight:600; }
.info-body .cl-date { font-family:var(--mono); font-size:10px; color:var(--text3); margin-left:8px; }
.info-body .cl-items { margin-top:6px; font-size:12px; color:var(--text2); line-height:1.8; }
.info-footer { padding:12px 20px; border-top:1px solid var(--border); text-align:right; flex-shrink:0; }
.info-credit {
  font-family:var(--mono); font-size:10px; color:var(--text3);
  padding:10px 0 0; text-align:center;
}

/* ── NOTIFICATION CENTER ─────────────────────────────────────────────────── */
.notif-bell {
  position:relative; background:transparent; border:1px solid var(--border);
  border-radius:var(--radius-sm); padding:4px 8px; cursor:pointer;
  font-size:14px; color:var(--text2); transition: border-color .15s, color .15s; line-height:1;
}
.notif-bell:hover { border-color:var(--border2); color:var(--text); }
.notif-badge {
  position:absolute; top:-4px; right:-4px; background:var(--danger);
  color:white; font-family:var(--mono); font-size:9px; font-weight:700;
  width:16px; height:16px; border-radius:50%; display:none;
  align-items:center; justify-content:center;
}
.notif-badge.on { display:flex; }
.notif-panel {
  position:absolute; top:calc(100% + 8px); right:0; z-index:500;
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:320px; max-height:400px;
  box-shadow:0 8px 24px rgba(0,0,0,.5); display:none; flex-direction:column;
}
.notif-panel.on { display:flex; }
.notif-panel-header {
  padding:10px 14px; border-bottom:1px solid var(--border);
  display:flex; align-items:center; gap:8px;
  font-family:var(--mono); font-size:11px; font-weight:600;
}
.notif-list { flex:1; overflow-y:auto; }
.notif-item {
  padding:10px 14px; border-bottom:1px solid var(--border);
  font-size:12px; line-height:1.5;
}
.notif-item-title { font-weight:600; color:var(--text); margin-bottom:2px; }
.notif-item-body { color:var(--text3); font-family:var(--mono); font-size:10px; }
.notif-item-time { color:var(--text3); font-family:var(--mono); font-size:10px; float:right; }
.notif-empty { padding:20px; text-align:center; font-family:var(--mono); font-size:11px; color:var(--text3); }
.notif-bell-wrap { position:relative; }

/* ── ONBOARDING OVERLAY ──────────────────────────────────────────────────── */
.onboard-backdrop {
  display:none; position:fixed; inset:0; background:rgba(0,0,0,.85);
  z-index:1000; align-items:center; justify-content:center;
}
.onboard-backdrop.on { display:flex; }
.onboard-card {
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:500px; max-width:95vw;
  padding:32px; text-align:center; box-shadow:0 20px 60px rgba(0,0,0,.7);
}
.onboard-icon { font-size:48px; margin-bottom:12px; }
.onboard-title { font-size:22px; font-weight:700; margin-bottom:8px; }
.onboard-sub { font-size:13px; color:var(--text2); margin-bottom:24px; line-height:1.6; }
.onboard-steps { display:flex; gap:12px; margin-bottom:28px; }
.onboard-step {
  flex:1; background:var(--surface2); border-radius:8px; padding:14px 10px;
  font-size:12px; color:var(--text2); line-height:1.5;
}
.onboard-step-num { font-size:20px; font-weight:800; color:var(--accent); display:block; margin-bottom:4px; }

/* ── MAP TAB ──────────────────────────────────────────────────────────────── */
#panelMap { display:none; flex-direction:column; }
#panelMap.on { display:flex; }
#mapContainer { flex:1; min-height:0; width:100%; }
.map-stats {
  position:absolute; top:8px; right:8px; z-index:500;
  background:var(--surface); border:1px solid var(--border2);
  border-radius:5px; padding:6px 10px;
  font-family:var(--mono); font-size:11px; color:var(--text2);
}

/* ── TAG CLOUD ────────────────────────────────────────────────────────────── */
.tag-cloud { display:flex; flex-wrap:wrap; gap:8px; padding:16px; }
.tag-chip {
  display:inline-flex; align-items:center; gap:4px;
  padding:4px 12px; border-radius:20px; cursor:pointer;
  font-family:var(--mono); font-size:11px; font-weight:600;
  border:1px solid transparent; transition: opacity .15s, transform .15s, box-shadow .15s;
}
.tag-chip:hover { opacity:.8; transform:scale(1.05); }
.tag-chip.active { box-shadow:0 0 0 2px var(--accent); }

/* ── LIVE CHART ───────────────────────────────────────────────────────────── */
.live-chart-wrap {
  position:absolute; bottom:52px; right:8px; z-index:10;
  background:var(--surface); border:1px solid var(--border);
  border-radius:6px; padding:8px; width:200px;
  font-family:var(--mono); font-size:10px; color:var(--text3);
  display:none;
}
.live-chart-wrap.on { display:block; }
#liveSpeedChart { width:100% !important; height:60px !important; }

/* ── GLOBAL SEARCH ────────────────────────────────────────────────────────── */
.global-search-wrap {
  position:relative; margin-left:auto; margin-right:8px;
}
.global-search-input {
  background:var(--surface2); border:1px solid var(--border);
  border-radius:var(--radius-sm); color:var(--text);
  font-family:var(--mono); font-size:11px; padding:5px 30px 5px 10px;
  width:200px; outline:none; transition: border-color .2s, width .2s, box-shadow .2s;
}
.global-search-input:focus-visible { border-color:var(--accent2); width:280px; box-shadow:0 0 0 3px var(--accent2-tint); outline:none; }
.global-search-results {
  position:absolute; top:calc(100% + 4px); right:0; z-index:600;
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:360px; max-height:360px;
  overflow-y:auto; box-shadow:var(--shadow-lg); display:none;
}
.global-search-results.on { display:block; }
.gs-item {
  padding:10px 14px; border-bottom:1px solid var(--border);
  cursor:pointer; transition:background .1s;
}
.gs-item:hover { background:var(--surface2); }
.gs-item-title { font-size:12px; font-weight:600; color:var(--text); }
.gs-item-meta  { font-family:var(--mono); font-size:10px; color:var(--text3); margin-top:2px; }
.gs-item-snip  { font-size:11px; color:var(--text2); margin-top:3px; }
.gs-section    { padding:6px 14px; font-family:var(--mono); font-size:10px;
  color:var(--text3); background:var(--surface2); text-transform:uppercase;
  letter-spacing:.06em; border-bottom:1px solid var(--border); position:sticky; top:0; }

/* ── BULK SELECT ──────────────────────────────────────────────────────────── */
.bulk-toolbar {
  display:none; align-items:center; gap:8px;
  padding:6px 12px; background:var(--accent2-tint);
  border-bottom:1px solid var(--border2); flex-shrink:0;
  font-family:var(--mono); font-size:11px; color:var(--accent2);
}
.bulk-toolbar.on { display:flex; }
.td-sel { width:32px; text-align:center; }
.row-checkbox { width:14px; height:14px; cursor:pointer; accent-color:var(--accent2); }

/* ── VALIDATION PANEL ────────────────────────────────────────────────────── */
.valid-issue {
  display:flex; align-items:baseline; gap:10px; padding:8px 0;
  border-bottom:1px solid var(--border); font-size:12px;
}
.valid-count {
  font-family:var(--mono); font-size:12px; font-weight:700;
  min-width:36px; text-align:right;
}
.valid-label { color:var(--text2); flex:1; }
.valid-bar-wrap { width:100px; height:4px; background:var(--border); border-radius:2px; }
.valid-bar { height:100%; border-radius:2px; }


.merge-backdrop {
  display:none; position:fixed; inset:0; background:rgba(0,0,0,.7);
  z-index:300; align-items:center; justify-content:center;
}
.merge-backdrop.on { display:flex; }
.merge-modal {
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:600px; max-width:95vw; max-height:82vh;
  display:flex; flex-direction:column; box-shadow:var(--shadow-lg);
}
.merge-header { padding:14px 18px; border-bottom:1px solid var(--border); flex-shrink:0; }
.merge-title { font-size:14px; font-weight:600; }
.merge-body { flex:1; overflow-y:auto; padding:14px 18px; display:flex; gap:16px; }
.merge-left { flex:1; min-width:0; }
.merge-right { flex:1; min-width:0; }
.merge-section-label { font-family:var(--mono); font-size:10px; text-transform:uppercase;
  letter-spacing:.08em; color:var(--text3); margin-bottom:8px; }
.merge-file-row {
  display:flex; align-items:center; gap:8px; padding:7px 10px;
  border:1px solid var(--border); border-radius:5px; margin-bottom:6px;
  cursor:pointer; transition: border-color .15s, background .15s;
}
.merge-file-row:hover { border-color:var(--accent2); }
.merge-file-row.selected { border-color:var(--accent); background:var(--accent-tint); }
.merge-file-name { flex:1; font-family:var(--mono); font-size:11px; color:var(--text2); }
.merge-file-cnt { font-family:var(--mono); font-size:10px; color:var(--text3); }
.merge-arrow { text-align:center; padding:16px 0; color:var(--text3); font-size:20px; }
.merge-output-input { width:100%; margin-bottom:8px; }
.merge-result {
  font-family:var(--mono); font-size:11px; padding:8px 10px;
  background:var(--surface2); border-radius:5px; display:none;
}
.merge-result.on { display:block; }
.merge-footer { padding:10px 18px; border-top:1px solid var(--border);
  display:flex; gap:8px; align-items:center; flex-shrink:0; }

/* ── CONTEXT PREVIEW (review hover) ──────────────────────────────────────── */
.cat-preview {
  position:fixed; z-index:500; background:var(--surface2);
  border:1px solid var(--border2); border-radius:6px;
  padding:10px 14px; max-width:320px; box-shadow:0 8px 24px rgba(0,0,0,.5);
  display:none; pointer-events:none;
}
.cat-preview.on { display:block; }
.cat-preview-title { font-size:12px; font-weight:600; color:var(--text); margin-bottom:4px; }
.cat-preview-meta { font-family:var(--mono); font-size:10px; color:var(--text3); margin-bottom:6px; }
.cat-preview-items { font-family:var(--mono); font-size:11px; color:var(--text2); line-height:1.6; }
.cat-preview-item { white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }

/* ── QUALITY BADGE ────────────────────────────────────────────────────────── */
.quality-badge {
  display:inline-block; font-family:var(--mono); font-size:9px;
  padding:1px 5px; border-radius:10px; font-weight:600;
}
.qb-high  { background:var(--accent-tint);  color:var(--accent); }
.qb-mid   { background:rgba(251,191,36,.12); color:var(--warn); }
.qb-low   { background:var(--danger-tint);  color:var(--danger); }

/* ── FULLTEXT SEARCH ──────────────────────────────────────────────────────── */
.ft-highlight { background:rgba(251,191,36,.3); border-radius:2px; }
.ft-snippet { font-size:11px; color:var(--text3); font-family:var(--mono);
  display:block; margin-top:2px; }
.dt-search-mode {
  font-family:var(--mono); font-size:10px; color:var(--text3);
  padding:2px 6px; border-radius:3px; background:var(--surface2);
  cursor:pointer; border:1px solid var(--border); white-space:nowrap;
}
.dt-search-mode:hover { border-color:var(--accent2); color:var(--accent2); }
.dt-search-mode.ft { border-color:var(--accent2); color:var(--accent2); background:var(--accent2-tint); }

.colpick-backdrop {
  display:none; position:fixed; inset:0; background:rgba(0,0,0,.7);
  z-index:300; align-items:center; justify-content:center;
}
.colpick-backdrop.on { display:flex; }
.colpick-modal {
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:520px; max-width:95vw; max-height:80vh;
  display:flex; flex-direction:column; box-shadow:var(--shadow-lg);
}
.colpick-header { padding:14px 18px; border-bottom:1px solid var(--border); flex-shrink:0; }
.colpick-title { font-size:14px; font-weight:600; margin-bottom:2px; }
.colpick-sub { font-family:var(--mono); font-size:11px; color:var(--text3); }
.colpick-toolbar {
  padding:8px 14px; border-bottom:1px solid var(--border);
  display:flex; gap:8px; align-items:center; flex-shrink:0; background:var(--bg);
}
.thresh-wrap { font-family:var(--mono); font-size:11px; color:var(--text2); display:flex; align-items:center; gap:6px; margin-left:auto; }
.thresh-wrap input { width:52px; margin-bottom:0; padding:4px 6px; }
.colpick-body { flex:1; overflow-y:auto; }
.colpick-body::-webkit-scrollbar { width:3px; }
.colpick-body::-webkit-scrollbar-thumb { background:var(--border2); }
.colpick-section-hdr {
  padding:5px 14px; font-family:var(--mono); font-size:10px;
  text-transform:uppercase; letter-spacing:.08em; color:var(--text3);
  background:var(--surface2); border-bottom:1px solid var(--border);
  position:sticky; top:0; z-index:1;
}
.colpick-row {
  display:flex; align-items:center; gap:10px; padding:7px 14px;
  border-bottom:1px solid var(--border); cursor:pointer; transition:background .1s;
}
.colpick-row:hover { background:var(--surface2); }
.colpick-row.on .cp-chk { background:var(--accent); border-color:var(--accent); color:var(--bg); }
.colpick-row.on .cp-name { color:var(--accent); }
.cp-chk { width:14px; height:14px; border-radius:3px; flex-shrink:0; border:1.5px solid var(--border2); display:flex; align-items:center; justify-content:center; font-size:9px; }
.cp-name { flex:1; font-family:var(--mono); font-size:12px; color:var(--text2); }
.cp-bar-wrap { width:70px; height:4px; background:var(--border); border-radius:2px; flex-shrink:0; }
.cp-bar { height:100%; border-radius:2px; background:var(--accent2); transition:width .2s; }
.cp-pct { font-family:var(--mono); font-size:10px; color:var(--text3); min-width:34px; text-align:right; }
.colpick-footer { padding:10px 14px; border-top:1px solid var(--border); display:flex; gap:8px; align-items:center; flex-shrink:0; }
.colpick-count { font-family:var(--mono); font-size:11px; color:var(--text2); flex:1; }

.project-bar {
  display: flex; align-items: center; gap: 0;
  background: var(--bg); border-bottom: 1px solid var(--border);
  padding: 0 4px; flex-shrink: 0; overflow-x: auto; scrollbar-width: none;
}
.project-bar::-webkit-scrollbar { display: none; }
.proj-tab {
  display: flex; align-items: center; gap: 6px;
  padding: 6px 14px; font-family: var(--mono); font-size: 11px;
  color: var(--text3); cursor: pointer; border-bottom: 2px solid transparent;
  white-space: nowrap; transition: color .15s, background .15s, border-color .15s;
  border-right: 1px solid var(--border); min-width: 80px;
}
.proj-tab:hover { color: var(--text2); background: var(--surface2); }
.proj-tab.active { color: var(--text); border-bottom-color: var(--accent2); background: var(--surface); }
.proj-name { flex: 1; outline: none; cursor: pointer; }
.proj-close { font-size: 13px; color: var(--text3); padding: 0 2px; }
.proj-close:hover { color: var(--danger); }
.proj-add { padding: 6px 12px; font-size: 16px; color: var(--text3); background: transparent; border: none; cursor: pointer; }
.proj-add:hover { color: var(--accent); }

/* ── SAVED REVIEW BANNER ──────────────────────────────────────────────────── */
.saved-review-banner {
  display: none; align-items: center; gap: 8px;
  padding: 7px 16px; background: var(--accent2-tint);
  border-bottom: 1px solid var(--border2);
  font-family: var(--mono); font-size: 11px; color: var(--accent2); flex-shrink: 0;
}
.saved-review-banner button {
  font-family: var(--mono); font-size: 10px; padding: 3px 9px;
  background: var(--accent2); color: var(--bg); border: none;
  border-radius: 3px; cursor: pointer;
}
.saved-review-banner .srv-dismiss { background: transparent; color: var(--text3); border: 1px solid var(--border); }

/* ── RENAME FILE ──────────────────────────────────────────────────────────── */
.rename-wrap { display: flex; gap: 6px; align-items: center; }
.rename-input {
  flex: 1; background: var(--bg); border: 1px solid var(--border2);
  border-radius: var(--radius-sm); color: var(--text);
  font-family: var(--mono); font-size: 12px; padding: 5px 8px; outline: none;
}
.rename-input:focus { border-color: var(--accent); }


/* ── LAYOUT ──────────────────────────────────────────────────────────────── */
.app { display: grid; grid-template-rows: 48px auto auto 1fr; height: 100vh; }
.body { display: grid; grid-template-columns: var(--sidebar-w,320px) 1fr; overflow: hidden; }

/* ── HEADER ──────────────────────────────────────────────────────────────── */
.header {
  display: flex; align-items: center; gap: 12px;
  padding: 0 20px;
  background: var(--surface);
  border-bottom: 1px solid var(--border);
  box-shadow: var(--shadow);
  grid-column: 1 / -1;
}
.header-logo { font-family: var(--mono); font-size: 15px; font-weight: 600; color: var(--text); }
.header-logo em { color: var(--accent); font-style: normal; }
.header-badge {
  font-family: var(--mono); font-size: 10px; color: var(--text3);
  border: 1px solid var(--border2); border-radius: 20px; padding: 1px 7px;
}
.header-right { margin-left: auto; display: flex; align-items: center; gap: 10px; }
.status-dot { width: 7px; height: 7px; border-radius: 50%; background: var(--text3); }
.status-dot.running { background: var(--accent); box-shadow: 0 0 6px var(--accent); animation: blink 1.5s infinite; }
.status-dot.done    { background: var(--accent2); }
.status-dot.error   { background: var(--danger); }
@keyframes blink { 0%,100%{opacity:1} 50%{opacity:.3} }
.status-label { font-family: var(--mono); font-size: 11px; color: var(--text2); }
.rl-chip {
  display: none; font-family: var(--mono); font-size: 10px;
  padding: 2px 8px; border-radius: 20px;
  background: rgba(251,191,36,.1); border: 1px solid rgba(251,191,36,.3); color: var(--warn);
}
.rl-chip.on { display: block; }

/* ── LEFT PANEL ──────────────────────────────────────────────────────────── */
.sidebar {
  background: var(--surface);
  border-right: 1px solid var(--border);
  box-shadow: var(--shadow);
  display: flex; flex-direction: column;
  overflow: hidden;
}
.sidebar-scroll {
  flex: 1; overflow-y: auto; padding: 16px;
  display: flex; flex-direction: column; gap: 20px;
}
.sidebar-scroll::-webkit-scrollbar { width: 4px; }
.sidebar-scroll::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 2px; }

/* Sidebar footer — buttons always visible */
.sidebar-footer {
  padding: 12px 16px;
  border-top: 1px solid var(--border);
  display: flex; flex-direction: column; gap: 8px;
  flex-shrink: 0;
}

/* ── SECTION LABEL ───────────────────────────────────────────────────────── */
.section-label {
  font-family: var(--mono); font-size: 10px; font-weight: 600;
  letter-spacing: .08em; text-transform: uppercase;
  color: var(--text3); margin-bottom: 8px;
}

/* ── FORM ELEMENTS ───────────────────────────────────────────────────────── */
.field-label {
  font-family: var(--mono); font-size: 11px; color: var(--text2);
  display: block; margin-bottom: 4px;
}
input[type=text], input[type=number], select {
  width: 100%;
  background: var(--bg);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  color: var(--text);
  font-family: var(--mono); font-size: 12px;
  padding: 7px 10px;
  outline: none;  /* nahrazeno :focus-visible globálně */
  transition: border-color .15s, box-shadow .15s;
  margin-bottom: 10px;
  -webkit-appearance: none;
}
input:focus-visible, select:focus-visible {
  border-color: var(--accent2);
  box-shadow: 0 0 0 3px rgba(96,165,250,.18);
}
html.light input:focus-visible, html.light select:focus-visible {
  box-shadow: 0 0 0 3px rgba(26,78,207,.15);
}
input::placeholder { color: var(--text3); }
.row-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }

/* ── FIELD TOGGLES ───────────────────────────────────────────────────────── */
.field-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 5px; }
.field-toggle {
  display: flex; align-items: center; gap: 8px;
  padding: 8px 10px;
  background: var(--bg);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  cursor: pointer;
  transition: border-color .15s, background .15s;
  user-select: none;
}
.field-toggle:hover { border-color: var(--border2); }
.field-toggle.on { border-color: var(--accent); background: var(--accent-tint); }
.field-toggle.on .ft-name { color: var(--accent); }
.ft-box {
  width: 14px; height: 14px; border-radius: 3px; flex-shrink: 0;
  border: 1.5px solid var(--border2);
  display: flex; align-items: center; justify-content: center;
  font-size: 9px; transition: background .15s, border-color .15s;
}
.field-toggle.on .ft-box { background: var(--accent); border-color: var(--accent); color: var(--bg); }
.ft-text { min-width: 0; }
.ft-name { font-family: var(--mono); font-size: 11px; font-weight: 600; display: block; color: var(--text2); transition: color .15s; }
.ft-desc { font-family: var(--mono); font-size: 10px; color: var(--text3); display: block; }
.slow-warn { font-family: var(--mono); font-size: 10px; color: var(--warn); margin-top: 4px; display: none; }
.slow-warn.on { display: block; }

/* ── QUICK EXAMPLES ──────────────────────────────────────────────────────── */
.ex-list { display: flex; flex-direction: column; gap: 4px; }
.ex-btn {
  display: flex; align-items: center; gap: 8px;
  padding: 8px 10px;
  background: var(--bg);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  cursor: pointer; text-align: left;
  transition: border-color .15s;
  width: 100%;
}
.ex-btn:hover { border-color: var(--border2); }
.ex-icon { font-size: 16px; flex-shrink: 0; width: 22px; text-align: center; }
.ex-info { min-width: 0; }
.ex-name { font-family: var(--sans); font-size: 12px; font-weight: 600; color: var(--text); display: block; }
.ex-url  { font-family: var(--mono); font-size: 10px; color: var(--text3); display: block; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.ex-shuffle { font-family: var(--mono); font-size: 10px; color: var(--accent2); cursor: pointer; float: right; }
.ex-shuffle:hover { text-decoration: underline; }

/* ── RESUME BOX ──────────────────────────────────────────────────────────── */
.resume-box {
  display: none;
  background: var(--accent2-tint);
  border: 1px solid var(--border2);
  border-radius: var(--radius);
  padding: 12px;
}
.resume-box.on { display: block; }
.resume-title { font-weight: 600; color: var(--accent2); font-size: 13px; margin-bottom: 4px; }
.resume-info  { font-family: var(--mono); font-size: 11px; color: var(--text2); margin-bottom: 10px; }
.resume-btns  { display: flex; gap: 6px; }

/* ── BUTTONS ──────────────────────────────────────────────────────────────── */
.btn {
  display: flex; align-items: center; justify-content: center; gap: 6px;
  padding: 9px 14px; border-radius: var(--radius-sm);
  font-family: var(--sans); font-size: 13px; font-weight: 600;
  border: none; cursor: pointer;
  transition: background .15s, color .15s, border-color .15s, box-shadow .15s;
  width: 100%;
}
.btn:active { transform: translateY(1px); }
.btn:disabled { opacity: .4; cursor: not-allowed; transform: none; }
.btn-primary  { background: var(--accent); color: #0f1117; }
.btn-primary:hover { box-shadow: 0 0 0 3px var(--accent-tint), 0 2px 8px rgba(0,0,0,.2); }
html.light .btn-primary { color: #fff; }
.btn-danger   { background: transparent; color: var(--danger); border: 1px solid var(--danger); }
.btn-danger:hover { background: var(--danger-tint); }
.btn-ghost    { background: transparent; color: var(--text2); border: 1px solid var(--border); }
.btn-ghost:hover { border-color: var(--border2); color: var(--text); background: var(--surface2); }
.btn-sm {
  padding: 5px 10px; font-size: 12px; font-family: var(--mono);
  background: var(--surface2); border: 1px solid var(--border);
  border-radius: var(--radius-sm); cursor: pointer; color: var(--text2);
  transition: border-color .15s, color .15s, background .15s;
}
.btn-sm:hover { border-color: var(--border2); color: var(--text); background: var(--surface); }
.hidden { display: none !important; }

/* ── RIGHT PANEL ──────────────────────────────────────────────────────────── */
.main {
  display: flex; flex-direction: column; overflow: hidden;
  background: var(--bg);
}

/* ── PROGRESS ────────────────────────────────────────────────────────────── */
.progress-section {
  background: var(--surface);
  border-bottom: 1px solid var(--border);
  padding: 8px 16px;
  flex-shrink: 0;
}
.progress-bars { display: flex; flex-direction: column; gap: 5px; margin-bottom: 7px; }
.progress-row { display: flex; align-items: center; gap: 10px; }
.progress-name { font-family: var(--mono); font-size: 10px; color: var(--text3); min-width: 100px; }
.progress-track { flex: 1; height: 3px; background: var(--border); border-radius: 2px; overflow: hidden; }
.progress-fill { height: 100%; border-radius: 2px; width: 0%; transition: width .4s; }
.progress-fill.p1 { background: var(--accent2); }
.progress-fill.p2 { background: var(--accent); }
.progress-pct { font-family: var(--mono); font-size: 10px; color: var(--text3); min-width: 30px; text-align: right; }

.current-task {
  display: flex; align-items: center; gap: 8px;
  padding: 5px 8px;
  background: var(--surface2);
  border-radius: var(--radius-sm);
  min-height: 28px;
}
.task-phase {
  font-family: var(--mono); font-size: 9px; font-weight: 600;
  padding: 1px 5px; border-radius: 3px;
  flex-shrink: 0;
}
.task-phase.p1 { background: var(--accent2-tint); color: var(--accent2); }
.task-phase.p2 { background: var(--accent-tint);  color: var(--accent); }
.task-name { flex: 1; font-family: var(--mono); font-size: 11px; color: var(--text); overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.task-eta  { font-family: var(--mono); font-size: 10px; color: var(--text3); flex-shrink: 0; }

/* ── STATS BAR ───────────────────────────────────────────────────────────── */
.stats-bar {
  display: flex; border-bottom: 1px solid var(--border);
  background: var(--surface); flex-shrink: 0;
  /* Schovat v idle stavu — zobrazí se JS při prvním PROGRESS eventu */
  max-height: 0; overflow: hidden; transition: max-height .25s ease;
}
.stats-bar.visible { max-height: 60px; }
.stat { padding: 6px 14px; border-right: 1px solid var(--border); }
.stat-val { font-family: var(--mono); font-size: 16px; font-weight: 600; color: var(--text); }
.stat-val.c-accent  { color: var(--accent); }
.stat-val.c-accent2 { color: var(--accent2); }
.stat-val.c-warn    { color: var(--warn); }
.stat-lbl { font-family: var(--mono); font-size: 10px; color: var(--text3); text-transform: uppercase; letter-spacing: .05em; }

/* ── TABS ────────────────────────────────────────────────────────────────── */
.tabs { display: flex; background: var(--surface); border-bottom: 1px solid var(--border); flex-shrink: 0; }
.tab-btn {
  padding: 8px 16px; font-family: var(--mono); font-size: 11px; text-transform: uppercase; letter-spacing: .06em;
  color: var(--text3); background: transparent; border: none; border-bottom: 2px solid transparent;
  cursor: pointer; transition: color .15s, border-color .15s; margin-bottom: -1px;
}
.tab-btn:hover { color: var(--text2); }
.tab-btn.on { color: var(--accent); border-bottom-color: var(--accent); }
.tab-actions { margin-left: auto; display: flex; align-items: center; gap: 6px; padding: 0 12px; }

/* ── CONSOLE ─────────────────────────────────────────────────────────────── */
.console-header {
  display: flex; align-items: center; gap: 8px;
  padding: 6px 16px;
  background: var(--surface); border-bottom: 1px solid var(--border);
  cursor: pointer; flex-shrink: 0;
}
.console-header:hover { background: var(--surface2); }
.console-caret { font-size: 10px; color: var(--text3); margin-left: auto; transition: transform .2s; }
.console-caret.open { transform: rotate(180deg); }
.console-label { font-family: var(--mono); font-size: 10px; text-transform: uppercase; letter-spacing: .08em; color: var(--text3); }
.console-body { overflow: hidden; max-height: 0; transition: max-height .25s ease; flex-shrink: 0; }
.console-body.open { max-height: 140px; }
.console-body.open.resized { max-height: var(--console-h, 140px); transition: none; }
.console-inner {
  height: 140px; overflow-y: auto;
  padding: 6px 16px;
  font-family: var(--mono); font-size: 11px; line-height: 1.65;
  background: var(--bg);
}
.console-inner::-webkit-scrollbar { width: 3px; }
.console-inner::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 2px; }
.log-row { display: flex; gap: 8px; }
.log-t { color: var(--text3); font-size: 10px; min-width: 46px; flex-shrink: 0; padding-top: 1px; }
.log-m { flex: 1; word-break: break-all; color: var(--text2); }
.log-m.ok   { color: var(--accent);  }
.log-m.warn { color: var(--warn);    }
.log-m.err  { color: var(--danger);  }
.log-m.dim  { color: var(--text3);   }
.log-m.hl   { color: var(--accent); font-weight: 600; }
.log-m.info { color: var(--accent2); }

/* ── TAB PANELS ──────────────────────────────────────────────────────────── */
.tab-panel { flex: 1; overflow: hidden; display: none; flex-direction: column; }
.tab-panel.on { display: flex; }

/* Overview */
.overview-panel { flex: 1; overflow-y: auto; padding: 20px; }
.overview-panel::-webkit-scrollbar { width: 4px; }
.overview-panel::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 2px; }
.overview-empty {
  display: flex; flex-direction: column; align-items: center; justify-content: center;
  height: 100%; gap: 10px; color: var(--text3); font-family: var(--mono); font-size: 12px; text-align: center;
}
.overview-icon { font-size: 36px; opacity: .25; }
.summary-cards { display: flex; flex-wrap: wrap; gap: 12px; }
.summary-card {
  background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius);
  padding: 14px 16px; min-width: 180px;
}
.sc-val { font-family: var(--mono); font-size: 28px; font-weight: 600; color: var(--accent); }
.sc-lbl { font-family: var(--mono); font-size: 11px; color: var(--text3); margin-top: 2px; }
.sc-sub { font-family: var(--mono); font-size: 10px; color: var(--text3); margin-top: 8px; line-height: 1.6; }

/* Data table */
.dt-toolbar {
  display: flex; align-items: center; gap: 8px;
  padding: 8px 12px; background: var(--surface); border-bottom: 1px solid var(--border); flex-shrink: 0;
}
.dt-search {
  flex: 1; background: var(--bg); border: 1px solid var(--border);
  border-radius: var(--radius-sm); color: var(--text);
  font-family: var(--mono); font-size: 12px; padding: 6px 10px; outline: none;
}
.dt-search:focus { border-color: var(--accent2); }
.dt-count { font-family: var(--mono); font-size: 11px; color: var(--text3); white-space: nowrap; }
.dt-wrap { flex: 1; overflow-y: auto; }
.dt-wrap::-webkit-scrollbar { width: 4px; }
.dt-wrap::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 2px; }
table { width: 100%; border-collapse: collapse; }
thead th {
  background: var(--surface); position: sticky; top: 0; z-index: 1;
  font-family: var(--mono); font-size: 10px; text-transform: uppercase; letter-spacing: .06em;
  color: var(--text3); padding: 8px 12px; text-align: left;
  border-bottom: 1px solid var(--border); cursor: pointer; user-select: none; white-space: nowrap;
}
thead th:hover { color: var(--text2); }
thead th.sorted { color: var(--accent); }
tbody tr { border-bottom: 1px solid var(--border); transition: background .1s; }
tbody tr:hover { background: var(--surface); }
tbody td { padding: 7px 12px; font-family: var(--mono); font-size: 11px; vertical-align: top; }
.td-num { color: var(--text3); width: 36px; text-align: right; }
.td-title a { color: var(--accent2); text-decoration: none; font-weight: 600; font-family: var(--sans); font-size: 12px; }
.td-title a:hover { text-decoration: underline; }
.td-text { color: var(--text2); max-width: 260px; overflow: hidden; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; }
.td-cats { color: var(--text3); font-size: 10px; max-width: 160px; }
.td-ib-count { color: var(--text3); text-align: center; }
.ib-link { color: var(--accent); cursor: pointer; font-size: 10px; }
.ib-link:hover { text-decoration: underline; }
.no-rows { text-align: center; padding: 40px; color: var(--text3); font-family: var(--mono); font-size: 12px; }

/* Infobox popup */
.ib-popup {
  display: none; position: fixed; z-index: 100;
  background: var(--surface); border: 1px solid var(--border2);
  border-radius: var(--radius); padding: 12px;
  max-width: 300px; max-height: 360px; overflow-y: auto;
  box-shadow: 0 8px 32px rgba(0,0,0,.5);
  font-family: var(--mono); font-size: 11px;
}
.ib-popup::-webkit-scrollbar { width: 3px; }
.ib-popup::-webkit-scrollbar-thumb { background: var(--border2); }
.ib-row { display: flex; gap: 8px; padding: 4px 0; border-bottom: 1px solid var(--border); }
.ib-k { color: var(--text3); min-width: 80px; flex-shrink: 0; }
.ib-v { color: var(--text); word-break: break-word; }

/* Viz */
.viz-scroll { flex: 1; overflow-y: auto; padding: 16px; display: flex; flex-wrap: wrap; gap: 16px; align-content: flex-start; }
.viz-scroll::-webkit-scrollbar { width: 4px; }
.viz-scroll::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 2px; }
.viz-card { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 14px; }
.viz-card-title { font-family: var(--mono); font-size: 10px; text-transform: uppercase; letter-spacing: .08em; color: var(--text3); margin-bottom: 12px; }
#mapWrap { height: 300px; border-radius: var(--radius-sm); overflow: hidden; }

/* Review modal — multi-column */
.modal-backdrop {
  display:none; position:fixed; inset:0; background:rgba(0,0,0,.7);
  z-index:200; align-items:center; justify-content:center;
}
.modal-backdrop.on { display:flex; }
.modal {
  background:var(--surface); border:1px solid var(--border2);
  border-radius:var(--radius); width:92vw; max-width:1200px; height:82vh;
  display:flex; flex-direction:column; box-shadow:var(--shadow-lg);
}
.modal-header {
  padding:14px 20px; border-bottom:1px solid var(--border);
  display:flex; align-items:center; gap:12px; flex-shrink:0;
}
.modal-title { font-size:15px; font-weight:600; }
.modal-sub { font-family:var(--mono); font-size:11px; color:var(--text3); }
.modal-footer {
  padding:10px 20px; border-top:1px solid var(--border);
  display:flex; align-items:center; gap:8px; flex-shrink:0;
}

/* Columns container */
.review-columns {
  flex:1; display:flex; overflow:hidden; min-height:0;
}
.review-col {
  display:flex; flex-direction:column;
  border-right:1px solid var(--border);
  min-width:280px; flex:1;
  overflow:hidden;
}
.review-col:last-child { border-right:none; }

/* Column header */
.col-header {
  padding:10px 14px; border-bottom:1px solid var(--border);
  background:var(--surface2); flex-shrink:0;
  display:flex; align-items:center; gap:8px;
}
.col-name-input {
  flex:1; background:transparent; border:none; outline:none;
  font-family:var(--mono); font-size:12px; font-weight:600;
  color:var(--text); min-width:0;
}
.col-name-input::placeholder { color:var(--text3); }
.col-stats { font-family:var(--mono); font-size:10px; color:var(--text3); flex-shrink:0; }
.col-delete {
  font-size:14px; color:var(--text3); cursor:pointer; padding:0 2px;
  flex-shrink:0; line-height:1;
}
.col-delete:hover { color:var(--danger); }

/* Column toolbar (search + select all) */
.col-toolbar {
  padding:6px 10px; border-bottom:1px solid var(--border);
  display:flex; gap:6px; background:var(--bg); flex-shrink:0;
}
.col-search {
  flex:1; background:var(--surface); border:1px solid var(--border);
  border-radius:3px; color:var(--text); font-family:var(--mono); font-size:11px;
  padding:4px 8px; outline:none;
}
.col-search:focus { border-color:var(--accent2); }
.col-toolbar-btn {
  font-family:var(--mono); font-size:10px; padding:3px 7px;
  background:transparent; border:1px solid var(--border);
  border-radius:3px; cursor:pointer; color:var(--text3); white-space:nowrap;
}
.col-toolbar-btn:hover { color:var(--text); border-color:var(--border2); }

/* Category items */
.col-body { flex:1; overflow-y:auto; scrollbar-width:thin; scrollbar-color:var(--border2) transparent; }
.col-body::-webkit-scrollbar { width:3px; }
.col-body::-webkit-scrollbar-thumb { background:var(--border2); border-radius:2px; }
/* Prevent scrollbar from overlapping arrow buttons */
.cat-arrows { flex-shrink:0; margin-right:3px; }

.cat-row {
  display:flex; align-items:center; gap:0;
  border-bottom:1px solid var(--border);
  transition:background .1s;
}
.cat-row:hover { background:var(--surface2); }
.cat-row.focused { background:rgba(96,165,250,.10); }
.cat-row.focused .cat-label { color:var(--text); }
.cat-row.focused .cat-badge { background:rgba(96,165,250,.15); color:var(--accent2); }

/* Checkbox — only in col 0 */
.cat-check-cell {
  width:36px; display:flex; align-items:center; justify-content:center;
  flex-shrink:0; cursor:pointer; padding:10px 0; align-self:stretch;
}
.cat-check-cell:hover { background: var(--accent-tint); }
.cat-cb {
  width:14px; height:14px; border-radius:3px;
  border:1.5px solid var(--border2);
  display:flex; align-items:center; justify-content:center;
  font-size:9px; transition: background .15s, border-color .15s;
}
.cat-row.checked .cat-cb { background:var(--accent); border-color:var(--accent); color:var(--bg); }
.cat-row.unchecked .cat-cb { opacity:.4; }

/* Name + count */
.cat-label {
  flex:1; padding:10px 8px; font-size:12px; color:var(--text2);
  overflow:hidden; text-overflow:ellipsis; white-space:nowrap; min-width:0;
  cursor:default;
}
.cat-row.unchecked .cat-label { color:var(--text3); text-decoration:line-through; }
.cat-row.unchecked .cat-label span { color:inherit; }
.cat-badge {
  font-family:var(--mono); font-size:10px; color:var(--text3);
  padding:1px 6px; border-radius:10px; background:var(--surface2);
  flex-shrink:0; margin-right:4px;
}

/* Arrow buttons */
.cat-arrows { display:flex; flex-shrink:0; }
.cat-arr {
  width:28px; height:100%; display:flex; align-items:center; justify-content:center;
  font-size:12px; color:var(--text3); cursor:pointer; padding:10px 4px;
  transition: color .15s, background .15s; align-self:stretch;
}
.cat-arr:hover { color:var(--accent2); background:var(--accent2-tint); }
.cat-arr.disabled { opacity:.2; cursor:default; pointer-events:none; }

/* Add column button */
.add-col-btn {
  display:flex; flex-direction:column; align-items:center; justify-content:center;
  gap:8px; padding:20px 16px; cursor:pointer; color:var(--text3);
  border-left:1px dashed var(--border); min-width:60px; flex-shrink:0;
  transition: color .2s, background .2s, border-color .2s;
}
.add-col-btn:hover { color:var(--accent); background:var(--accent-tint); border-color:var(--accent); }
.add-col-icon { font-size:22px; }
.add-col-label { font-family:var(--mono); font-size:10px; writing-mode:vertical-rl; letter-spacing:.1em; }

/* Summary footer */
.review-summary {
  font-family:var(--mono); font-size:11px; color:var(--text2);
  display:flex; gap:16px; flex:1;
}
.rs-item { display:flex; align-items:center; gap:6px; }
.rs-dot { width:8px; height:8px; border-radius:50%; flex-shrink:0; }

/* Indeterminate progress for phase 1 (unknown total) */
.progress-track { overflow: hidden; }
.progress-fill.indeterminate {
  width: 35% !important;
  animation: slide-ind 1.5s ease-in-out infinite;
}
@keyframes slide-ind {
  0%   { transform: translateX(-200%); }
  100% { transform: translateX(400%); }
}

.export-box {
  display: none;
  background: var(--accent-tint);
  border: 1px solid var(--border2);
  border-radius: var(--radius); padding: 12px;
}
.export-box.on { display: block; }
.export-title { font-weight: 600; color: var(--accent); font-size: 13px; margin-bottom: 8px; }
.export-status { font-family: var(--mono); font-size: 11px; margin-top: 6px; min-height: 16px; }
/* ── RIPPLE ───────────────────────────────────────────────────────────────── */
.btn, .btn-sm { position: relative; overflow: hidden; }
.ripple-circle {
  position: absolute; border-radius: 50%; pointer-events: none;
  background: currentColor; opacity: .12; transform: scale(0);
  animation: ripple-anim .55s linear; z-index: 0;
}
@keyframes ripple-anim { to { transform: scale(4); opacity: 0; } }

/* ── COL RESIZER (review columns) ────────────────────────────────────────── */
.col-resizer {
  width: 5px; flex-shrink: 0; background: transparent;
  cursor: col-resize; transition: background .15s; align-self: stretch;
  border-right: 1px solid var(--border);
}
.col-resizer:hover, .col-resizer.active { background: var(--accent2); border-color: var(--accent2); }

/* ── COLUMN HIDER ─────────────────────────────────────────────────────────── */
.col-hider-wrap { position: relative; flex-shrink: 0; }
.col-hider-panel {
  display: none; position: absolute; top: calc(100% + 4px); right: 0; z-index: 200;
  background: var(--surface); border: 1px solid var(--border2);
  border-radius: var(--radius); padding: 6px; min-width: 150px;
  box-shadow: 0 8px 24px rgba(0,0,0,.4);
}
.col-hider-panel.on { display: block; }
.col-hider-row {
  display: flex; align-items: center; gap: 8px; padding: 5px 8px;
  cursor: pointer; border-radius: 3px;
  font-family: var(--mono); font-size: 11px; color: var(--text2); user-select: none;
}
.col-hider-row:hover { background: var(--surface2); }
.col-hider-row input { accent-color: var(--accent2); cursor: pointer; }

/* ── TABLE KEYBOARD FOCUS ────────────────────────────────────────────────── */
tbody tr.tr-focused { background: var(--accent2-tint) !important; outline: 1px solid var(--border2); }

/* ── HISTORY COLLAPSE ────────────────────────────────────────────────────── */
.history-more-btn {
  display: block; text-align: center; padding: 5px;
  font-family: var(--mono); font-size: 10px; color: var(--accent2);
  cursor: pointer; border-radius: var(--radius-sm); margin-top: 2px;
}
.history-more-btn:hover { background: var(--surface2); }

/* ── REVIEW UNDO BAR ─────────────────────────────────────────────────────── */
.review-undo-bar {
  display: none; align-items: center; gap: 6px;
  padding: 5px 20px; border-bottom: 1px solid var(--border);
  background: var(--surface2); flex-shrink: 0;
}
.review-undo-bar.on { display: flex; }
.review-undo-label { font-family: var(--mono); font-size: 10px; color: var(--text3); margin-left: auto; }

/* ── SHAKE ANIMATION ──────────────────────────────────────────────────────── */
@keyframes shake {
  0%,100% { transform: translateX(0); }
  20%     { transform: translateX(-7px); }
  40%     { transform: translateX(7px); }
  60%     { transform: translateX(-4px); }
  80%     { transform: translateX(4px); }
}
.shake { animation: shake 0.45s ease; }

/* ── SIDEBAR COLLAPSE ─────────────────────────────────────────────────────── */
.body { position: relative; transition: grid-template-columns .3s ease; }
.body.sidebar-collapsed { grid-template-columns: 0 1fr; }
.body.sidebar-collapsed .sidebar { pointer-events: none; overflow: hidden; }
.sidebar-toggle-btn {
  position: absolute; left: var(--sidebar-w,320px); top: 50%; transform: translateY(-50%);
  width: 12px; height: 44px; z-index: 20;
  background: var(--surface2); border: 1px solid var(--border);
  border-left: none; border-radius: 0 5px 5px 0;
  cursor: col-resize; display: flex; align-items: center; justify-content: center;
  font-size: 9px; color: var(--text3);
  transition: left .3s ease, color .15s, background .15s;
  line-height: 1; padding: 0;
}
.sidebar-toggle-btn:hover { color: var(--text); background: var(--surface); }
.body.sidebar-collapsed .sidebar-toggle-btn {
  left: 0 !important; border-left: 1px solid var(--border);
  border-radius: 0 5px 5px 0; cursor: pointer;
}

/* ── VIRTUAL SCROLL SPACER ────────────────────────────────────────────────── */
.vscroll-spacer td { padding: 0 !important; border: none !important; background: none !important; }

/* ── REVIEW DRAG & DROP ───────────────────────────────────────────────────── */
.review-col.drag-over { background: var(--accent2-tint); outline: 2px dashed var(--accent2); outline-offset: -2px; }
.col-header { cursor: grab; user-select: none; }
.col-header:active { cursor: grabbing; }
</style>
</head>
<body>
<div class="app">

<!-- HEADER -->
<header class="header">
  <div class="header-logo">Wiki<em>Scraper</em></div>
  <div class="header-badge">v7</div>
  <div class="header-badge" id="sessionBadge" title="Tvoje session ID" style="color:var(--text3);cursor:default;font-size:9px"></div>
  <a href="/admin/users" class="logout-btn" style="margin-left:auto" title="Správa uživatelů (jen admin)">👥</a>
  <a href="/profile" class="logout-btn" title="Změnit heslo">🔑</a>
  <a href="/logout" class="logout-btn" title="Odhlásit se">Odhlásit</a>
  <!-- STEPPER -->
  <div class="stepper" id="stepper">
    <div class="step" id="step0"><div class="step-dot" id="sd0">1</div>Sbírání</div>
    <div class="step-line" id="sl0"></div>
    <div class="step" id="step1"><div class="step-dot" id="sd1">2</div>Review</div>
    <div class="step-line" id="sl1"></div>
    <div class="step" id="step2"><div class="step-dot" id="sd2">3</div>Stahování</div>
    <div class="step-line" id="sl2"></div>
    <div class="step" id="step3"><div class="step-dot" id="sd3">4</div>Hotovo</div>
  </div>
  <div class="header-right">
    <div class="rl-chip" id="rlChip">⏱ <span id="rlVal">—</span>s delay</div>
    <!-- Global search -->
    <div class="global-search-wrap" id="globalSearchWrap">
      <input class="global-search-input" id="globalSearchInput"
        placeholder="🔍 Hledat ve všech souborech…"
        oninput="globalSearch(this.value)"
        onfocus="document.getElementById('globalResults').classList.add('on')"
        onblur="setTimeout(()=>document.getElementById('globalResults').classList.remove('on'),200)">
      <div class="global-search-results" id="globalResults"></div>
    </div>
    <!-- Notification bell -->
    <div class="notif-bell-wrap">
      <button class="notif-bell" onclick="toggleNotifPanel()" title="Notifikace">🔔<span class="notif-badge" id="notifBadge"></span></button>
      <div class="notif-panel" id="notifPanel">
        <div class="notif-panel-header">
          Notifikace
          <button class="btn-sm" style="margin-left:auto" onclick="clearNotifs()" title="Smazat všechny notifikace">Vymazat</button>
        </div>
        <div class="notif-list" id="notifList"><div class="notif-empty">Zatím žádné notifikace</div></div>
      </div>
    </div>
    <!-- Info button -->
    <button class="theme-btn" onclick="document.getElementById('infoModal').classList.add('on')" title="O aplikaci a nápověda">ℹ</button>
    <button class="theme-btn" id="themeBtn" onclick="toggleTheme()" title="Přepnout světlý/tmavý režim">🌙</button>
    <div class="status-dot" id="stDot"></div>
    <div class="status-label" id="stLbl">Idle</div>
  </div>
</header>

<!-- PROJECT TABS -->
<div class="project-bar" id="projectTabs"></div>

<!-- SAVED REVIEW BANNER -->
<div class="saved-review-banner" id="savedReviewBanner">
  📋 <span id="srvText"></span>
  <button onclick="resumeSavedReview()" title="Otevřít uložený stav review dialogu">↩ Pokračovat v review</button>
  <button class="srv-dismiss" onclick="clearReviewState();this.closest('.saved-review-banner').style.display='none'" title="Zavřít a zahodit uložený stav">×</button>
</div>

<div class="body" id="bodyGrid">
<button class="sidebar-toggle-btn" id="sidebarToggleBtn" onclick="toggleSidebar()" title="Skrýt panel">◀</button>

<!-- ═══ SIDEBAR ═══ -->
<aside class="sidebar">
<div class="sidebar-scroll">

  <!-- URL -->
  <div class="collapsible-section" data-sec="url">
    <div class="section-label sec-header">Cílová URL<span class="sec-caret">▾</span></div>
    <div class="sec-body">
    <input type="text" id="url" placeholder="https://cs.wikipedia.org/wiki/Kategorie:…"
      oninput="onUrlChange()" ondragover="urlDragOver(event)" ondragleave="urlDragLeave(event)" ondrop="urlDrop(event)">
    <div style="display:flex;gap:6px;margin-top:-8px;margin-bottom:12px">
      <button class="btn-sm" id="btnTest" onclick="testUrl()" style="flex:1" title="Otestovat URL — zkontroluje typ stránky a odhadne počet hesel">🔍 Otestovat</button>
    </div>
    <div class="test-result" id="testResult"></div>
    </div>
  </div>

  <!-- BLACKLIST -->
  <div class="collapsible-section" data-sec="blacklist">
    <div class="section-label sec-header">Blacklist kategorií<span class="sec-caret">▾</span></div>
    <div class="sec-body">
    <div class="blacklist-wrap">
      <input type="text" id="blacklist" placeholder="sportovci, ocenění, medailé, …"
        oninput="updateBlacklistCount()" style="margin-bottom:4px">
      <div class="bl-count" id="blCount"></div>
    </div>
    <div style="font-family:var(--mono);font-size:10px;color:var(--text3);margin-bottom:12px">
      Slova oddělená čárkou — kategorie obsahující tato slova se v review automaticky odškrtnou
    </div>
    </div>
  </div>

  <!-- FIELDS -->
  <div class="collapsible-section" data-sec="fields">
    <div class="section-label sec-header">Co stahovat<span class="sec-caret">▾</span></div>
    <div class="sec-body">
    <div class="field-grid" id="fieldGrid"></div>
    <div class="slow-warn" id="slowWarn">⚠ Plný text / sekce prodlouží stahování</div>
    </div>
  </div>

  <!-- LOAD REVIEW -->
  <div class="collapsible-section" data-sec="loadreview">
    <div class="section-label sec-header">Načíst review<span class="sec-caret">▾</span></div>
    <div class="sec-body">
      <div style="font-family:var(--mono);font-size:10px;color:var(--text3);margin-bottom:8px;line-height:1.6">Načte uložený .review.json soubor a otevře review dialog.</div>
      <label style="display:flex;align-items:center;gap:8px;padding:7px 10px;background:var(--bg);border:1px solid var(--border);border-radius:var(--radius-sm);cursor:pointer" id="loadReviewLabel">
        <span>📂</span>
        <span style="font-family:var(--mono);font-size:11px;color:var(--text2)">Vybrat .review.json…</span>
        <input type="file" accept=".json" style="display:none" id="loadReviewInput" onchange="importReviewFile(this)">
      </label>
      <div id="loadReviewMsg" style="font-family:var(--mono);font-size:10px;margin-top:4px;min-height:14px"></div>
    </div>
  </div>

  <!-- RESUME -->
  <div class="resume-box" id="resumeBox">
    <div class="resume-title" id="resumeTitle">⏸ Nedokončený scraping</div>
    <div class="resume-info" id="resumeInfo">—</div>
    <div class="resume-btns">
      <button class="btn btn-ghost"   style="flex:1;font-size:12px" onclick="discardCP()" title="Smazat checkpoint a začít od nuly">🗑 Zahodit</button>
      <button class="btn btn-ghost"   style="flex:1;font-size:12px" onclick="go(false)" title="Zahodit checkpoint a spustit znovu od začátku">↺ Znovu</button>
      <button class="btn btn-primary" style="flex:1;font-size:12px" onclick="go(true)" title="Navázat na přerušené stahování z checkpointu">♻ Pokračovat</button>
    </div>
  </div>

  <!-- HISTORY -->
  <div id="historySection" class="collapsible-section" data-sec="history" style="display:none">
    <div class="section-label sec-header">Naposledy dokončeno<span class="sec-caret">▾</span></div>
    <div class="sec-body">
    <div class="ex-list" id="historyList"></div>
    </div>
  </div>

  <!-- QUICK SELECT -->
  <div class="collapsible-section" data-sec="quickselect">
    <div class="section-label sec-header">
      Rychlý výběr
      <span style="display:flex;align-items:center;gap:6px">
        <span class="ex-shuffle" id="shuffleBtn">↺ zamíchat</span>
        <span class="sec-caret">▾</span>
      </span>
    </div>
    <div class="sec-body">
    <div class="ex-list" id="exList"></div>
    </div>
  </div>

  <!-- PARAMS -->
  <div class="collapsible-section" data-sec="params">
    <div class="section-label sec-header">Parametry<span class="sec-caret">▾</span></div>
    <div class="sec-body">
    <div class="row-2">
      <div data-tip="Hloubka rekurze do podkategorií.
Příklad: hloubka 2 u Chemie projde
Chemie → Prvky → (stop).
Větší hloubka = více hesel, ale i
více nesouvisejících kategorií.">
        <span class="field-label">Hloubka podkat.</span>
        <input type="number" id="depth" value="5" min="1" max="10">
      </div>
      <div data-tip="Maximální počet hesel k procházení.
0 = žádný limit.
Doporučení: nastav 20–50 pro rychlý
test před plným scrapingem.">
        <span class="field-label">Max hesel (0 = ∞)</span>
        <input type="number" id="limit" value="0" min="0">
      </div>
    </div>
    <div class="row-2">
      <div data-tip="Prodleva mezi stahováním hesel.
Wikipedie má rate limit — příliš
rychlé stahování způsobí dočasné
zablokování (429 Too Many Requests).
0.5s je bezpečný kompromis.">
        <span class="field-label">Delay (s)</span>
        <input type="number" id="delay" value="0.5" min="0.1" step="0.1" max="5">
      </div>
      <div data-tip="Výstupní formáty dat.
JSON = strukturovaná data pro programy.
CSV = tabulka, otevře Excel/Sheets.
SQLite = databáze pro SQL dotazy.
README = popis datasetu v Markdownu.">
        <span class="field-label">Formát</span>
        <select id="format" title="Výstupní formát souborů">
          <option value="both">JSON + CSV</option>
          <option value="json">Jen JSON</option>
          <option value="csv">Jen CSV</option>
          <option value="all">JSON + CSV + SQLite + README</option>
        </select>
      </div>
    </div>
    <div data-tip="Název výstupního souboru bez přípony.
Po dokončení vzniknou soubory:
  nazev.json, nazev.csv atd.
Použí jen písmena a podtržítka.">
    <span class="field-label">Název souboru</span>
    <input type="text" id="output" value="wiki_data" oninput="checkCP()">
    </div>
    </div>
  </div>

  <!-- Výkon & API -->
  <div class="collapsible-section" data-sec="apiparams">
    <div class="section-label sec-header">Výkon &amp; API<span class="sec-caret">▾</span></div>
    <div class="sec-body">
    <div class="params-grid" style="margin-bottom:6px">
      <div>
        <span class="field-label">Vlákna</span>
        <input type="number" id="workers" value="1" min="1" max="5" data-tip="Paralelní stahování hesel.
1 vlákno = sekvenční, nejbezpečnější.
2–5 vláken = rychlejší, ale Wikipedia
může dříve aktivovat rate limit.
Doporučení: 2–3 vlákna s delay 0.5s.">
      </div>
    </div>
    <div style="display:flex;flex-direction:column;gap:6px;margin-bottom:8px">
      <label style="display:flex;align-items:center;gap:8px;font-family:var(--mono);font-size:11px;cursor:pointer" data-tip="MediaWiki API vrací čistá data bez
parsování HTML — až 3× rychlejší.
Nevýhoda: infobox se stahuje
zálohově přes HTML (pomalejší).
Doporučení: zapnout pro velké datasety.">
        <input type="checkbox" id="apiMode"> MediaWiki API režim (rychlejší)
      </label>
      <label style="display:flex;align-items:center;gap:8px;font-family:var(--mono);font-size:11px;cursor:pointer" data-tip="Přidá pole _tags ke každému heslu
na základě jeho kategorií.
Příklad: heslo v kat. Města dostane
tag 'město', prvky dostanou 'prvek'.
Potřebné pro záložku Tagy v GUI.">
        <input type="checkbox" id="autoTags"> Auto-tagy z kategorií
      </label>
      <label style="display:flex;align-items:center;gap:8px;font-family:var(--mono);font-size:11px;cursor:pointer" data-tip="Po scrapingu doplní strukturovaná data
z Wikidata: QID, stát, oficiální web,
datum vzniku, souřadnice a další.
Zpomalí stahování o desítky sekund.
Doporučení: jen pro datasety do 500 hesel.">
        <input type="checkbox" id="wikidata"> Wikidata obohacení (pomalé)
      </label>
      <label style="display:flex;align-items:center;gap:8px;font-family:var(--mono);font-size:11px;cursor:pointer" data-tip="Přeskočí hesla, která se od minulého
scrapingu nezměnila — porovnává
timestamp poslední revize na Wikipedii.
Vyžaduje existující soubor se stejným
názvem z předchozího scrapingu.">
        <input type="checkbox" id="incremental"> Incremental (jen změněná hesla)
      </label>
    </div>
    </div>
  </div>

  <!-- BULK + UNDO toolbar (hidden until data loaded) -->
  <div class="bulk-toolbar" id="bulkToolbar">
    <span id="bulkCount">0 vybráno</span>
    <button class="btn-sm" onclick="bulkExport()" title="Exportovat vybrané záznamy jako JSON soubor">⬇ Exportovat výběr</button>
    <button class="btn-sm" onclick="bulkDelete()" style="color:var(--danger)" title="Odstranit vybrané záznamy z tabulky (soubor se nemění)">🗑 Odstranit</button>
    <span style="flex:1"></span>
    <button class="btn-sm" id="btnUndo" onclick="undo()" disabled title="Vrátit poslední akci v tabulce (Ctrl+Z)">↩</button>
    <button class="btn-sm" id="btnRedo" onclick="redo()" disabled title="Zopakovat vrácenou akci (Ctrl+Y)">↪</button>
    <button class="btn-sm" onclick="toggleBulkMode()" style="margin-left:4px" title="Zrušit hromadný výběr">✕ Zrušit výběr</button>
  </div>

  <!-- EXPORT -->
  <div class="export-box collapsible-section" id="exportBox" data-sec="export">
    <div class="export-title sec-header" style="cursor:pointer;user-select:none">📊 Stáhnout výsledky<span class="sec-caret" style="font-size:12px">▾</span></div>
    <div class="sec-body">
    <div class="export-status" id="sizeEstimate" style="margin-bottom:6px;color:var(--text2)"></div>
    <div id="exportFileList" style="margin-bottom:8px"></div>

    <!-- Rychlé stažení -->
    <div class="export-group">
      <div class="export-group-label">Rychlé stažení</div>
      <button class="btn btn-ghost export-btn" onclick="dlZip()" title="JSON + CSV jako ZIP">⬇ ZIP (JSON + CSV)</button>
      <button class="btn btn-primary export-btn" id="btnXlsx" onclick="dlXlsx()" title="Excel — otevře výběr infobox sloupců">⬇ Excel (.xlsx)</button>
    </div>

    <!-- Databáze a strukturovaná data -->
    <details class="export-details">
      <summary class="export-summary">🗄 Databáze &amp; strukturovaná data</summary>
      <div class="export-details-body">
        <button class="btn btn-ghost export-btn" id="btnSqlite" onclick="exportSqlite()" title="SQLite DB — tabulky articles + infobox_values">⬇ SQLite (.db)</button>
        <button class="btn btn-ghost export-btn" id="btnParquet" onclick="exportParquet()" title="Apache Parquet — ideální pro pandas/polars, snappy komprese">⬇ Parquet (.parquet)</button>
      </div>
    </details>

    <!-- Sémantická data -->
    <details class="export-details">
      <summary class="export-summary">🌐 Sémantická a dokumentační data</summary>
      <div class="export-details-body">
        <button class="btn btn-ghost export-btn" id="btnJsonLd" onclick="exportJsonLd()" title="JSON-LD se schema.org anotacemi — pro SEO a linked data">⬇ JSON-LD (.jsonld)</button>
        <button class="btn btn-ghost export-btn" id="btnReadme" onclick="exportReadme()" title="Automatický README.md s popisem datasetu">📄 README.md</button>
      </div>
    </details>

    <!-- Ostatní -->
    <details class="export-details">
      <summary class="export-summary">📁 Ostatní formáty a nástroje</summary>
      <div class="export-details-body">
        <button class="btn btn-ghost export-btn" id="btnTxtZip" onclick="exportTxtZip()" title="Jeden .txt soubor na heslo v ZIP archivu">📁 ZIP (textové soubory)</button>
        <button class="btn btn-ghost export-btn" onclick="openMerge()" title="Sloučit více výstupů do jednoho souboru">🔀 Sloučit soubory…</button>
      </div>
    </details>

    <button class="btn btn-danger" style="margin-top:10px;font-size:11px" onclick="openDiscardModal()">🗑 Zahodit výsledky…</button>
    <div class="export-status" id="xlsxSt"></div>
    <div class="export-status" id="extraSt"></div>
    </div>
  </div>

</div><!-- sidebar-scroll -->

<!-- SIDEBAR FOOTER — always visible -->
<div class="sidebar-footer">
  <button class="btn btn-primary hidden" id="btnRun" onclick="go(false)" title="Spustit scraper — fáze 1 sbírá kategorie, pak review (Space)">▶ Spustit scraper</button>
  <div class="hidden" id="btnStopGroup" style="display:none;flex-direction:column;gap:6px">
    <button class="btn btn-ghost hidden" id="btnPause" onclick="pauseScraping()" title="Pozastavit — uloží checkpoint, lze navázat">⏸ Pauza</button>
    <button class="btn btn-danger hidden" id="btnStop" onclick="stop()" title="Zastavit a zrušit scraping">■ Zastavit</button>
  </div>
</div>
  <div class="resize-handle-x" id="sidebarResizeHandle"></div>
</aside>

<!-- MOBILE SIDEBAR BACKDROP -->
<div id="mobSidebarBackdrop" onclick="mobCloseSidebar()"></div>

<!-- MOBILE BOTTOM NAV -->
<nav id="mobileNav" style="display:none">
  <button class="mob-nav-btn active" id="mobNavHome" onclick="mobTab('overview')" title="Přehled">
    <span class="nav-icon">📊</span>
    <span>Přehled</span>
  </button>
  <button class="mob-nav-btn" id="mobNavData" onclick="mobTab('data')" title="Výsledky">
    <span class="nav-icon">📋</span>
    <span>Výsledky</span>
  </button>
  <button class="mob-nav-btn" id="mobNavRun" onclick="go(false)" title="Spustit">
    <span class="nav-icon">▶</span>
    <span>Spustit</span>
  </button>
  <button class="mob-nav-btn" id="mobNavSidebar" onclick="mobToggleSidebar()" title="Nastavení">
    <span class="nav-icon">⚙</span>
    <span>Nastavení</span>
  </button>
  <button class="mob-nav-btn" id="mobNavViz" onclick="mobTab('viz')" title="Grafy">
    <span class="nav-icon">📈</span>
    <span>Grafy</span>
  </button>
</nav>

<!-- ═══ MAIN ═══ -->
<main class="main">

  <!-- PROGRESS -->
  <div class="progress-section">
    <div class="progress-bars">
      <div class="progress-row">
        <span class="progress-name">Fáze 1 · sběr URL</span>
        <div class="progress-track"><div class="progress-fill p1" id="pb1"></div></div>
        <span class="progress-pct" id="pp1">—</span>
      </div>
      <div class="progress-row">
        <span class="progress-name">Fáze 2 · stahování</span>
        <div class="progress-track"><div class="progress-fill p2" id="pb2"></div></div>
        <span class="progress-pct" id="pp2">—</span>
      </div>
    </div>
    <div class="current-task">
      <span class="task-phase p1" id="tPhase">IDLE</span>
      <span class="task-name"  id="tName">Čeká na spuštění…</span>
      <span class="task-eta"   id="tEta"></span>
    </div>
  </div>

  <!-- STATS -->
  <div class="stats-bar">
    <div class="stat"><div class="stat-val c-accent2" id="stCats">—</div><div class="stat-lbl">Kategorie</div></div>
    <div class="stat"><div class="stat-val c-accent"  id="stFound">—</div><div class="stat-lbl">Hesel</div></div>
    <div class="stat"><div class="stat-val c-accent2" id="stDone">—</div><div class="stat-lbl">Staženo</div></div>
    <div class="stat"><div class="stat-val"           id="stTime">—</div><div class="stat-lbl">Čas (s)</div></div>
    <div class="stat"><div class="stat-val c-warn"    id="stEta">—</div><div class="stat-lbl">Zbývá</div></div>
  </div>

  <!-- TABS -->
  <div class="tabs">
    <button class="tab-btn on" id="tabOverview" onclick="tab('overview')" title="Přehled statistik a top kategorií">Přehled</button>
    <button class="tab-btn"    id="tabData"     onclick="tab('data')"     title="Tabulka stažených hesel s filtrováním a vyhledáváním">Výsledky <span id="dataCnt"></span></button>
    <button class="tab-btn"    id="tabViz"      onclick="tab('viz')"      title="Grafy a vizualizace dat">Vizualizace</button>
    <button class="tab-btn"    id="tabMap"      onclick="tab('map')"      title="Geografická mapa hesel se souřadnicemi" style="display:none">🗺 Mapa</button>
    <button class="tab-btn"    id="tabTags"     onclick="tab('tags')"     title="Tag cloud z automaticky přiřazených tagů">🏷 Tagy</button>
    <button class="tab-btn"    id="tabValidate" onclick="tab('validate')" title="Validační report — chybějící pole, duplicity, stub články">✅ Validace</button>
    <div class="tab-actions">
      <button class="btn-sm" onclick="clearConsole()" title="Smazat výstup konzole">Vymazat log</button>
    </div>
  </div>

  <!-- CONSOLE (collapsible, below tabs, above tab panel) -->
  <div class="console-header" onclick="toggleConsole()">
    <span class="console-label">Konzole</span>
    <span class="console-caret" id="conCaret">▾</span>
  </div>
  <div class="console-body" id="conBody">
    <div class="console-inner" id="conInner"></div>
  </div>
  <div class="resize-handle-y" id="consoleResizeHandle" title="Táhni pro změnu výšky konzole"></div>

  <!-- TAB: OVERVIEW -->
  <div class="tab-panel on" id="panelOverview">
    <div class="overview-panel">
      <div class="overview-empty" id="overviewEmpty">
        <div class="overview-icon">📚</div>
        <div>Zadej URL a spusť scraper</div>
        <div style="font-size:11px;margin-top:4px;color:var(--text3)">Přehled a vizualizace se zobrazí po dokončení</div>
      </div>
      <div class="summary-cards hidden" id="summaryCards"></div>
    </div>
  </div>

  <!-- TAB: DATA -->
  <div class="tab-panel" id="panelData">
    <div class="dt-toolbar">
      <input class="dt-search" id="dtQ" placeholder="Hledat…" oninput="filterTbl()">
      <button class="dt-search-mode" id="btnFtMode" onclick="toggleFtMode()" title="Přepnout fulltext režim (AND/OR/NOT)">FT</button>
      <div class="col-hider-wrap">
        <button class="btn-sm" onclick="toggleColHiderPanel()" title="Skrýt/zobrazit sloupce">Sl ▾</button>
        <div class="col-hider-panel" id="colHiderPanel"></div>
      </div>
      <button class="btn-sm" onclick="toggleBulkMode()" title="Hromadný výběr" style="flex-shrink:0">☑</button>
      <span class="dt-count" id="dtCnt"></span>
    </div>
    <div style="display:flex;align-items:center;gap:8px;padding:6px 12px;background:var(--surface);border-bottom:1px solid var(--border);flex-shrink:0;flex-wrap:wrap">
      <div class="result-filters" id="resultFilters">
        <span style="font-family:var(--mono);font-size:10px;color:var(--text3)">Filtr:</span>
        <span class="rf-chip" id="rf-infobox"   onclick="toggleFilter('infobox')"   title="Jen hesla s infoboxem">📋 s infoboxem</span>
        <span class="rf-chip" id="rf-coords"    onclick="toggleFilter('coords')"    title="Jen hesla se souřadnicemi">📍 se souřadnicemi</span>
        <span class="rf-chip" id="rf-fulltext"  onclick="toggleFilter('fulltext')"  title="Jen hesla s plným textem">📄 s textem</span>
        <span class="rf-chip" id="rf-error"     onclick="toggleFilter('error')"     title="Hesla s chybou stažení">⚠ chyby</span>
        <span class="rf-chip" id="rf-images"    onclick="toggleFilter('images')"    title="Jen hesla s obrázky">🖼 s obrázky</span>
        <span class="rf-chip" id="rf-tables"    onclick="toggleFilter('tables')"    title="Jen hesla s tabulkami">📊 s tabulkami</span>
      </div>
      <span style="flex:1"></span>
      <span style="font-family:var(--mono);font-size:10px;color:var(--text3)" id="renameWrap" style="display:none">
        <span id="renameSection" style="display:none">
          Přejmenovat soubor: <input class="rename-input" id="renameInput" placeholder="nový název" style="width:120px">
          <button class="btn-sm" onclick="renameFile()" title="Potvrdit přejmenování">OK</button>
        </span>
        <button class="btn-sm" onclick="toggleRename()" id="btnRename" title="Přejmenovat výstupní soubor">✏ Přejmenovat</button>
      </span>
    </div>
    <div class="dt-wrap">
      <table>
        <thead><tr>
          <th class="col-c0 td-num">#</th>
          <th class="col-c1" onclick="sortBy('title')" title="Řadit podle názvu hesla">Název <span id="sTitle">↕</span><div class="th-resize" data-col="1"></div></th>
          <th class="col-c2" onclick="sortBy('intro')" title="Řadit podle perexu">Perex <span id="sIntro">↕</span><div class="th-resize" data-col="2"></div></th>
          <th class="col-c3" onclick="sortBy('categories')" title="Řadit podle kategorií">Kategorie <span id="sCats">↕</span><div class="th-resize" data-col="3"></div></th>
          <th class="col-c4" onclick="sortBy('_introLen')" title="Řadit podle délky perexu (znaky)">Délka ↕<div class="th-resize" data-col="4"></div></th>
          <th class="col-c5" onclick="sortBy('_quality')" title="Quality score 0–100 (perex, infobox, kategorie, souřadnice)">Q ↕<div class="th-resize" data-col="5"></div></th>
          <th class="col-c6" title="Počet polí v infoboxu">Infobox<div class="th-resize" data-col="6"></div></th>
          <th class="col-c7"></th>
        </tr></thead>
        <tbody id="dtBody"></tbody>
      </table>
      <div class="no-rows hidden" id="noRows">Nic nenalezeno</div>
      <div class="overview-empty" id="dataEmpty">
        <div class="overview-icon">📋</div>
        <div>Data se zobrazí po dokončení</div>
      </div>
    </div>
  </div>

  <!-- TAB: VIZ -->
  <div class="tab-panel" id="panelViz">
    <div class="viz-scroll" id="vizArea">
      <div class="overview-empty" style="width:100%">
        <div class="overview-icon">📊</div>
        <div>Vizualizace se zobrazí po dokončení</div>
      </div>
    </div>
    <!-- Live speed chart overlay -->
    <div class="live-chart-wrap" id="liveChartWrap">
      <div style="margin-bottom:4px">⚡ Rychlost stahování</div>
      <canvas id="liveSpeedChart"></canvas>
    </div>
  </div>

  <!-- MAP TAB -->
  <div class="tab-panel" id="panelMap" style="position:relative">
    <div id="mapContainer"></div>
    <div class="map-stats" id="mapStats" style="display:none"></div>
  </div>

  <!-- TAGS TAB -->
  <div class="tab-panel" id="panelTags">
    <div class="overview-panel">
      <div style="font-family:var(--mono);font-size:11px;color:var(--text3);margin-bottom:12px">
        Kliknutím na tag filtrovat tabulku výsledků
      </div>
      <div class="tag-cloud" id="tagCloud">
        <div class="overview-empty">
          <div class="overview-icon">🏷</div>
          <div>Tagy se zobrazí po načtení dat</div>
        </div>
      </div>
    </div>
  </div>

  <!-- VALIDATION TAB -->
  <div class="tab-panel" id="panelValidate">
    <div class="overview-panel" id="validateContent">
      <div class="overview-empty">
        <div class="overview-icon">✅</div>
        <div>Validace se zobrazí po načtení dat</div>
      </div>
    </div>
  </div>

</main>
</div><!-- body -->
</div><!-- app -->

<div class="ib-popup" id="ibPopup" onclick="this.style.display='none'"></div>
<div id="globalTooltip"></div>

<!-- INFO MODAL -->
<div class="info-backdrop" id="infoModal">
  <div class="info-modal">
    <div class="info-tabs">
      <div class="info-tab active" id="infoTab0" onclick="switchInfoTab(0)">ℹ O aplikaci</div>
      <div class="info-tab" id="infoTab1" onclick="switchInfoTab(1)">📋 Changelog</div>
    </div>
    <div class="info-body" id="infoBody0">
      <h2>📚 WikiScraper v6</h2>
      <p>Nástroj pro hromadné stahování dat z Wikipedie do strukturovaných formátů (JSON, CSV, XLSX, SQLite).</p>
      <p class="info-credit">Vytvořil <strong>Tomáš Alvarez</strong> s přispěním <strong>Claude (Anthropic)</strong>, 2026</p>
      <h3>Jak začít</h3>
      <p>1. Vlož URL Wikipedie kategorie nebo seznamu do pole <kbd>Cílová URL</kbd><br>
         2. Klikni <kbd>🔍 Otestovat</kbd> pro ověření URL a odhad dat<br>
         3. Zvol pole která chceš stáhnout a nastav parametry<br>
         4. Klikni <kbd>▶ Spustit</kbd> — fáze 1 sbírá kategorie<br>
         5. V Review dialogu roztřiď kategorie do souborů<br>
         6. Fáze 2 stáhne hesla — po dokončení exportuj</p>
      <h3>Klávesové zkratky</h3>
      <p><kbd>Space</kbd> spustit/zastavit &nbsp; <kbd>Esc</kbd> zavřít dialog &nbsp; <kbd>Ctrl+Enter</kbd> potvrdit review<br>
         <kbd>Ctrl+F</kbd> hledat v tabulce &nbsp; <kbd>↑↓</kbd> navigace v review &nbsp; <kbd>→+</kbd> přesunout do nového sloupce</p>
      <h3>Pokročilé funkce</h3>
      <p>
        <strong>API režim</strong> — rychlejší stahování přes MediaWiki API (nastavení → Workers & API)<br>
        <strong>Paralelní stahování</strong> — až 5 vláken současně, nastavitelné<br>
        <strong>Wikidata</strong> — obohacení dat o Wikidata properties<br>
        <strong>Fulltext search</strong> — klikni FT v tabulce, podporuje AND/OR/NOT a "fráze"<br>
        <strong>Sloučit soubory</strong> — v export boxu po dokončení<br>
        <strong>SQLite export</strong> — tabulky articles + infobox_values pro přímou SQL analýzu
      </p>
    </div>
    <div class="info-body" id="infoBody1" style="display:none">
      <h2>Changelog</h2>
      <div class="cl-entry">
        <span class="cl-version">v8</span><span class="cl-date">2026</span>
        <div class="cl-items">
          ✦ Rate limiting — per-domain adaptivní backoff, exponential retry s jitterem<br>
          ✦ Sanitizace dat — wiki markup, unicode normalizace, čištění infoboxu<br>
          ✦ Detekce anomálií — 10 pravidel, badge ⚑ v tabulce, filtry ve Validaci<br>
          ✦ Export Parquet — Apache Parquet (snappy), ideální pro pandas/polars<br>
          ✦ Export JSON-LD — schema.org anotace pro linked data a SEO<br>
          ✦ Export UI redesign — skupiny formátů v rozklikávacích <details><br>
          ✦ Strukturované logy — JSON rotující soubory, úrovně, separátní audit log<br>
          ✦ OpenAPI 3.0 spec — kompletní dokumentace všech routes<br>
          ✦ Swagger UI — interaktivní API dokumentace na /api/docs<br>
          ✦ Unit testy — extraktory, sanitizace, rate limiting, Flask routes (pytest)
        </div>
      </div>
      <div class="cl-entry">
        <span class="cl-version">v7</span><span class="cl-date">2026</span>
        <div class="cl-items">
          ✦ Strom kategorií — nový krok před review, hierarchický výběr větví<br>
          ✦ Obrázky — thumbnail strip, lightbox s licencí a autorem (Wikimedia API)<br>
          ✦ Tabulky — wikitable → JSON, popup s náhledem, záložky pro více tabulek<br>
          ✦ Wildcard v blacklistu — hydro*, *natý, *chemi*<br>
          ✦ Export / import stavu review (.review.json)<br>
          ✦ Tlačítko Pauza — uloží checkpoint, scraping jde obnovit<br>
          ✦ Přepnutí projektu nepřerušuje běžící scraping<br>
          ✦ Smazání položky z historie stahování<br>
          ✦ Vizuální feedback při výběru rychlého příkladu (URL)<br>
          ✦ Výřazené kategorie blacklistem označeny 🚫 v review<br>
          ✦ Počet blacklistem vyřazených kategorií v review dialogu<br>
          ✦ Smart tooltip — skutečné rozměry, nepřeteče viewport<br>
          ✦ Konzole roztažitelná na celou výšku obrazovky<br>
          ✦ Záložka Mapa skryta pokud nejsou GPS souřadnice<br>
          ✦ Resize sloupců tabulky, sidebar a konzole drag & drop<br>
          ✦ Kolapsovatelné sekce sidebaru (stav pamatuje localStorage)<br>
          ✦ Odrážky (ul/ol/li) správně extrahované v HTML i API módu<br>
          ✦ Izolace sessions pro víceuživatelský provoz (UUID prefix)<br>
          ✦ Automatické mazání souborů starších 60 minut<br>
          ✦ Opraveny: filtry křížové resetování, resumé UI, Esc pro modály,<br>
          &nbsp;&nbsp;checkCP session prefix, drag v search poli review, UNDO_MAX=7
        </div>
      </div>
      <div class="cl-entry">
        <span class="cl-version">v6</span><span class="cl-date">2026</span>
        <div class="cl-items">
          ✦ Paralelní stahování (2–5 vláken)<br>
          ✦ MediaWiki API režim (rychlejší, méně rate limitů)<br>
          ✦ Wikidata obohacení<br>
          ✦ Incremental update (jen změněná hesla)<br>
          ✦ Fuzzy dedup titulů (Levenshtein)<br>
          ✦ Auto-tagy z kategorií<br>
          ✦ Validační report<br>
          ✦ Geografická mapa výsledků (Leaflet)<br>
          ✦ Tag cloud s filtrováním<br>
          ✦ Live chart rychlosti stahování<br>
          ✦ Globální search přes všechny soubory<br>
          ✦ Notifikační centrum<br>
          ✦ Onboarding pro nové uživatele<br>
          ✦ Hromadný výběr v tabulce<br>
          ✦ Undo/redo editací<br>
          ✦ Info modal + Changelog
        </div>
      </div>
      <div class="cl-entry">
        <span class="cl-version">v5</span><span class="cl-date">2026</span>
        <div class="cl-items">
          ✦ Multi-projekt záložky<br>
          ✦ Uložení stavu review<br>
          ✦ Drag & drop URL<br>
          ✦ Test URL před spuštěním<br>
          ✦ Blacklist kategorií<br>
          ✦ Reveal animace po fázi 1<br>
          ✦ Tooltips na polích<br>
          ✦ Quality score (0–100)<br>
          ✦ Normalizace souřadnic<br>
          ✦ Detekce stub článků<br>
          ✦ SQLite + README export<br>
          ✦ Merge souborů<br>
          ✦ Column picker pro XLSX<br>
          ✦ Fulltext AND/OR/NOT search
        </div>
      </div>
      <div class="cl-entry">
        <span class="cl-version">v4</span><span class="cl-date">2026</span>
        <div class="cl-items">
          ✦ Multi-sloupce v review dialogu<br>
          ✦ Stepper průběhu<br>
          ✦ Notifikace po dokončení<br>
          ✦ Historie spuštění<br>
          ✦ Odhad velikosti souboru<br>
          ✦ Klávesové zkratky<br>
          ✦ Dark/light mode<br>
          ✦ Automatická deduplikace<br>
          ✦ Klávesová navigace v review
        </div>
      </div>
      <div class="cl-entry">
        <span class="cl-version">v3</span><span class="cl-date">2026</span>
        <div class="cl-items">
          ✦ Dvoustupňový scraping (fáze 1 + review + fáze 2)<br>
          ✦ Selektivní pole<br>
          ✦ Export XLSX s infoboxem<br>
          ✦ Rate limit handling<br>
          ✦ Checkpoint & resume
        </div>
      </div>
      <div class="cl-entry">
        <span class="cl-version">v1–v2</span><span class="cl-date">2026</span>
        <div class="cl-items">
          ✦ Základní Flask GUI<br>
          ✦ Rekurzivní scraping kategorií<br>
          ✦ JSON + CSV výstup<br>
          ✦ SSE live log
        </div>
      </div>
    </div>
    <div class="info-footer">
      <button class="btn btn-ghost" style="width:auto;padding:7px 16px" onclick="document.getElementById('infoModal').classList.remove('on')" title="Zavřít nápovědu">Zavřít</button>
    </div>
  </div>
</div>

<!-- ONBOARDING OVERLAY -->
<div class="onboard-backdrop" id="onboardModal">
  <div class="onboard-card">
    <div class="onboard-icon">📚</div>
    <div class="onboard-title">Vítej ve WikiScraperu</div>
    <div class="onboard-sub">Nástroj pro stahování dat z Wikipedie bez jediného řádku kódu. Tři kroky a máš data.</div>
    <div class="onboard-steps">
      <div class="onboard-step"><span class="onboard-step-num">1</span>Vlož URL Wikipedie kategorie a klikni Otestovat</div>
      <div class="onboard-step"><span class="onboard-step-num">2</span>Spusť → vyber kategorie v Review dialogu</div>
      <div class="onboard-step"><span class="onboard-step-num">3</span>Stáhni výsledky jako Excel, JSON nebo SQLite</div>
    </div>
    <button class="btn btn-primary" style="width:auto;padding:10px 28px" onclick="closeOnboard()" title="Zavřít uvítací obrazovku a začít">Začít →</button>
    <div style="margin-top:12px;font-family:var(--mono);font-size:10px;color:var(--text3)">
      Potřebuješ nápovědu? Klikni na ℹ v pravém horním rohu.
    </div>
  </div>
</div>
<div class="merge-backdrop" id="mergeModal">
  <div class="merge-modal">
    <div class="merge-header">
      <div class="merge-title">🔀 Sloučit soubory</div>
    </div>
    <div class="merge-body">
      <div class="merge-left">
        <div class="merge-section-label">Dostupné soubory</div>
        <div id="mergeFileList"></div>
      </div>
      <div style="display:flex;align-items:center;padding:0 8px;color:var(--text3);font-size:20px">→</div>
      <div class="merge-right">
        <div class="merge-section-label">Výstupní soubor</div>
        <input class="merge-output-input" type="text" id="mergeOutput" value="merged" placeholder="název výsledného souboru">
        <div class="merge-section-label" style="margin-top:12px">Vybrané soubory</div>
        <div id="mergeSelected" style="font-family:var(--mono);font-size:11px;color:var(--text3);min-height:60px">
          Klikni na soubory vlevo…
        </div>
        <div class="merge-result" id="mergeResult"></div>
      </div>
    </div>
    <div class="merge-footer">
      <span style="font-family:var(--mono);font-size:11px;color:var(--text3)" id="mergeCnt">0 souborů vybráno</span>
      <span style="flex:1"></span>
      <button class="btn btn-ghost" style="width:auto;padding:7px 14px" onclick="document.getElementById('mergeModal').classList.remove('on')" title="Zavřít bez sloučení">Zavřít</button>
      <button class="btn btn-primary" style="width:auto;padding:7px 18px" onclick="doMerge()" title="Sloučit vybrané soubory do jednoho — duplicity se odstraní podle URL">🔀 Sloučit</button>
    </div>
  </div>
</div>

<!-- CONTEXT PREVIEW (cat hover in review) -->
<div class="cat-preview" id="catPreview">
  <div class="cat-preview-title" id="cpTitle"></div>
  <div class="cat-preview-meta" id="cpMeta"></div>
  <div class="cat-preview-items" id="cpItems"></div>
</div>

<!-- COLPICK MODAL -->
<div class="colpick-backdrop" id="colpickModal">
  <div class="colpick-modal">
    <div class="colpick-header">
      <div class="colpick-title">Vybrat sloupce pro Excel export</div>
      <div class="colpick-sub" id="cpSub">Načítám…</div>
    </div>
    <div class="colpick-toolbar">
      <button class="btn-sm" onclick="cpSelectAll(true)" title="Vybrat všechny sloupce">✓ Vše</button>
      <button class="btn-sm" onclick="cpSelectAll(false)" title="Odebrat výběr všech sloupců">✗ Nic</button>
      <button class="btn-sm" onclick="cpSelectAboveThresh()" title="Vybrat jen sloupce které jsou přítomny u více než X % hesel">✓ Nad prahem</button>
      <div class="thresh-wrap">
        Prah: <input type="number" id="cpThresh" value="20" min="1" max="100" step="1" oninput="cpUpdateThresh()">%
      </div>
    </div>
    <div class="colpick-body" id="cpBody"></div>
    <div class="colpick-footer">
      <div class="colpick-count" id="cpCount"></div>
      <button class="btn btn-ghost" style="width:auto;padding:7px 14px" onclick="document.getElementById('colpickModal').classList.remove('on')" title="Zavřít bez exportu">Zrušit</button>
      <button class="btn btn-primary" style="width:auto;padding:7px 18px" onclick="cpConfirmExport()" title="Exportovat vybrané sloupce do XLSX">⬇ Exportovat</button>
    </div>
  </div>
</div>


<!-- IMAGE LIGHTBOX -->
<div class="img-lightbox" id="imgLightbox" onclick="closeLightbox()">
  <img id="lightboxImg" src="" alt="">
  <div class="img-lightbox-cap" id="lightboxCap"></div>
  <div class="img-lightbox-meta" id="lightboxMeta"></div>
</div>

<!-- TABLE POPUP -->
<div class="tbl-popup" id="tblPopup">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
    <span style="font-family:var(--mono);font-size:11px;font-weight:600;color:var(--text)">Tabulky hesla</span>
    <button class="btn-sm" onclick="document.getElementById('tblPopup').classList.remove('on')" style="padding:2px 8px">✕</button>
  </div>
  <div class="tbl-nav" id="tblNav"></div>
  <div id="tblContent"></div>
</div>

<!-- CATEGORY TREE MODAL -->
<div class="tree-backdrop" id="treeModal">
  <div class="tree-modal">
    <div class="tree-header">
      <div class="tree-title">Strom kategorií</div>
      <div class="tree-sub" id="treeSub">Načítám…</div>
    </div>
    <div class="tree-toolbar">
      <button class="btn-sm" onclick="treeExpandAll()">+ Rozbalit vše</button>
      <button class="btn-sm" onclick="treeCollapseAll()">− Sbalit vše</button>
      <button class="btn-sm" onclick="treeCheckAll(true)">✓ Vše</button>
      <button class="btn-sm" onclick="treeCheckAll(false)">✗ Nic</button>
      <span style="flex:1"></span>
      <span style="font-family:var(--mono);font-size:10px;color:var(--text3)" id="treeColMode">
        Každá zaškrtnutá větev = 1 sloupec
      </span>
    </div>
    <div class="tree-body" id="treeBody"></div>
    <div class="tree-footer">
      <div class="tree-summary" id="treeSummary"></div>
      <button class="btn btn-ghost" style="width:auto;padding:8px 16px"
        onclick="treeSkip()" title="Přeskočit strom — otevřít klasický review se vším">
        Přeskočit → klasický review
      </button>
      <button class="btn btn-primary" style="width:auto;padding:8px 20px"
        id="btnTreeConfirm" onclick="treeConfirm()">
        Potvrdit výběr →
      </button>
    </div>
  </div>
</div>

<!-- CATEGORY REVIEW MODAL -->
<div class="modal-backdrop" id="reviewModal">
  <div class="modal">
    <div class="modal-header">
      <div>
        <div class="modal-title">Roztřiď kategorie do souborů</div>
        <div class="modal-sub" id="reviewSub">—</div>
      </div>
    </div>
    <!-- UNDO/REDO BAR -->
    <div class="review-undo-bar" id="reviewUndoBar">
      <button class="btn-sm" id="btnReviewUndo" onclick="undoReview()" disabled title="Ctrl+Z">↩ Zpět</button>
      <button class="btn-sm" id="btnReviewRedo" onclick="redoReview()" disabled title="Ctrl+Y">↪ Znovu</button>
      <span class="review-undo-label" id="reviewUndoLabel"></span>
    </div>
    <!-- REVEAL ANIMATION (shown first, then transitions to columns) -->
    <div class="review-reveal" id="reviewReveal">
      <div class="rv-number" id="rvNumber">0</div>
      <div class="rv-label">kategorií nalezeno</div>
      <div class="rv-sub" id="rvSub"></div>
      <button class="btn btn-primary rv-btn" onclick="showReviewColumns()" title="Přejít na výběr a třídění kategorií">
        Roztřídit → <span id="rvHesel"></span>
      </button>
    </div>
    <!-- Columns rendered by JS (hidden until reveal done) -->
    <div class="review-columns" id="reviewColumns" style="display:none"></div>
    <div class="modal-footer" id="reviewFooter" style="display:none">
      <div class="review-summary" id="reviewSummary"></div>
      <div style="font-family:var(--mono);font-size:10px;color:var(--text3)" id="reviewSizeEst"></div>
      <div style="font-family:var(--mono);font-size:10px;color:var(--text3);margin-right:auto;line-height:1.7">
        <kbd style="background:var(--surface2);border:1px solid var(--border2);border-radius:3px;padding:1px 5px">↑↓</kbd> navigace &nbsp;
        <kbd style="background:var(--surface2);border:1px solid var(--border2);border-radius:3px;padding:1px 5px">Space</kbd> zaškrtnout/odškrtnout &nbsp;
        <kbd style="background:var(--surface2);border:1px solid var(--border2);border-radius:3px;padding:1px 5px">→</kbd> přesunout do vedlejšího sloupce
      </div>
      <button class="btn btn-ghost" style="width:auto;padding:7px 12px" onclick="exportReviewFile()" title="Stáhnout review jako .json soubor — načíst zpět přes Načíst review">💾 Uložit review</button>
      <button class="btn btn-ghost" style="width:auto;padding:8px 16px" onclick="closeReview()" title="Zavřít review a vrátit se (stav se uloží)">Zrušit</button>
      <button class="btn btn-primary" style="width:auto;padding:8px 20px" id="btnStartScraping" onclick="startPhase2()" title="Spustit stahování vybraných hesel (Ctrl+Enter)">
        ▶ Spustit stahování
      </button>
    </div>
  </div>
</div>

<!-- DISCARD RESULTS MODAL -->
<div class="modal-backdrop" id="discardModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:400;align-items:center;justify-content:center">
  <div style="background:var(--surface);border:1px solid var(--border2);border-radius:var(--radius);width:420px;max-width:95vw;padding:24px;box-shadow:0 16px 48px rgba(0,0,0,.6)">
    <div style="font-size:15px;font-weight:600;margin-bottom:8px">Zahodit výsledky?</div>
    <div style="font-family:var(--mono);font-size:11px;color:var(--text2);margin-bottom:20px;line-height:1.6">
      Soubory na serveru budou smazány.<br>Co chceš udělat dál?
    </div>
    <div style="display:flex;flex-direction:column;gap:8px">
      <button class="btn btn-ghost" onclick="discardAndReview()" style="justify-content:flex-start;text-align:left">
        ↩ Vrátit se do review — vybrat jiné kategorie
      </button>
      <button class="btn btn-ghost" onclick="discardCompletely()" style="justify-content:flex-start;text-align:left;color:var(--danger)">
        🗑 Zahodit úplně — začít znovu od URL
      </button>
      <button class="btn btn-ghost" onclick="document.getElementById('discardModal').style.display='none'" style="margin-top:4px">
        Zrušit
      </button>
    </div>
  </div>
</div>

<script>
/* ── EXAMPLES ─────────────────────────────────────────────────────────────── */
const EX=[
  {i:'⚗️',t:'Chemické prvky',      u:'https://cs.wikipedia.org/wiki/Kategorie:Chemické_prvky',              n:'chemicke_prvky'},
  {i:'🧪',t:'Chem. sloučeniny',    u:'https://cs.wikipedia.org/wiki/Kategorie:Chemické_sloučeniny',         n:'chem_slouceniny'},
  {i:'🦕',t:'Dinosauři',           u:'https://cs.wikipedia.org/wiki/Kategorie:Dinosauři',                   n:'dinosauri'},
  {i:'🐦',t:'Ptáci',               u:'https://cs.wikipedia.org/wiki/Kategorie:Ptáci',                       n:'ptaci'},
  {i:'🌿',t:'Léčivé rostliny',     u:'https://cs.wikipedia.org/wiki/Kategorie:Léčivé_rostliny',             n:'rostliny'},
  {i:'🍄',t:'Houby',               u:'https://cs.wikipedia.org/wiki/Kategorie:Houby',                       n:'houby'},
  {i:'🐟',t:'Ryby',                u:'https://cs.wikipedia.org/wiki/Kategorie:Ryby',                        n:'ryby'},
  {i:'🦈',t:'Žraloci',             u:'https://cs.wikipedia.org/wiki/Kategorie:Žraloci',                     n:'zraloci'},
  {i:'🌋',t:'Sopky světa',         u:'https://cs.wikipedia.org/wiki/Kategorie:Sopky',                       n:'sopky'},
  {i:'🏔',t:'Pohoří světa',        u:'https://cs.wikipedia.org/wiki/Kategorie:Pohoří',                      n:'pohori'},
  {i:'🌊',t:'Řeky světa',          u:'https://cs.wikipedia.org/wiki/Kategorie:Řeky',                        n:'reky'},
  {i:'🏝',t:'Ostrovy',             u:'https://cs.wikipedia.org/wiki/Kategorie:Ostrovy',                     n:'ostrovy'},
  {i:'🔬',t:'Fyzikální jevy',      u:'https://cs.wikipedia.org/wiki/Kategorie:Fyzikální_jevy',              n:'fyzikalni_jevy'},
  {i:'🧬',t:'Nemoci',              u:'https://cs.wikipedia.org/wiki/Kategorie:Nemoci',                      n:'nemoci'},
  {i:'💊',t:'Léky',                u:'https://cs.wikipedia.org/wiki/Kategorie:Léky',                        n:'leky'},
  {i:'🌌',t:'Astronomické obj.',   u:'https://cs.wikipedia.org/wiki/Kategorie:Astronomické_objekty',        n:'astronomie'},
  {i:'🪐',t:'Planety',             u:'https://cs.wikipedia.org/wiki/Kategorie:Planety',                     n:'planety'},
  {i:'⭐',t:'Hvězdy',              u:'https://cs.wikipedia.org/wiki/Kategorie:Hvězdy',                      n:'hvezdy'},
  {i:'🦠',t:'Bakterie',            u:'https://cs.wikipedia.org/wiki/Kategorie:Bakterie',                    n:'bakterie'},
  {i:'🧠',t:'Psychické poruchy',   u:'https://cs.wikipedia.org/wiki/Kategorie:Psychické_poruchy',           n:'psych_poruchy'},
  {i:'🏙',t:'Česká města',         u:'https://cs.wikipedia.org/wiki/Kategorie:Česká_města',                 n:'ceska_mesta'},
  {i:'🏰',t:'Hrady v Čechách',     u:'https://cs.wikipedia.org/wiki/Kategorie:Hrady_v_Čechách',             n:'hrady'},
  {i:'🏯',t:'Zámky v Čechách',     u:'https://cs.wikipedia.org/wiki/Kategorie:Zámky_v_Čechách',             n:'zamky'},
  {i:'⛪',t:'Kláštery v Česku',    u:'https://cs.wikipedia.org/wiki/Kategorie:Kláštery_v_Česku',            n:'klastery'},
  {i:'🌍',t:'Státy světa',         u:'https://cs.wikipedia.org/wiki/Kategorie:Státy_světa',                 n:'staty_sveta'},
  {i:'🗺',t:'Hlavní města',        u:'https://cs.wikipedia.org/wiki/Kategorie:Hlavní_města',                n:'hlavni_mesta'},
  {i:'🏔',t:'Národní parky',       u:'https://cs.wikipedia.org/wiki/Kategorie:Národní_parky',               n:'nar_parky'},
  {i:'🎬',t:'České filmy',         u:'https://cs.wikipedia.org/wiki/Kategorie:České_filmy',                 n:'ceske_filmy'},
  {i:'🎭',t:'Čeští herci',         u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_herci',                 n:'cesti_herci'},
  {i:'🎸',t:'Čs. hudebníci',       u:'https://cs.wikipedia.org/wiki/Kategorie:Českoslovenští_hudebníci',    n:'hudebnici'},
  {i:'📚',t:'Čeští spisovatelé',   u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_spisovatelé',           n:'spisovatele'},
  {i:'🎨',t:'Čeští malíři',        u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_malíři',                n:'maliri'},
  {i:'🏛',t:'Starověký Řím',       u:'https://cs.wikipedia.org/wiki/Kategorie:Starověký_Řím',               n:'rim'},
  {i:'🏺',t:'Starověké Řecko',     u:'https://cs.wikipedia.org/wiki/Kategorie:Starověké_Řecko',             n:'recko'},
  {i:'⚔️',t:'Historické bitvy',   u:'https://cs.wikipedia.org/wiki/Kategorie:Bitvy',                       n:'bitvy'},
  {i:'👑',t:'Čeští panovníci',     u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_panovníci',             n:'panovnici'},
  {i:'🤝',t:'Čeští politici',      u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_politici',              n:'cesti_politici'},
  {i:'⚽',t:'Čeští fotbalisté',    u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_fotbalisté',            n:'fotbaliste'},
  {i:'🏒',t:'Čeští hokejisté',     u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_lední_hokejisté',       n:'hokejiste'},
  {i:'🎾',t:'Čeští tenisté',       u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_tenisté',               n:'teniste'},
  {i:'🚴',t:'Čeští cyklisté',      u:'https://cs.wikipedia.org/wiki/Kategorie:Čeští_cyklisté',              n:'cykliste'},
  {i:'💻',t:'Prog. jazyky',        u:'https://en.wikipedia.org/wiki/List_of_programming_languages',         n:'prog_langs'},
  {i:'🤖',t:'AI projekty',         u:'https://en.wikipedia.org/wiki/List_of_artificial_intelligence_projects', n:'ai_projects'},
  {i:'🚀',t:'Mise NASA',           u:'https://en.wikipedia.org/wiki/List_of_NASA_missions',                 n:'nasa'},
  {i:'🛸',t:'Vesmírné sondy',      u:'https://en.wikipedia.org/wiki/List_of_Solar_System_probes',           n:'sondy'},
  {i:'🏅',t:'Nobel — fyzika',      u:'https://en.wikipedia.org/wiki/List_of_Nobel_laureates_in_Physics',    n:'nobel_fyzika'},
  {i:'🏅',t:'Nobel — chemie',      u:'https://en.wikipedia.org/wiki/List_of_Nobel_laureates_in_Chemistry',  n:'nobel_chemie'},
  {i:'🏅',t:'Nobel — medicína',    u:'https://en.wikipedia.org/wiki/List_of_Nobel_laureates_in_Physiology_or_Medicine', n:'nobel_medicina'},
  {i:'🌍',t:'Countries by pop.',   u:'https://en.wikipedia.org/wiki/List_of_countries_by_population_(United_Nations)', n:'countries_pop'},
  {i:'🏙',t:'Largest cities',      u:'https://en.wikipedia.org/wiki/List_of_largest_cities',                n:'largest_cities'},
  {i:'🌉',t:'Longest bridges',     u:'https://en.wikipedia.org/wiki/List_of_longest_bridges',               n:'bridges'},
  {i:'🏗',t:'Tallest buildings',   u:'https://en.wikipedia.org/wiki/List_of_tallest_buildings',             n:'buildings'},
  {i:'🦁',t:'Endangered species',  u:'https://en.wikipedia.org/wiki/List_of_endangered_species',            n:'endangered'},
  {i:'🎬',t:'Oscar best picture',  u:'https://en.wikipedia.org/wiki/Academy_Award_for_Best_Picture',        n:'oscar'},
  {i:'🏎',t:'F1 World Champions',  u:'https://en.wikipedia.org/wiki/List_of_Formula_One_World_Drivers%27_Champions', n:'f1'},
  {i:'⚽',t:'FIFA World Cup',      u:'https://en.wikipedia.org/wiki/FIFA_World_Cup',                        n:'worldcup'},
];

function shuffle(a){
  const b=[...a];
  for(let i=b.length-1;i>0;i--){const j=Math.floor(Math.random()*(i+1));const t=b[i];b[i]=b[j];b[j]=t;}
  return b;
}

let _lastEx=[];
function renderEx(){
  const s=shuffle(EX);
  // Prefer items not shown last time
  s.sort((a,b)=>(_lastEx.includes(a.u)?1:0)-(_lastEx.includes(b.u)?1:0));
  const picks=s.slice(0,4);
  _lastEx=picks.map(p=>p.u);

  const list=document.getElementById('exList');
  list.innerHTML=picks.map(e=>`
    <button class="ex-btn">
      <span class="ex-icon">${e.i}</span>
      <span class="ex-info">
        <span class="ex-name">${e.t}</span>
        <span class="ex-url">${e.u.replace('https://','')}</span>
      </span>
      <span class="ex-pick-indicator" style="font-family:var(--mono);font-size:9px;color:var(--accent);opacity:0;transition:opacity .2s;margin-left:auto;flex-shrink:0">✓ vybráno</span>
    </button>`).join('');
  list.querySelectorAll('.ex-btn').forEach((btn,i)=>btn.onclick=()=>{
    document.getElementById('url').value=picks[i].u;
    document.getElementById('output').value=picks[i].n;
    checkCP();
    // Vizuální feedback — zvýraznit vybranou položku
    list.querySelectorAll('.ex-btn').forEach(b=>{
      b.style.borderColor='';
      b.style.background='';
      const ind=b.querySelector('.ex-pick-indicator');
      if(ind) ind.style.opacity='0';
    });
    btn.style.borderColor='var(--accent)';
    btn.style.background='rgba(110,231,183,.06)';
    const ind=btn.querySelector('.ex-pick-indicator');
    if(ind) ind.style.opacity='1';
    // Pulzující efekt na URL poli
    const urlEl=document.getElementById('url');
    urlEl.style.transition='box-shadow .15s';
    urlEl.style.boxShadow='0 0 0 2px rgba(110,231,183,.35)';
    setTimeout(()=>{ urlEl.style.boxShadow=''; },600);
  });
}

/* ── FIELDS ───────────────────────────────────────────────────────────────── */
const FIELDS=[
  {id:'title',      label:'Název',       desc:'h1',              def:true,  slow:false, tip:'Název hesla\nPříklad: "Kyslík"'},
  {id:'intro',      label:'Perex',       desc:'1. odstavec',     def:true,  slow:false, tip:'První odstavec článku\nPříklad: "Kyslík je chemický prvek…"'},
  {id:'infobox',    label:'Infobox',     desc:'tabulka dat',     def:true,  slow:false, tip:'Strukturovaná data z pravého panelu\nPříklad: {"Protonové číslo": "8", "Hmotnost": "15.999"}'},
  {id:'categories', label:'Kategorie',   desc:'wiki kategorie',  def:true,  slow:false, tip:'Wiki kategorie hesla\nPříklad: ["Chemické prvky", "Plyny"]'},
  {id:'coordinates',label:'Souřadnice',  desc:'GPS poloha',      def:true,  slow:false, tip:'GPS souřadnice (jen u míst)\nPříklad: "50°5′N 14°25′E"'},
  {id:'full_text',  label:'Plný text',   desc:'všechny odstavce',def:false, slow:true,  tip:'⚠ Pomalé — celý text článku\nJeden dlouhý string všech odstavců'},
  {id:'sections',   label:'Sekce',       desc:'dle nadpisů',     def:false, slow:true,  tip:'⚠ Pomalé — text rozdělený dle nadpisů\nPříklad: {"Výskyt": "…", "Vlastnosti": "…"}'},
  {id:'links',      label:'Ext. odkazy', desc:'mimo Wikipedii',  def:false, slow:false, tip:'Ext. odkazyz článku\nPříklad: [{"url": "https://…", "text": "…"}]'},
  {id:'images',     label:'Obrázky',     desc:'url + popis + licence', def:false, slow:true,  tip:'Obrázky z článku s thumbnail URL,\npopiskem a metadaty (licence, autor)\nz Wikimedia API.\nZpomalí stahování u článků s obrázky.'},
  {id:'tables',     label:'Tabulky',     desc:'wikitable → data',      def:false, slow:false, tip:'Strukturovaná data z wikitable tabulek\n(záhlaví + řádky jako JSON).\nPřeskočí infobox tabulky.\nVhodné pro datasety se srovnávacími tabulkami.'},
];
const fState={};
FIELDS.forEach(f=>{fState[f.id]=f.def;});

function renderFields(){
  document.getElementById('fieldGrid').innerHTML=FIELDS.map(f=>`
    <div class="field-toggle${f.def?' on':''}" id="ft_${f.id}" data-tip="${escAttr(f.tip||'')}">
      <div class="ft-box" id="fb_${f.id}">${f.def?'✓':''}</div>
      <div class="ft-text">
        <span class="ft-name">${f.label}</span>
        <span class="ft-desc">${f.desc}</span>
      </div>
    </div>`).join('');
  FIELDS.forEach(f=>{
    document.getElementById('ft_'+f.id).onclick=()=>{
      if(f.id==='title') return;
      fState[f.id]=!fState[f.id];
      document.getElementById('ft_'+f.id).classList.toggle('on',fState[f.id]);
      document.getElementById('fb_'+f.id).textContent=fState[f.id]?'✓':'';
      document.getElementById('slowWarn').classList.toggle('on',
        FIELDS.filter(x=>x.slow).some(x=>fState[x.id]));
    };
  });
}
function getFields(){ return FIELDS.filter(f=>fState[f.id]).map(f=>f.id).join(','); }

/* ── STATUS / TIMER ───────────────────────────────────────────────────────── */
let evtSrc=null, t0=null, timerInt=null;
let allData=[], sk='', sasc=true;
let _phase2Running=false;  // true během fáze 2 stahování

function setStatus(s){
  const labels={running:'Běží',done:'Hotovo',error:'Chyba',idle:'Idle'};
  document.getElementById('stDot').className='status-dot '+s;
  document.getElementById('stLbl').textContent=labels[s]||s;
}
function startTimer(){t0=Date.now();timerInt=setInterval(()=>document.getElementById('stTime').textContent=((Date.now()-t0)/1000).toFixed(1),200);}
function stopTimer(){clearInterval(timerInt);}
function eta(s){if(s==null||s<0||!isFinite(s))return '—';if(s<60)return Math.round(s)+'s';return (s/60).toFixed(1)+'min';}

function setPhase(p,name,sub){
  const ph=document.getElementById('tPhase');
  ph.textContent=p===1?'FÁZE 1':'FÁZE 2';
  ph.className='task-phase '+(p===1?'p1':'p2');
  document.getElementById('tName').textContent=name;
  document.getElementById('tEta').textContent=sub||'';
}

/* ── CONSOLE ──────────────────────────────────────────────────────────────── */
let conOpen=false;
function toggleConsole(){
  conOpen=!conOpen;
  document.getElementById('conBody').classList.toggle('open',conOpen);
  document.getElementById('conCaret').classList.toggle('open',conOpen);
}
function clog(msg,type=''){
  const w=document.getElementById('conInner');
  const t=new Date().toTimeString().slice(0,8);
  const d=document.createElement('div'); d.className='log-row';
  d.innerHTML=`<span class="log-t">${t}</span><span class="log-m ${type}">${esc(msg)}</span>`;
  w.appendChild(d); w.scrollTop=w.scrollHeight;
}
function clearConsole(){document.getElementById('conInner').innerHTML='';}

/* ── CHECKPOINT ───────────────────────────────────────────────────────────── */
function checkCP(){
  const out=document.getElementById('output').value.trim();
  if(!out) return;
  // Hledat checkpoint nejprve s session prefixem, pak bez
  const prefixedOut=sessionOutput(out);
  fetch('/check_cp?output='+encodeURIComponent(prefixedOut))
    .then(r=>r.json()).then(d=>{
      const box=document.getElementById('resumeBox');
      if(d.exists){
        if(d.phase===1){
          document.getElementById('resumeInfo').textContent=
            `Fáze 1 přerušena: ${d.cats} kat · ${d.articles} hesel · ${d.saved_at}`;
          document.getElementById('resumeTitle').textContent='⏸ Přerušená fáze 1';
        } else {
          document.getElementById('resumeInfo').textContent=
            `Hotovo: ${d.done} / ${d.total} · ${d.saved_at}`;
          document.getElementById('resumeTitle').textContent='⏸ Nedokončený scraping';
        }
        box.classList.add('on');
      } else box.classList.remove('on');
    }).catch(()=>{});
}
function onUrlChange(){
  clearTimeout(window._uct);
  window._uct=setTimeout(()=>{
    checkCP();
    if(_activeProject){
      const idx=_projects.findIndex(p=>p.id===_activeProject);
      if(idx>=0){
        _projects[idx].url=document.getElementById('url').value;
        saveProjects();
      }
    }
  },400);
}

// Auto-save output name to active project
document.addEventListener('DOMContentLoaded',()=>{
  const outEl=document.getElementById('output');
  if(outEl) outEl.addEventListener('change',()=>{
    if(_activeProject){
      const idx=_projects.findIndex(p=>p.id===_activeProject);
      if(idx>=0){ _projects[idx].output=outEl.value; saveProjects(); }
    }
    checkCP();
  });
});

/* ── TABS ─────────────────────────────────────────────────────────────────── */
function tab(id){
  const tabs=['overview','data','viz','map','tags','validate'];
  tabs.forEach(t=>{
    const cap=t.charAt(0).toUpperCase()+t.slice(1);
    const btn=document.getElementById('tab'+cap);
    const pnl=document.getElementById('panel'+cap);
    if(btn) btn.classList.toggle('on',t===id);
    if(pnl) pnl.classList.toggle('on',t===id);
  });
  // Invalidate Leaflet map size when map tab opens
  if(id==='map'){
    if(window._mapInitPending){
      setTimeout(()=>{ window._mapInitPending(); window._mapInitPending=null; }, 50);
    } else if(_map){
      setTimeout(()=>_map.invalidateSize(), 50);
    }
  }
}

/* ── RUN — FÁZE 1 ────────────────────────────────────────────────────────── */
function go(resume){
  const rawUrl=document.getElementById('url').value.trim();
  if(!rawUrl){
    const _urlEl=document.getElementById('url');
    const _runBtn=document.getElementById('btnRun');
    _urlEl.style.borderColor='var(--danger)';
    _runBtn.classList.remove('shake'); void _runBtn.offsetWidth; // reflow → restart animace
    _runBtn.classList.add('shake');
    setTimeout(()=>{ _urlEl.style.borderColor=''; _runBtn.classList.remove('shake'); },600);
    clog('❌ Zadej URL!','err');if(!conOpen)toggleConsole();return;
  }
  // Dekódovat URL aby se zabránilo double-encoding (%25C3 místo %C3)
  let url;
  try { url = decodeURIComponent(rawUrl); } catch(e) { url = rawUrl; }
  allData=[];
  _activeTagFilter=null;
  _validIssueFilter=null;
  _activeTag=null;

  ['stFound','stDone','stEta','stCats'].forEach(id=>document.getElementById(id).textContent='—');
  document.getElementById('stTime').textContent='—';
  document.querySelector('.stats-bar')?.classList.remove('visible');
  document.getElementById('pb1').style.width='0%';
  document.getElementById('pb1').classList.remove('indeterminate');
  document.getElementById('pp1').textContent='—';
  document.getElementById('pb2').style.width='0%'; document.getElementById('pp2').textContent='—';
  document.getElementById('overviewEmpty').classList.remove('hidden');
  document.getElementById('summaryCards').classList.add('hidden');
  document.getElementById('summaryCards').innerHTML='';

  const _rawOutput=document.getElementById('output').value||'wiki_data';
  const p={
    url,
    depth:document.getElementById('depth').value,
    limit:document.getElementById('limit').value,
    delay:document.getElementById('delay').value,
    format:document.getElementById('format').value,
    output:sessionOutput(_rawOutput),
    fields:getFields(),
    workers:getWorkers(),
    api:getApiMode(),
    tags:getAutoTags(),
    wikidata:getWikidata(),
    incremental:getIncremental(),
    s:getSessionId(),
  };
  window._runParams=p;  // uložit pro fázi 2
  window._origUrl=url;  // pro historii
  _completedJobs=[];    // reset pro nové spuštění

  document.getElementById('btnRun').classList.add('hidden');
  document.getElementById('btnStopGroup').classList.add('on');
  document.getElementById('btnStop').classList.remove('hidden');
  document.getElementById('btnPause').classList.remove('hidden');
  document.getElementById('exportBox').classList.remove('on');
  document.getElementById('resumeBox').classList.remove('on');
  document.getElementById('savedReviewBanner').style.display='none';
  document.getElementById('xlsxSt').textContent='';
  setStatus('running'); startTimer();
  setPhase(1,'Sbírám kategorie a URL…','');
  setStep(1);
  initLiveChart();
  addNotif('Scraping zahajen', url.slice(0,60), 'info');
  requestNotifPermission();
  // Otevřít konzoli automaticky
  if(!conOpen) toggleConsole();
  // Spustit indeterminate animaci pro fázi 1
  document.getElementById('pb1').classList.add('indeterminate');
  document.getElementById('pp1').textContent='…';
  tab('overview');

  // Resume přeskočí fázi 1
  if(resume){
    runPhase2Direct(p);
    return;
  }

  // Fáze 1 — jen sbírání
  evtSrc=new EventSource('/run_phase1?'+new URLSearchParams({
    url:p.url, depth:p.depth, limit:p.limit, delay:p.delay, output:p.output, s:p.s||getSessionId()
  }));

  evtSrc.onmessage=e=>{
    const d=JSON.parse(e.data);
    if(d.type==='raw'){
      let t='';
      const m=d.msg;
      if(m.includes('✅')||m.includes('✓')) t='ok';
      else if(m.includes('❌')||m.includes('✗')) t='err';
      else if(m.includes('⚠')||m.includes('⏱')) t='warn';
      else if(m.startsWith('📂')||m.startsWith('💾')||m.startsWith('📋')) t='dim';
      clog(m,t);
    }
    else if(d.type==='phase1_cat'){
      const sub=`${d.cats} kat · ${d.subcats} podkat · ${d.articles} hesel`;
      setPhase(1, d.name, sub);
      document.getElementById('stCats').textContent=d.cats+(d.subcats>0?' +'+d.subcats:'');
      document.getElementById('stFound').textContent=d.articles;
      document.querySelector('.stats-bar')?.classList.add('visible');
      clog(`📂 [${d.cats}] ${d.name}  →  ${d.articles} hesel`, 'dim');
    }
    else if(d.type==='phase1_list'){
      setPhase(1, d.name, `${d.count} hesel`);
      document.getElementById('stFound').textContent=d.count;
      clog(`🔗 ${d.name}  (${d.count})`, 'dim');
    }
    else if(d.type==='ratelimit'){
      document.getElementById('rlVal').textContent=d.new_delay;
      document.getElementById('rlChip').classList.add('on');
    }
    else if(d.type==='phase1_complete'){
      evtSrc.close();
      document.getElementById('pb1').classList.remove('indeterminate');
      document.getElementById('pb1').style.width='100%';
      document.getElementById('pp1').textContent='100%';
      setPhase(1,'Faze 1 hotova — nacitam prehled...','');
      clog('Nacitam pending soubor: '+p.output,'dim');
      fetch('/get_pending?output='+encodeURIComponent(p.output))
        .then(r=>{
          if(!r.ok) throw new Error('HTTP '+r.status+' — pending soubor nenalezen (output='+p.output+')');
          return r.json();
        })
        .then(pending=>{
          const cats=pending.cat_articles||{};
          const catCount=Object.keys(cats).length;
          const total=Object.values(cats).reduce((s,v)=>s+v.length,0);
          clog('Pending: '+catCount+' kategorii, '+total+' hesel','dim');
          if(!catCount){
            clog('Faze 1 nenasla zadne kategorie s hesly. Zkontroluj URL a konzoli.','warn');
            if(!conOpen) toggleConsole();
          }
          // Pokud pending obsahuje cat_tree → ukázat strom, jinak rovnou review
          if(pending.cat_tree && Object.keys(pending.cat_tree).length > 1){
            showCategoryTree(pending, p);
          } else {
            showReview(pending, p);
          }
        })
        .catch(err=>{
          clog('Nelze nacist pending soubor: '+err.message,'err');
          if(!conOpen) toggleConsole();
          setStatus('error');
          document.getElementById('btnRun').classList.remove('hidden');
          document.getElementById('btnStop').classList.add('hidden');
          setPhase(1,'Chyba pri nacteni vysledku faze 1','');
        });
    }
    else if(d.type==='error'){
      stopTimer(); evtSrc.close(); setStatus('error');
      document.getElementById('btnRun').classList.remove('hidden');
      document.getElementById('btnStop').classList.add('hidden');
      setPhase(1,'❌ Chyba', d.msg);
      clog('❌ '+d.msg,'err');
      if(!conOpen)toggleConsole();
    }
  };

  // Timeout — pokud do 5s nepřijde žádná zpráva, něco je špatně
  const _p1timeout=setTimeout(()=>{
    if(!evtSrc||evtSrc.readyState===EventSource.CLOSED) return;
    clog('⚠ Žádná odpověď po 5s — zkontroluj terminál serveru', 'warn');
    if(!conOpen) toggleConsole();
  }, 5000);

  evtSrc.onopen=()=>{
    clearTimeout(_p1timeout);
    clog('🔌 Fáze 1 zahájena', 'dim');
  };

  evtSrc.onerror=(e)=>{
    clearTimeout(_p1timeout);
    const url=evtSrc.url;
    if(evtSrc.readyState===EventSource.CLOSED){
      fetch(url,{method:'HEAD'})
        .then(r=>{
          if(r.status===404)      clog('❌ 404 — route /run_phase1 nenalezena. Máš správnou verzi wiki_gui.py? Zastav server (Ctrl+C), vyměň soubor a spusť znovu.','err');
          else if(r.status===500) clog('❌ 500 — chyba serveru, zkontroluj terminál','err');
          else                    clog(`❌ HTTP ${r.status} — spojení selhalo`,'err');
        })
        .catch(()=>clog('❌ Server neodpovídá — je spuštěný?','err'));
    } else {
      clog('❌ Chyba SSE — zkontroluj terminál','err');
    }
    if(!conOpen) toggleConsole();
    stopTimer(); setStatus('error');
    document.getElementById('btnRun').classList.remove('hidden');
    document.getElementById('btnStop').classList.add('hidden');
    document.getElementById('pb1').classList.remove('indeterminate');
    document.getElementById('pp1').textContent='—';
    setPhase(1,'❌ Chyba při startu','');
  };
}

/* ── REVIEW MODAL — multi-column ─────────────────────────────────────────── */
let _reviewPending=null;
// _cols = [{name, color, cats: {catName: true/false}}]
// Col 0 = hlavní soubor s checkboxy, col 1+ = extra soubory
let _cols=[];
let _colSearch=[];  // search query per column

const COL_COLORS=['#6ee7b7','#60a5fa','#fbbf24','#f87171','#a78bfa','#fb923c','#34d399','#818cf8'];

function showReview(pending, params){
  // Guard: pending muze byt prazdny object nebo obsahovat error z 404
  if(!pending || pending.error || typeof pending.cat_articles === 'undefined'){
    const reason = pending?.error || 'cat_articles chybi v odpovedi';
    clog('Chyba: showReview dostalo neplatna data — '+reason,'err');
    setStatus('error');
    document.getElementById('btnRun').classList.remove('hidden');
    document.getElementById('btnStop').classList.add('hidden');
    if(!conOpen) toggleConsole();
    return;
  }

  _reviewPending=pending;
  const rawCats=pending.cat_articles||{};

  const {result:cats, removed}=dedupCatArticles(rawCats);
  _reviewPending.cat_articles=cats;
  const total=Object.values(cats).reduce((s,v)=>s+v.length,0);

  if(removed>0) clog('Odstraneno '+removed+' duplicitnich hesel pred review','warn');

  _cols=[{
    name: window._runParams?.output || 'výsledky',
    color: COL_COLORS[0],
    cats: Object.fromEntries(Object.keys(cats).map(c=>[c,true]))
  }];
  _colSearch=[''];

  const blCount=_blacklistedCats.size;
  document.getElementById('reviewSub').textContent=
    `${Object.keys(cats).length} kategorií · ${total} hesel`
    +(removed>0?` · −${removed} duplikátů`:'')
    +(blCount>0?` · 🚫 ${blCount} blacklistem`:'');

  // Aplikovat blacklist automaticky
  applyBlacklist();

  document.getElementById('btnStop').classList.add('hidden');
  // Během review: tlačítko Spustit schovat — uživatel používá "Spustit stahování" v dialogu
  document.getElementById('btnRun').classList.add('hidden');
  setStatus('idle');
  stopTimer();
  setPhase(1,'Čeká na výběr kategorií…','');
  setStep(2);

  // Zobrazit reveal animaci
  showReviewReveal(Object.keys(cats).length, total, removed);
  document.getElementById('reviewModal').classList.add('on');
}

let _renderReviewPending=false;
function renderReview(){
  if(_renderReviewPending) return;
  _renderReviewPending=true;
  requestAnimationFrame(()=>{
    _renderReviewPending=false;
    _renderReviewImpl();
  });
}
function _renderReviewImpl(){
  const cats=_reviewPending.cat_articles||{};
  const container=document.getElementById('reviewColumns');

  // Preserve scroll positions per column body before re-render
  const scrollTops=[];
  container.querySelectorAll('.col-body').forEach((b,i)=>{ scrollTops[i]=b.scrollTop; });

  // Preserve focused col-search index and cursor
  let focusedColIdx=null, focusSel=[null,null];
  container.querySelectorAll('.col-search').forEach((inp,i)=>{
    if(document.activeElement===inp){
      focusedColIdx=i;
      try{ focusSel=[inp.selectionStart,inp.selectionEnd]; }catch(e){}
    }
  });

  const html=_cols.map((col,ci)=>{
    const q=(_colSearch[ci]||'').toLowerCase();
    // Get cats for this column
    const entries=Object.entries(col.cats)
      .filter(([name,on])=>on)
      .filter(([name])=>!q||name.toLowerCase().includes(q))
      .sort(([a],[b])=>a.localeCompare(b,'cs'));

    const totalHesel=Object.entries(col.cats)
      .filter(([,on])=>on)
      .reduce((s,[n])=>s+(cats[n]?.length||0),0);

    const isMain=ci===0;
    const canLeft=ci>0;
    const canRight=ci<_cols.length-1;

    // For main col also show unchecked (discarded)
    const discarded=isMain ? Object.entries(col.cats)
      .filter(([,on])=>!on)
      .filter(([name])=>!q||name.toLowerCase().includes(q))
      .sort(([a],[b])=>a.localeCompare(b,'cs')) : [];

    const rowsHtml=[...entries,...discarded].map(([name,on])=>{
      const count=cats[name]?.length||0;
      const rowClass=isMain?(on?'checked':'unchecked'):'checked';
      const safeIdx=encodeURIComponent(name);
      const arrowLeft=canLeft?`<div class="cat-arr" data-action="move" data-name="${safeIdx}" data-from="${ci}" data-to="${ci-1}" title="Přesunout do sloupce vlevo">←</div>`:'';
      const arrowRight=canRight
        ?`<div class="cat-arr" data-action="move" data-name="${safeIdx}" data-from="${ci}" data-to="${ci+1}" title="Přesunout do sloupce vpravo">→</div>`
        :`<div class="cat-arr" data-action="tonew" data-name="${safeIdx}" data-from="${ci}" title="Přesunout do nového sloupce">→+</div>`;
      const isBlacklisted=isMain&&!on&&_blacklistedCats.has(name);
      const checkbox=isMain?`<div class="cat-check-cell" data-action="toggle" data-name="${safeIdx}" data-ci="${ci}"><div class="cat-cb">${on?'✓':''}</div></div>`:'';
      return `<div class="cat-row ${rowClass}">
        ${checkbox}
        <div class="cat-label" title="${esc(name)}${isMain&&!on&&_blacklistedCats.has(name)?' — vyřazeno blacklistem':''}"><span>${esc(name)}</span>${isMain&&!on&&_blacklistedCats.has(name)?'<span style="font-family:var(--mono);font-size:9px;color:var(--warn);margin-left:5px;opacity:.8">🚫</span>':''}</div>
        <div class="cat-badge">${count}</div>
        <div class="cat-arrows">
          ${arrowLeft}
          ${arrowRight}
        </div>
      </div>`;
    }).join('');

    const deleteBtn=!isMain?`<div class="col-delete" onclick="deleteCol(${ci})" title="Smazat sloupec">×</div>`:'';
    return `<div class="review-col collapsible-section" id="col_${ci}" data-ci="${ci}">
      <div class="col-header" style="border-top:3px solid ${col.color}">
        <input class="col-name-input" value="${esc(col.name)}" placeholder="Název souboru"
          oninput="renameCol(${ci},this.value)" onclick="event.stopPropagation()">
        <span class="col-stats">${Object.values(col.cats).filter(Boolean).length} kat · ${totalHesel} hesel</span>
        ${deleteBtn}
      </div>
      <div class="sec-body" style="display:flex;flex-direction:column;flex:1;overflow:hidden">
      <div class="col-toolbar">
        <input class="col-search" placeholder="Hledat…" value="${esc(_colSearch[ci]||'')}"
          oninput="searchCol(${ci},this.value)"
          ondragstart="event.stopPropagation();event.preventDefault()"
          onmousedown="event.stopPropagation()">
        ${isMain?`<button class="col-toolbar-btn" onclick="selectAllInCol(${ci},true)" title="Zaškrtnout všechny kategorie v tomto sloupci">✓ vše</button>
          <button class="col-toolbar-btn" onclick="selectAllInCol(${ci},false)" title="Odškrtnout všechny kategorie v tomto sloupci">✗ vše</button>`:''}
      </div>
      <div class="col-body">${rowsHtml||'<div style="padding:20px;text-align:center;font-family:var(--mono);font-size:11px;color:var(--text3)">Prázdný sloupec</div>'}</div>
      </div>
    </div>`;
  }).join('');

  container.innerHTML=html+`
    <div class="add-col-btn" onclick="addCol()" title="Přidat nový výstupní soubor — kategorie lze přetáhnout do nového sloupce">
      <div class="add-col-icon">+</div>
      <div class="add-col-label">Nový soubor</div>
    </div>`;

  // Event delegation — handles all cat-row interactions via data attributes
  container.querySelectorAll('[data-action]').forEach(el=>{
    el.addEventListener('click', e=>{
      e.stopPropagation();
      const action=el.dataset.action;
      const name=decodeURIComponent(el.dataset.name||'');
      const ci=parseInt(el.dataset.ci||el.dataset.from||'0');
      const to=parseInt(el.dataset.to||'0');
      if(action==='toggle'){
        pushReviewUndo(_cols[ci].cats[name]?`Odškrtnuto: ${name}`:`Zaškrtnuto: ${name}`);
        _cols[ci].cats[name]=!_cols[ci].cats[name];
        renderReview();
      } else if(action==='move'){
        const fromCi=parseInt(el.dataset.from);
        if(to<0||to>=_cols.length) return;
        pushReviewUndo(`Přesunuto: ${name} → ${_cols[to].name}`);
        if(fromCi===0) _cols[0].cats[name]=false;
        else delete _cols[fromCi].cats[name];
        _cols[to].cats[name]=true;
        renderReview();
      } else if(action==='tonew'){
        const fromCi=parseInt(el.dataset.from);
        pushReviewUndo(`Nový sloupec pro: ${name}`);
        addCol();
        const newCi=_cols.length-1;
        if(fromCi===0) _cols[0].cats[name]=false;
        else delete _cols[fromCi].cats[name];
        _cols[newCi].cats[name]=true;
        renderReview();
      }
    });
  });

  // Drag & drop přeřazení sloupců
  let _dragCiActive=null;
  container.querySelectorAll('.review-col[draggable]').forEach(colEl=>{
    colEl.addEventListener('dragstart',e=>{
      _dragCiActive=parseInt(colEl.dataset.ci);
      e.dataTransfer.effectAllowed='move';
      setTimeout(()=>colEl.style.opacity='0.45',0);
    });
    colEl.addEventListener('dragend',()=>{
      colEl.style.opacity='';
      container.querySelectorAll('.review-col').forEach(c=>c.classList.remove('drag-over'));
    });
    colEl.addEventListener('dragover',e=>{
      e.preventDefault(); e.dataTransfer.dropEffect='move';
      container.querySelectorAll('.review-col').forEach(c=>c.classList.remove('drag-over'));
      colEl.classList.add('drag-over');
    });
    colEl.addEventListener('dragleave',()=>colEl.classList.remove('drag-over'));
    colEl.addEventListener('drop',e=>{
      e.preventDefault(); colEl.classList.remove('drag-over'); colEl.style.opacity='';
      const toCi=parseInt(colEl.dataset.ci);
      if(_dragCiActive===null||_dragCiActive===toCi){_dragCiActive=null;return;}
      pushReviewUndo(`Přehozeny sloupce: ${_cols[_dragCiActive].name} ↔ ${_cols[toCi].name}`);
      [_cols[_dragCiActive],_cols[toCi]]=[_cols[toCi],_cols[_dragCiActive]];
      [_colSearch[_dragCiActive],_colSearch[toCi]]=[_colSearch[toCi],_colSearch[_dragCiActive]];
      _dragCiActive=null;
      renderReview();
    });
  });

  // Split pane mezi sloupci
  addColResizers(container);

  // Draggable sloupce — aktivovat jen při tahu za header (ne za input)
  container.querySelectorAll('.review-col').forEach(col=>{
    col.addEventListener('mousedown', e=>{
      // Draggable jen pokud mousedown na header, ale NE na input nebo button
      const header=e.target.closest('.col-header');
      const isInput=e.target.tagName==='INPUT'||e.target.tagName==='BUTTON'||e.target.closest('.col-delete');
      if(header && !isInput){
        col.setAttribute('draggable','true');
      } else {
        col.removeAttribute('draggable');
      }
    });
    col.addEventListener('dragend', ()=>col.removeAttribute('draggable'));
  });

  // Context preview on cat-label hover
  container.querySelectorAll('.cat-label').forEach(el=>{
    const row=el.closest('.cat-row');
    const actionEl=row?.querySelector('[data-name]');
    const name=actionEl?decodeURIComponent(actionEl.dataset.name||''):'';
    if(!name) return;
    el.style.cursor='default';
    el.addEventListener('mouseenter',e=>catPreviewShow(name,e.clientX,e.clientY));
    el.addEventListener('mousemove',e=>{
      const p=_previewEl();
      if(p.classList.contains('on')){p.style.left=(e.clientX+16)+'px';p.style.top=(e.clientY-10)+'px';}
    });
    el.addEventListener('mouseleave',()=>catPreviewHide());
  });

  // Restore scroll positions
  container.querySelectorAll('.col-body').forEach((b,i)=>{
    if(scrollTops[i]!=null) b.scrollTop=scrollTops[i];
  });

  // Restore focus on col-search
  if(focusedColIdx!=null){
    const inputs=container.querySelectorAll('.col-search');
    const inp=inputs[focusedColIdx];
    if(inp){
      inp.focus();
      try{ inp.setSelectionRange(focusSel[0],focusSel[1]); }catch(e){}
    }
  }

  updateSummary();
}

// Keep these as wrappers for keyboard nav compatibility
function toggleCat(e,name,ci){
  e.stopPropagation();
  _cols[ci].cats[name]=!_cols[ci].cats[name];
  renderReview();
}

function moveCat(e,name,fromCi,toCi){
  e.stopPropagation();
  if(toCi<0||toCi>=_cols.length) return;
  if(fromCi===0) _cols[0].cats[name]=false;
  else delete _cols[fromCi].cats[name];
  _cols[toCi].cats[name]=true;
  renderReview();
}




function addCol(){
  const ci=_cols.length;
  _cols.push({
    name:`soubor_${ci}`,
    color:COL_COLORS[ci%COL_COLORS.length],
    cats:{}
  });
  _colSearch.push('');
  renderReview();
}


function toggleColCollapse(e, ci){
  if(e.target.tagName==='INPUT'||e.target.tagName==='BUTTON'||e.target.closest('.col-delete')) return;
  const col=document.getElementById('col_'+ci);
  if(col) col.classList.toggle('collapsed');
}
function deleteCol(ci){
  if(ci===0) return;
  // Přesunout zpátky do hlavního sloupce jako vybrané
  Object.keys(_cols[ci].cats).forEach(name=>{
    _cols[0].cats[name]=true;
  });
  _cols.splice(ci,1);
  _colSearch.splice(ci,1);
  renderReview();
}

function renameCol(ci,val){
  _cols[ci].name=val;
  updateSummary();
  saveReviewState();
}

function searchCol(ci,val){
  _colSearch[ci]=val;
  renderReview();
}

function selectAllInCol(ci,val){
  Object.keys(_cols[ci].cats).forEach(k=>{ _cols[ci].cats[k]=val; });
  renderReview();
}

function updateSummary(){
  const cats=_reviewPending.cat_articles||{};
  const parts=_cols.map((col,ci)=>{
    const n=Object.values(col.cats).filter(Boolean).length;
    const h=Object.entries(col.cats).filter(([,on])=>on).reduce((s,[c])=>s+(cats[c]?.length||0),0);
    return `<div class="rs-item"><div class="rs-dot" style="background:${col.color}"></div><b>${col.name}</b>: ${n} kat · ${h} hesel</div>`;
  });
  // Přidat info o blacklistovaných
  const blCount=_blacklistedCats.size;
  const blInfo=blCount>0
    ?`<div class="rs-item" style="color:var(--warn);margin-top:4px">🚫 ${blCount} kategorií vyřazeno blacklistem</div>`
    :'';
  document.getElementById('reviewSummary').innerHTML=parts.join('')+blInfo;
  const anySelected=_cols.some(col=>Object.values(col.cats).some(Boolean));
  document.getElementById('btnStartScraping').disabled=!anySelected;
  updateReviewSizeEst();
}

function closeReview(){
  if(_reviewPending) saveReviewState();
  _focusedCatName=null;
  _reviewUndoStack=[]; _reviewRedoStack=[]; _updateReviewUndoUI();
  document.querySelectorAll('.cat-row').forEach(r=>r.style.outline='');
  document.getElementById('reviewModal').classList.remove('on');
  // Ukázat btnRun jen pokud neprobíhá scraping
  if(!_phase2Running){
    document.getElementById('btnRun').classList.remove('hidden');
  }
  checkSavedReview();
}

function selectAllCats(val){
  // Legacy — still works on col 0
  Object.keys(_cols[0].cats).forEach(c=>{ _cols[0].cats[c]=val; });
  renderReview();
}

function startPhase2(){
  _phase2Running=true;  // nastavit PŘED closeReview aby se btnRun neschoval
  document.getElementById("savedReviewBanner").style.display="none";  // skrýt banner při spuštění
  closeReview();
  clearReviewState();
  const p=window._runParams;
  const cats=_reviewPending.cat_articles||{};

  // Build jobs: [{output, urls}]
  const jobs=_cols.map(col=>{
    const selected=Object.entries(col.cats).filter(([,on])=>on).map(([c])=>c);
    const urls=[...new Set(selected.flatMap(c=>cats[c]||[]))];
    return {output:col.name, urls};
  }).filter(j=>j.urls.length>0);

  if(!jobs.length){ clog('⚠ Žádné kategorie vybrány','warn'); return; }

  clog(`▶ Stahování: ${jobs.length} souborů`,'hl');
  jobs.forEach(j=>clog(`  📁 ${j.output}: ${j.urls.length} hesel`,'dim'));

  // Spustit jobs sekvenčně
  runJobsSequentially(jobs, p, 0);
}

function runJobsSequentially(jobs, p, idx){
  if(idx>=jobs.length){
    clog('✅ Všechny soubory hotové','ok');
    return;
  }
  const job=jobs[idx];
  clog(`\n▶ Soubor ${idx+1}/${jobs.length}: ${job.output} (${job.urls.length} hesel)`,'hl');
  // Override output name for this job
  const jobParams={...p, output:job.output};
  window._runParams=jobParams;

  // Uložit URL a spustit
  fetch('/save_urls',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({output:job.output, urls:job.urls})
  })
  .then(r=>r.json())
  .then(()=>{
    const qp=new URLSearchParams({
      output:job.output,
      format:p.format,
      delay:p.delay,
      fields:p.fields,
      workers:p.workers||1,
      api:p.api?'1':'0',
      tags:p.tags?'1':'0',
      wikidata:p.wikidata?'1':'0',
      incremental:p.incremental?'1':'0',
      s:p.s||getSessionId(),
    });
    evtSrc=new EventSource('/run_phase2_sse?'+qp);
    // Po dokončení spustit další job
    const originalDone=()=>{
      stopTimer(); evtSrc.close(); 
      document.getElementById('rlChip').classList.remove('on');
      document.getElementById('pb2').style.width='100%';
      document.getElementById('pp2').textContent='100%';
      if(idx===jobs.length-1){
        // Poslední job — uložit historii, poslat notifikaci
        const duration=t0?Math.round((Date.now()-t0)/1000):0;
        const totalHesel=jobs.reduce((s,j)=>s+j.urls.length,0);
        saveToHistory({
          url: window._origUrl||p.url||'',
          name: jobs.length===1 ? job.output : jobs.map(j=>j.output).join(', '),
          output: job.output,
          hesel: totalHesel,
          duration_s: duration,
          date: new Date().toLocaleDateString('cs-CZ')
        });
        sendNotification('WikiScraper — hotovo! 📚',
          `${totalHesel} hesel staženo za ${fmtDuration(duration)}`);
        // 🎉 Konfety!
        if(typeof confetti!=='undefined'){
          confetti({particleCount:130,spread:80,origin:{y:0.6}});
          setTimeout(()=>confetti({particleCount:70,spread:55,origin:{y:0.4},
            colors:['#6ee7b7','#60a5fa','#fbbf24','#f87171','#a78bfa']}),700);
        }
        updateSizeEstimate(totalHesel);
        setStatus('done'); setStep(4);
        _phase2Running=false;
        _currentFile=job.output;
        document.getElementById('renameWrap').style.display='';
        document.getElementById('btnRename').style.display='';
        document.getElementById('btnRun').classList.remove('hidden');
        document.getElementById('btnStopGroup').classList.remove('on');
        document.getElementById('btnStop').classList.add('hidden');
        document.getElementById('btnPause').classList.add('hidden');
        setPhase(2,'✅ Vše hotovo!','');
        // Zobrazit export box se VŠEMI joby
        showExportBox(jobs.map(j=>j.output));
        // Načíst sloučená data ze všech jobů do tabulky
        loadDataMultiple(jobs.map(j=>j.output));
      } else {
        // Další job
        runJobsSequentially(jobs, p, idx+1);
      }
    };
    attachPhase2Listeners(job.output, originalDone);
  })
  .catch(e=>{
    setStatus('error');
    clog('❌ '+e,'err');
    document.getElementById('btnRun').classList.remove('hidden');
    document.getElementById('btnStop').classList.add('hidden');
  });
}

function runPhase2Direct(p){
  // Resume — přeskočit fázi 1, použít checkpoint fáze 2
  setStatus('running'); startTimer();
  setPhase(2,'Obnovuji z checkpointu…','');
  setStep(3);
  document.getElementById('btnRun').classList.add('hidden');
  document.getElementById('btnStop').classList.remove('hidden');
  const params={...p, resume:'1', s:p.s||getSessionId()};
  evtSrc=new EventSource('/run?'+new URLSearchParams(params));
  attachPhase2Listeners(p.output);
}

function runPhase2WithUrls(urls, p){
  setStatus('running');
  startTimer();
  setPhase(2,'Připravuji stahování…',`${urls.length} hesel`);
  setStep(3);
  _phase2Running=true;
  document.getElementById('btnRun').classList.add('hidden');
  document.getElementById('btnStop').classList.remove('hidden');

  // Nejdříve POST uloží URL file, pak GET stream spustí fázi 2
  fetch('/save_urls',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({output:p.output, urls:urls})
  })
  .then(r=>r.json())
  .then(()=>{
    const qp=new URLSearchParams({
      output:p.output,
      format:p.format,
      delay:p.delay,
      fields:p.fields
    });
    evtSrc=new EventSource('/run_phase2_sse?'+qp);
    attachPhase2Listeners(p.output);
  })
  .catch(e=>{
    setStatus('error');
    clog('❌ '+e,'err');
    document.getElementById('btnRun').classList.remove('hidden');
    document.getElementById('btnStop').classList.add('hidden');
  });
}

function attachPhase2Listeners(output, onDone){
  evtSrc.onmessage=e=>{
    const d=JSON.parse(e.data);
    if(d.type==='raw'){
      let t='';
      const m=d.msg;
      if(m.includes('✅')||m.includes('✓')) t='ok';
      else if(m.includes('❌')||m.includes('✗')) t='err';
      else if(m.includes('⚠')) t='warn';
      else if(m.match(/\[\d+\/\d+\]/)) t='info';
      else if(m.startsWith('💾')) t='dim';
      clog(m,t);
    }
    else if(d.type==='phase1_done'){
      document.getElementById('stFound').textContent=d.total;
      document.getElementById('stEta').textContent=eta(d.eta);
    }
    else if(d.type==='progress'){
      const pct=d.total>0?Math.round(d.i/d.total*100):0;
      document.getElementById('pb2').style.width=pct+'%'; document.getElementById('pp2').textContent=pct+'%';
      document.getElementById('stFound').textContent=d.total;
      document.getElementById('stDone').textContent=d.i;
      document.getElementById('stEta').textContent=eta(d.eta);
      setPhase(2, d.name, `${d.i} / ${d.total} · ${eta(d.eta)}`);
      _done_count_global=d.i;
      updateLiveChart(d.i);
      setTitleProgress(pct);
      document.querySelector('.stats-bar')?.classList.add('visible');
    }
    else if(d.type==='ratelimit'){
      document.getElementById('rlVal').textContent=d.new_delay;
      document.getElementById('rlChip').classList.add('on');
    }
    else if(d.type==='delay_ok'){
      document.getElementById('rlVal').textContent=d.delay;
      if(parseFloat(d.delay)<=parseFloat(document.getElementById('delay').value)+0.1)
        document.getElementById('rlChip').classList.remove('on');
    }
    else if(d.type==='done'){
      setTitleProgress(null);
      if(onDone){ onDone(); return; }
      stopTimer(); evtSrc.close(); setStatus('done');
      document.getElementById('btnRun').classList.remove('hidden');
      document.getElementById('btnStop').classList.add('hidden');
      document.getElementById('rlChip').classList.remove('on');
      document.getElementById('pb2').style.width='100%'; document.getElementById('pp2').textContent='100%';
      setPhase(2,'✅ Hotovo!','');
      if(d.has_files){
        document.getElementById('exportBox').classList.add('on');
        loadData(output);
      }
    }
    else if(d.type==='error'){
      setTitleProgress(null);
      stopTimer(); evtSrc.close(); setStatus('error');
      document.getElementById('btnRun').classList.remove('hidden');
      document.getElementById('btnStop').classList.add('hidden');
      setPhase(2,'❌ Chyba', d.msg);
      clog('❌ '+d.msg,'err');
      if(!conOpen)toggleConsole();
    }
  };
  evtSrc.onerror=()=>{
    if(evtSrc.readyState===EventSource.CLOSED) return;
    stopTimer(); setStatus('error');
    document.getElementById('btnRun').classList.remove('hidden');
    document.getElementById('btnStop').classList.add('hidden');
  };
}

function discardCP(){
  const output=sessionOutput(document.getElementById('output').value||'wiki_data');
  fetch('/discard_checkpoint',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({output})})
    .then(()=>{
      document.getElementById('resumeBox').classList.remove('on');
      clog('🗑 Checkpoint zahozen','warn');
    });
}

function stop(){
  fetch('/stop',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({s:getSessionId()})});
  if(evtSrc)evtSrc.close();
  stopTimer(); setStatus('idle'); setStep(0);
  _phase2Running=false;
  document.getElementById('btnRun').classList.remove('hidden');
  document.getElementById('btnStopGroup').classList.remove('on');
  document.getElementById('btnStop').classList.add('hidden');
  document.getElementById('btnPause').classList.add('hidden');
  document.getElementById('pb1').classList.remove('indeterminate');
  document.getElementById('pp1').textContent='—';
  setPhase(1,'Zastaveno','');
  checkCP();
  checkSavedReview();  // zobrazit banner pokud existuje uložený review
}





/* ── MOBILE ──────────────────────────────────────────────────────────────── */
const _isMobile = () => window.innerWidth <= 768;

function mobToggleSidebar(){
  const sidebar = document.querySelector('aside.sidebar');
  const backdrop = document.getElementById('mobSidebarBackdrop');
  const navBtn = document.getElementById('mobNavSidebar');
  const isOpen = sidebar.classList.contains('mob-open');
  if(isOpen){
    sidebar.classList.remove('mob-open');
    backdrop.classList.remove('on');
    navBtn.classList.remove('active');
  } else {
    sidebar.classList.add('mob-open');
    backdrop.classList.add('on');
    navBtn.classList.add('active');
  }
}

function mobCloseSidebar(){
  document.querySelector('aside.sidebar')?.classList.remove('mob-open');
  document.getElementById('mobSidebarBackdrop')?.classList.remove('on');
  document.getElementById('mobNavSidebar')?.classList.remove('active');
}

function mobTab(name){
  // Přepnout tab a aktualizovat spodní nav
  tab(name);
  mobCloseSidebar();
  // Zvýraznit aktivní nav tlačítko
  document.querySelectorAll('.mob-nav-btn').forEach(b => b.classList.remove('active'));
  const map = {overview:'mobNavHome', data:'mobNavData', viz:'mobNavViz'};
  if(map[name]) document.getElementById(map[name])?.classList.add('active');
}

// Inicializace mobile layoutu
function initMobile(){
  // Zobrazit/skrýt mobilní nav podle šířky
  function applyMobileLayout(){
    const mob = _isMobile();
    const nav = document.getElementById('mobileNav');
    if(nav) nav.style.display = mob ? 'flex' : 'none';
    if(!mob) mobCloseSidebar();
  }
  applyMobileLayout();
  window.addEventListener('resize', applyMobileLayout, {passive:true});
  if(!_isMobile()) return;

  // Zavřít sidebar swipe dolů
  let touchStartY = 0;
  const sidebar = document.querySelector('aside.sidebar');
  sidebar?.addEventListener('touchstart', e => { touchStartY = e.touches[0].clientY; }, {passive:true});
  sidebar?.addEventListener('touchmove', e => {
    if(!sidebar.classList.contains('mob-open')) return;
    const dy = e.touches[0].clientY - touchStartY;
    if(dy > 60) mobCloseSidebar();
  }, {passive:true});

  // Swipe ze spodního okraje → otevřít sidebar
  document.addEventListener('touchstart', e => {
    if(e.touches[0].clientY > window.innerHeight - 80 && !sidebar?.classList.contains('mob-open')){
      touchStartY = e.touches[0].clientY;
    }
  }, {passive:true});

  // Aktualizovat spodní nav při přepnutí taby přes desktop tlačítka
  const observer = new MutationObserver(() => {
    const activeTab = document.querySelector('.tab-btn.on');
    if(!activeTab) return;
    document.querySelectorAll('.mob-nav-btn').forEach(b => b.classList.remove('active'));
    const id = activeTab.id;
    const navMap = {tabOverview:'mobNavHome', tabData:'mobNavData', tabViz:'mobNavViz'};
    if(navMap[id]) document.getElementById(navMap[id])?.classList.add('active');
  });
  const tabsEl = document.querySelector('.tabs');
  if(tabsEl) observer.observe(tabsEl, {subtree:true, attributeFilter:['class']});
}

// Reagovat na změnu orientace / resize
window.addEventListener('resize', () => {
  if(!_isMobile()) mobCloseSidebar();
}, {passive:true});

/* ── CATEGORY TREE ───────────────────────────────────────────────────────── */
let _treeData = null;      // pending data z fáze 1 (obsahuje cat_tree)
let _treeParams = null;    // run params
let _treeChecked = {};     // {cat_name: true/false}
let _treeExpanded = {};    // {cat_name: true/false}

function showCategoryTree(pending, params){
  _treeData = pending;
  _treeParams = params;
  const tree = pending.cat_tree || {};
  const root = pending.root_cat || Object.keys(tree).find(k=>!tree[k]?.parent) || '';
  const cats = pending.cat_articles || {};

  if(!root || Object.keys(tree).length <= 1){
    // Žádný strom — přejít rovnou na review
    showReview(pending, params);
    return;
  }

  // Výchozí stav: vše zaškrtnuto, root rozbalený, zbytek sbalený
  _treeChecked = {};
  _treeExpanded = {};
  Object.keys(tree).forEach(n=>{
    _treeChecked[n] = true;
    _treeExpanded[n] = false;
  });
  if(root) _treeExpanded[root] = true;

  const totalCats = Object.keys(cats).length;
  const totalHesel = Object.values(cats).reduce((s,v)=>s+v.length, 0);
  document.getElementById('treeSub').textContent =
    `${totalCats} kategorií · ${totalHesel} hesel`;

  document.getElementById('treeModal').classList.add('on');
  renderTree();
}

function _treeGetArticleCount(catName, tree, cats, visited=new Set()){
  // Rekurzivní součet hesel celé větve
  if(visited.has(catName)) return 0;
  visited.add(catName);
  let count = (cats[catName]||[]).length;
  (tree[catName]?.children||[]).forEach(child=>{
    count += _treeGetArticleCount(child, tree, cats, visited);
  });
  return count;
}

function renderTree(){
  const tree = _treeData.cat_tree || {};
  const cats = _treeData.cat_articles || {};
  const root = _treeData.root_cat || Object.keys(tree).find(k=>!tree[k]?.parent) || '';
  const body = document.getElementById('treeBody');

  function renderNode(name, depth){
    const node = tree[name] || {children:[]};
    const children = node.children || [];
    const hasChildren = children.length > 0;
    const expanded = _treeExpanded[name];
    const checked = _treeChecked[name];
    const count = _treeGetArticleCount(name, tree, cats);
    const ownCount = (cats[name]||[]).length;

    // Zjistit zda jsou děti částečně zaškrtnuty
    const anyChildChecked = hasChildren && children.some(c=>_treeChecked[c]);
    const allChildrenChecked = hasChildren && children.every(c=>_treeChecked[c]);
    const cbClass = checked ? 'checked' : (anyChildChecked ? 'partial' : '');

    const indent = depth * 18;
    let html = `<div class="tree-node${checked?' selected':''}" data-name="${encodeURIComponent(name)}" style="padding-left:${indent}px">
      <div class="tree-toggle${hasChildren?'':' leaf'}" data-toggle="${encodeURIComponent(name)}">
        ${hasChildren ? (expanded?'▾':'▸') : ''}
      </div>
      <div class="tree-cb-wrap" data-check="${encodeURIComponent(name)}">
        <div class="tree-cb ${cbClass}">${checked?'✓':''}</div>
      </div>
      <div class="tree-label" data-label="${encodeURIComponent(name)}">${esc(name)}</div>
      <div class="tree-count">${count>0?count:'—'}</div>
    </div>`;

    if(hasChildren && expanded){
      children.forEach(child=>{ html += renderNode(child, depth+1); });
    }
    return html;
  }

  body.innerHTML = root ? renderNode(root, 0) : '<div style="padding:20px;color:var(--text3);font-family:var(--mono);font-size:11px">Strom kategorií není k dispozici</div>';

  // Event delegation
  body.querySelectorAll('[data-toggle]').forEach(el=>{
    el.addEventListener('click', e=>{
      const name = decodeURIComponent(el.dataset.toggle);
      if(!(tree[name]?.children?.length)) return;
      _treeExpanded[name] = !_treeExpanded[name];
      renderTree();
    });
  });

  body.querySelectorAll('[data-check]').forEach(el=>{
    el.addEventListener('click', e=>{
      const name = decodeURIComponent(el.dataset.check);
      _treeSetChecked(name, !_treeChecked[name], tree);
      renderTree();
      updateTreeSummary();
    });
  });

  body.querySelectorAll('[data-label]').forEach(el=>{
    el.addEventListener('click', e=>{
      const name = decodeURIComponent(el.dataset.label);
      if(tree[name]?.children?.length){
        _treeExpanded[name] = !_treeExpanded[name];
        renderTree();
      }
    });
  });

  updateTreeSummary();
}

function _treeSetChecked(name, val, tree){
  _treeChecked[name] = val;
  // Rekurzivně nastavit děti
  (tree[name]?.children||[]).forEach(child=>{
    _treeSetChecked(child, val, tree);
  });
}

function updateTreeSummary(){
  const tree = _treeData.cat_tree || {};
  const cats = _treeData.cat_articles || {};
  // Spočítat kolik větví je zaškrtnuto (top-level pod rootem)
  const root = _treeData.root_cat || '';
  const topLevel = root ? (tree[root]?.children||[]) : Object.keys(tree).filter(k=>!tree[k]?.parent);
  const checkedBranches = topLevel.filter(n=>_treeChecked[n]);
  const totalHesel = checkedBranches.reduce((s,n)=>s+_treeGetArticleCount(n,tree,cats), 0);

  document.getElementById('treeSummary').textContent =
    `${checkedBranches.length} větví vybráno · ${totalHesel} hesel`;
  document.getElementById('btnTreeConfirm').disabled = checkedBranches.length === 0;
}

function treeExpandAll(){
  Object.keys(_treeData.cat_tree||{}).forEach(n=>{ _treeExpanded[n]=true; });
  renderTree();
}

function treeCollapseAll(){
  const root = _treeData.root_cat||'';
  Object.keys(_treeData.cat_tree||{}).forEach(n=>{ _treeExpanded[n]=false; });
  if(root) _treeExpanded[root]=true;
  renderTree();
}

function treeCheckAll(val){
  Object.keys(_treeChecked).forEach(n=>{ _treeChecked[n]=val; });
  renderTree();
  updateTreeSummary();
}

function treeSkip(){
  // Přeskočit strom — klasický review se vším
  document.getElementById('treeModal').classList.remove('on');
  showReview(_treeData, _treeParams);
}

function treeConfirm(){
  // Postavit _cols z vybraných větví
  document.getElementById('treeModal').classList.remove('on');
  const tree = _treeData.cat_tree || {};
  const cats = _treeData.cat_articles || {};
  const root = _treeData.root_cat || '';
  const topLevel = root ? (tree[root]?.children||[]) : Object.keys(tree).filter(k=>!tree[k]?.parent);
  const checkedBranches = topLevel.filter(n=>_treeChecked[n]);

  if(!checkedBranches.length){
    treeSkip(); return;
  }

  // Každá zaškrtnutá top-level větev = jeden sloupec
  // Sesbírat všechny URL celé větve rekurzivně
  function getBranchCats(name, visited=new Set()){
    if(visited.has(name)) return [];
    visited.add(name);
    const result = [name];
    (tree[name]?.children||[]).filter(c=>_treeChecked[c]).forEach(child=>{
      result.push(...getBranchCats(child, visited));
    });
    return result;
  }

  // Nastavit _cols a _reviewPending
  _reviewPending = _treeData;
  const rawCats = cats;
  const {result:dedupCats, removed} = dedupCatArticles(rawCats);
  _reviewPending.cat_articles = dedupCats;

  _cols = checkedBranches.map((branch, i)=>{
    const branchCats = getBranchCats(branch);
    const colCats = {};
    branchCats.forEach(c=>{ if(dedupCats[c]) colCats[c]=true; });
    return {
      name: branch,
      color: COL_COLORS[i % COL_COLORS.length],
      cats: colCats
    };
  });
  _colSearch = _cols.map(()=>'');

  // Otevřít review rovnou na sloupce (přeskočit reveal animaci)
  applyBlacklist();
  document.getElementById('btnStop').classList.add('hidden');
  document.getElementById('btnRun').classList.add('hidden');
  setStatus('idle'); stopTimer();
  setPhase(1,'Čeká na výběr kategorií…','');
  setStep(2);

  document.getElementById('reviewSub').textContent =
    `${checkedBranches.length} větví · ${Object.values(dedupCats).reduce((s,v)=>s+v.length,0)} hesel`;

  document.getElementById('reviewReveal').classList.remove('on');
  document.getElementById('reviewColumns').style.display='';
  document.getElementById('reviewFooter').style.display='';
  document.getElementById('reviewModal').classList.add('on');
  renderReview();
  updateSummary();
}


/* ── IMAGES & TABLES VIEWER ──────────────────────────────────────────────── */
function openLightbox(src, caption, meta){
  document.getElementById('lightboxImg').src = src;
  document.getElementById('lightboxCap').textContent = caption || '';
  document.getElementById('lightboxMeta').textContent = meta || '';
  document.getElementById('imgLightbox').classList.add('on');
}
function closeLightbox(){
  document.getElementById('imgLightbox').classList.remove('on');
}
// Klik na thumbnail → lightbox plné verze
document.addEventListener('click', e=>{
  const thumb = e.target.closest('.img-thumb');
  if(!thumb) return;
  e.stopPropagation();
  const row = thumb.closest('tr');
  if(!row) return;
  const idx = parseInt(row.dataset.idx);
  const r = _vtRows[idx];
  if(!r||!r.images) return;
  // Najít správný obrázek podle thumb src
  const src = thumb.src;
  const img = r.images.find(i=>(i.thumb||i.url)===src||i.url===src||i.thumb===src) || r.images[0];
  if(!img) return;
  const meta = [img.author, img.licence].filter(Boolean).join(' · ');
  openLightbox(img.url, img.caption, meta);
});
// Esc zavírá lightbox
document.addEventListener('keydown', e=>{
  if(e.key==='Escape') closeLightbox();
});

let _tblCurrentRows = null;
function showTables(e, rowIdx){
  e.stopPropagation();
  const r = _vtRows[rowIdx];
  if(!r||!r.tables||!r.tables.length) return;
  _tblCurrentRows = r.tables;

  const popup = document.getElementById('tblPopup');
  const nav   = document.getElementById('tblNav');
  const cont  = document.getElementById('tblContent');

  // Nav tlačítka pokud více tabulek
  if(r.tables.length > 1){
    nav.innerHTML = r.tables.map((t,i)=>
      `<button class="btn-sm${i===0?' active':''}" onclick="renderTableTab(${i})" id="tblTab${i}">${t.caption||('Tabulka '+(i+1))}</button>`
    ).join('');
    nav.style.display = 'flex';
  } else {
    nav.innerHTML = '';
    nav.style.display = 'none';
  }

  renderTableTab(0);

  // Umístit popup u tlačítka
  const rect = e.target.getBoundingClientRect();
  popup.style.left = Math.min(rect.left, window.innerWidth-690)+'px';
  popup.style.top  = Math.min(rect.bottom+4, window.innerHeight-400)+'px';
  popup.classList.add('on');
}

function renderTableTab(i){
  if(!_tblCurrentRows||!_tblCurrentRows[i]) return;
  const t = _tblCurrentRows[i];
  // Zvýraznit aktivní tab
  document.querySelectorAll('[id^="tblTab"]').forEach((btn,j)=>
    btn.classList.toggle('active', j===i)
  );
  const cont = document.getElementById('tblContent');
  const caption = t.caption ? `<div class="tbl-caption">${esc(t.caption)}</div>` : '';
  const headers = t.headers && t.headers.length
    ? `<tr>${t.headers.map(h=>`<th>${esc(h)}</th>`).join('')}</tr>`
    : '';
  const rows = (t.rows||[]).map(row=>
    `<tr>${row.map(cell=>`<td>${esc(cell)}</td>`).join('')}</tr>`
  ).join('');
  cont.innerHTML = caption + `<table><thead>${headers}</thead><tbody>${rows}</tbody></table>`;
}

// Zavřít table popup kliknutím mimo
document.addEventListener('click', e=>{
  const popup = document.getElementById('tblPopup');
  if(popup && popup.classList.contains('on') && !popup.contains(e.target) && !e.target.closest('.tbl-badge')){
    popup.classList.remove('on');
  }
});

/* ── REVIEW EXPORT / IMPORT ──────────────────────────────────────────────── */
function exportReviewFile(){
  if(!_reviewPending){ clog('Žádná review data k uložení','warn'); return; }
  const data={
    version: 1,
    savedAt: new Date().toISOString(),
    params: window._runParams||{},
    pending: _reviewPending,
    cols: _cols,
    colSearch: _colSearch,
  };
  const blob=new Blob([JSON.stringify(data,null,2)],{type:'application/json'});
  const a=document.createElement('a');
  const name=(window._runParams?.output||'review').replace(/^[a-z0-9]+_/,'')||'review';
  a.href=URL.createObjectURL(blob);
  a.download=name+'.review.json';
  a.click();
  clog('💾 Review uloženo jako '+name+'.review.json','ok');
}

function importReviewFile(input){
  const file=input.files[0];
  if(!file) return;
  const msg=document.getElementById('loadReviewMsg');
  if(msg) msg.textContent='Načítám…';

  // Guard: nepřerušovat běžící scraping
  if(_phase2Running){
    if(msg){ msg.textContent='⚠ Nelze načíst — stahování právě probíhá'; msg.style.color='var(--warn)'; }
    input.value='';
    return;
  }

  const reader=new FileReader();
  reader.onload=e=>{
    try{
      const data=JSON.parse(e.target.result);
      if(!data.pending||!data.cols){
        if(msg){ msg.textContent='❌ Neplatný soubor review'; msg.style.color='var(--danger)'; }
        input.value=''; return;
      }
      // Obnovit stav
      _reviewPending=data.pending;
      _cols=data.cols;
      _colSearch=data.colSearch||data.cols.map(()=>'');
      if(data.params) window._runParams=data.params;
      // Otevřít review rovnou na sloupce (bez reveal animace)
      document.getElementById('reviewModal').classList.add('on');
      document.getElementById('reviewReveal').classList.remove('on');
      document.getElementById('reviewColumns').style.display='';
      document.getElementById('reviewFooter').style.display='';
      renderReview();
      updateSummary();
      setStep(2);
      if(msg){ msg.textContent='✅ Review načteno'; msg.style.color='var(--accent)'; }
      setTimeout(()=>{ if(msg) msg.textContent=''; },3000);
      clog('📂 Review načteno ze souboru: '+file.name,'ok');
    } catch(err){
      if(msg){ msg.textContent='❌ Chyba čtení: '+err.message; msg.style.color='var(--danger)'; }
    }
    input.value='';
  };
  reader.readAsText(file);
}


/* ── PARQUET + JSON-LD EXPORT ────────────────────────────────────────────── */
function exportParquet(){
  const outputs=_completedJobs.length>0?_completedJobs:[document.getElementById('output').value||'wiki_data'];
  const btn=document.getElementById('btnParquet');
  const st=document.getElementById('extraSt');
  if(btn) btn.disabled=true;
  st.textContent='⏳ Generuji Parquet…'; st.className='export-status';
  const doNext=(i)=>{
    if(i>=outputs.length){
      if(btn) btn.disabled=false;
      st.textContent=`✅ Parquet staženo (${outputs.length} soubor${outputs.length===1?'':'ů'})`;
      st.className='export-status ok'; return;
    }
    const params=new URLSearchParams({output:outputs[i],s:getSessionId()});
    fetch('/export_parquet?'+params).then(r=>{
      if(!r.ok) return r.json().then(d=>{ throw new Error(d.error||r.status); });
      return r.blob();
    }).then(blob=>{
      const a=document.createElement('a');
      a.href=URL.createObjectURL(blob);
      a.download=outputs[i]+'.parquet'; a.click();
      doNext(i+1);
    }).catch(e=>{
      if(btn) btn.disabled=false;
      st.textContent='❌ '+e.message; st.className='export-status err';
    });
  };
  doNext(0);
}

function exportJsonLd(){
  const outputs=_completedJobs.length>0?_completedJobs:[document.getElementById('output').value||'wiki_data'];
  const btn=document.getElementById('btnJsonLd');
  const st=document.getElementById('extraSt');
  if(btn) btn.disabled=true;
  st.textContent='⏳ Generuji JSON-LD…'; st.className='export-status';
  const doNext=(i)=>{
    if(i>=outputs.length){
      if(btn) btn.disabled=false;
      st.textContent=`✅ JSON-LD staženo`;
      st.className='export-status ok'; return;
    }
    const params=new URLSearchParams({output:outputs[i],s:getSessionId()});
    fetch('/export_jsonld?'+params).then(r=>{
      if(!r.ok) return r.json().then(d=>{ throw new Error(d.error||r.status); });
      return r.blob();
    }).then(blob=>{
      const a=document.createElement('a');
      a.href=URL.createObjectURL(blob);
      a.download=outputs[i]+'.jsonld'; a.click();
      doNext(i+1);
    }).catch(e=>{
      if(btn) btn.disabled=false;
      st.textContent='❌ '+e.message; st.className='export-status err';
    });
  };
  doNext(0);
}

/* ── PAUSE / RESUME ──────────────────────────────────────────────────────── */
function pauseScraping(){
  // Pauza = zastavit subprocess (checkpoint se uloží automaticky)
  // ale zachovat UI stav = zobrazit "Pokračovat" tlačítko
  fetch('/stop',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({s:getSessionId()})});
  if(evtSrc){ evtSrc.close(); evtSrc=null; }
  stopTimer();
  _phase2Running=false;
  setStatus('idle');
  document.getElementById('btnRun').classList.remove('hidden');
  document.getElementById('btnStopGroup').classList.remove('on');
  document.getElementById('btnStop').classList.add('hidden');
  document.getElementById('btnPause').classList.add('hidden');
  setPhase(2,'⏸ Pozastaveno — checkpoint uložen','');
  clog('⏸ Pozastaveno — pokračuj přes ♻ Pokračovat v levém menu','warn');
  checkCP();
}

/* ── DOWNLOAD ─────────────────────────────────────────────────────────────── */
// Track all completed jobs in current session
let _completedJobs = [];

function dlZip(){
  if(_completedJobs.length===1){
    const o=_completedJobs[0];
    window.location='/download?output='+encodeURIComponent(o);
  } else if(_completedJobs.length>1){
    // Stáhnout všechny jako jeden ZIP
    window.location='/download_multiple?outputs='+encodeURIComponent(_completedJobs.join(','));
  } else {
    const o=document.getElementById('output').value||'wiki_data';
    window.location='/download?output='+encodeURIComponent(o);
  }
}

function dlXlsx(){
  const jobs=_completedJobs.length>0 ? _completedJobs : [document.getElementById('output').value||'wiki_data'];
  // Vždy otevřít column picker (používá první soubor pro analýzu sloupců)
  openColPicker(jobs[0], jobs);
}

function _dlXlsxSingle(output, ibCols, onDone){
  const params=new URLSearchParams({output});
  if(ibCols) params.set('ib_cols', ibCols);
  fetch('/export_xlsx?'+params)
    .then(r=>{
      if(!r.ok) return r.json().then(d=>{throw new Error(d.error||r.statusText);});
      return r.blob();
    })
    .then(blob=>{
      const a=document.createElement('a');
      a.href=URL.createObjectURL(blob);
      a.download=output+'.xlsx'; a.click();
      if(onDone) onDone();
    })
    .catch(e=>{
      document.getElementById('btnXlsx').disabled=false;
      document.getElementById('xlsxSt').textContent='❌ '+e.message;
      document.getElementById('xlsxSt').className='export-status err';
    });
}

function showExportBox(jobs){
  _completedJobs=jobs;
  const box=document.getElementById('exportBox');
  box.classList.add('on');

  // Zobrazit seznam souborů
  const listEl=document.getElementById('exportFileList');
  if(listEl){
    if(jobs.length===1){
      listEl.textContent='';
    } else {
      listEl.innerHTML=jobs.map(j=>
        `<div style="display:flex;align-items:center;gap:8px;padding:3px 0;font-family:var(--mono);font-size:11px">
          <span style="color:var(--accent)">📄</span>
          <span style="color:var(--text2);flex:1">${esc(j)}</span>
          <button class="btn-sm" onclick="_dlXlsxSingle('${esc(j)}')" title="Exportovat do Excelu">XLSX</button>
          <button class="btn-sm" onclick="window.location='/download?output='+encodeURIComponent('${esc(j)}')" title="Stáhnout jako ZIP">ZIP</button>
        </div>`
      ).join('');
    }
  }
}


/* ── FULLTEXT ENGINE ──────────────────────────────────────────────────────── */
let _ftMode=false;  // true = AND/OR/NOT mode

function toggleFtMode(){
  _ftMode=!_ftMode;
  document.getElementById('btnFtMode').classList.toggle('ft',_ftMode);
  document.getElementById('dtQ').placeholder=_ftMode
    ? 'Fulltext: kyslík AND reakce, NOT sport, "periodická tabulka"'
    : 'Hledat v názvech, perexu, kategoriích, infoboxu…';
  renderTbl();
}

function ftMatch(record, query){
  // Build searchable text
  const text=[
    record.title||'',
    record.intro||'',
    (record.categories||[]).join(' '),
    Object.values(record.infobox||{}).join(' '),
    Object.values(record.sections||{}).join(' '),
  ].join(' ').toLowerCase();

  if(!_ftMode) return text.includes(query.toLowerCase());

  // Parse AND/OR/NOT — simple tokenizer
  // Supports: word, "phrase", AND, OR, NOT, word1 NOT word2
  const q=query.trim();
  if(!q) return true;

  // Split by OR first (lowest precedence)
  const orParts=q.split(/\bOR\b/i).map(s=>s.trim()).filter(Boolean);
  return orParts.some(orPart=>{
    const andParts=orPart.split(/\bAND\b/i).map(s=>s.trim()).filter(Boolean);
    return andParts.every(andPart=>{
      const nots=andPart.split(/\bNOT\b/i).map(s=>s.trim()).filter(Boolean);
      const mustHave=nots[0];
      const mustNot=nots.slice(1);
      const matchPhrase=(phrase)=>{
        const p=phrase.replace(/^"|"$/g,'').toLowerCase();
        return text.includes(p);
      };
      return matchPhrase(mustHave) && mustNot.every(n=>!matchPhrase(n));
    });
  });
}

function ftSnippet(record, query){
  if(!query||_ftMode) return '';
  const q=query.toLowerCase();
  const sources=[record.intro||'', Object.values(record.sections||{}).join(' ')];
  for(const src of sources){
    const idx=src.toLowerCase().indexOf(q);
    if(idx>=0){
      const start=Math.max(0,idx-40);
      const end=Math.min(src.length,idx+q.length+60);
      let snip=src.slice(start,end);
      if(start>0) snip='…'+snip;
      if(end<src.length) snip+='…';
      // Bold the match
      const hi=snip.replace(new RegExp(q.replace(/[.*+?^${}()|[\]\\]/g,'\\$&'),'gi'),
        m=>`<mark class="ft-highlight">${esc(m)}</mark>`);
      return `<span class="ft-snippet">${hi}</span>`;
    }
  }
  return '';
}

/* ── QUALITY BADGE ────────────────────────────────────────────────────────── */
function qualityBadge(r){
  const q=r._quality;
  if(!q) return '—';
  const s=typeof q==='object'?q.score:Number(q)||0;
  const cls=s>=70?'qb-high':s>=40?'qb-mid':'qb-low';
  const title=(q.details||[]).join(', ');
  return `<span class="quality-badge ${cls}" title="${esc(title)}">${s}</span>`;
}


/* ── TXT ZIP EXPORT ───────────────────────────────────────────────────────── */
function exportTxtZip(){
  const outputs=_completedJobs.length>0?_completedJobs:[document.getElementById('output').value||'wiki_data'];
  const btn=document.getElementById('btnTxtZip');
  const st=document.getElementById('extraSt');
  btn.disabled=true;
  st.textContent='Generuji ZIP s textovymi soubory...'; st.className='export-status';
  fetch('/export_txt_zip',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({outputs})
  })
  .then(r=>{
    if(!r.ok) return r.json().then(d=>{throw new Error(d.error||r.statusText);});
    return r.blob();
  })
  .then(blob=>{
    const name=(outputs.length===1?outputs[0]:'wiki_export')+'_texty.zip';
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    a.download=name; a.click();
    st.textContent='ZIP stazeno ('+outputs.reduce((s,o)=>s,0)+' hesel)';
    st.className='export-status ok';
    btn.disabled=false;
  })
  .catch(e=>{
    st.textContent='Chyba: '+e.message; st.className='export-status err';
    btn.disabled=false;
  });
}


/* ── DISCARD RESULTS ─────────────────────────────────────────────────────── */
function openDiscardModal(){
  document.getElementById('discardModal').style.display='flex';
}

function _deleteCompletedFiles(cb){
  // Smazat soubory vsech dokoncených jobů na serveru
  const jobs=_completedJobs.length>0?_completedJobs:[document.getElementById('output').value||'wiki_data'];
  let done=0;
  if(!jobs.length){ if(cb)cb(); return; }
  jobs.forEach(output=>{
    fetch('/discard_checkpoint',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({output, delete_results:true})})
      .finally(()=>{ done++; if(done>=jobs.length && cb) cb(); });
  });
}

function discardAndReview(){
  document.getElementById('discardModal').style.display='none';
  _deleteCompletedFiles(()=>{
    _completedJobs=[];
    document.getElementById('exportBox').classList.remove('on');
    allData=[];
    document.getElementById('dataCnt').textContent='';
    renderTbl();
    // Zobrazit review pokud existuje uložený stav
    const saved=localStorage.getItem('wikiReviewState');
    if(saved){
      resumeSavedReview();
    } else {
      clog('Zahozeno — spusť nový scraping','warn');
      setStatus('idle'); setStep(0);
    }
  });
}

function discardCompletely(){
  document.getElementById('discardModal').style.display='none';
  _deleteCompletedFiles(()=>{
    _completedJobs=[];
    allData=[];
    clearReviewState();
    document.getElementById('exportBox').classList.remove('on');
    document.getElementById('dataCnt').textContent='';
    renderTbl();
    document.getElementById('savedReviewBanner').style.display='none';
    document.getElementById('pb1').style.width='0%';
    document.getElementById('pb1').classList.remove('indeterminate');
    document.getElementById('pp1').textContent='—';
    document.getElementById('pb2').style.width='0%';
    document.getElementById('pp2').textContent='—';
    ['stFound','stDone','stEta','stCats'].forEach(id=>document.getElementById(id).textContent='—');
    document.getElementById('stTime').textContent='—';
    setStatus('idle'); setStep(0);
    setPhase(1,'Zahozeno — připraven pro nový scraping','');
    clog('Výsledky zahozeny','warn');
    document.getElementById('btnRun').classList.remove('hidden');
    document.getElementById('btnStop').classList.add('hidden');
  });
}

/* ── SQLITE + README EXPORT ───────────────────────────────────────────────── */
function exportSqlite(){
  const outputs=_completedJobs.length>0?_completedJobs:[document.getElementById('output').value||'wiki_data'];
  const btn=document.getElementById('btnSqlite');
  const st=document.getElementById('extraSt');
  btn.disabled=true; st.textContent='⏳ Generuji SQLite…'; st.className='export-status';
  const doNext=(i)=>{
    if(i>=outputs.length){
      btn.disabled=false;
      st.textContent=`✅ SQLite staženo (${outputs.length} soubor${outputs.length===1?'':'ů'})`;
      st.className='export-status ok'; return;
    }
    fetch('/export_sqlite',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({output:outputs[i]})})
      .then(r=>r.json()).then(d=>{
        if(d.ok) window.location=`/download_file?file=${encodeURIComponent(outputs[i]+'.db')}`;
        else{ st.textContent='❌ '+d.error; st.className='export-status err'; btn.disabled=false; return; }
        setTimeout(()=>doNext(i+1),400);
      }).catch(e=>{ st.textContent='❌ '+e; st.className='export-status err'; btn.disabled=false; });
  };
  doNext(0);
}

function exportReadme(){
  const output=(_completedJobs[0])||document.getElementById('output').value||'wiki_data';
  const url=document.getElementById('url').value||'';
  const st=document.getElementById('extraSt');
  st.textContent='⏳ Generuji README…'; st.className='export-status';
  fetch('/export_readme',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({output,url})})
    .then(r=>r.json()).then(d=>{
      if(d.ok){
        st.textContent='✅ README vygenerován → stahování';
        st.className='export-status ok';
        setTimeout(()=>window.location=`/download_file?file=${encodeURIComponent(output+'.README.md')}`,300);
      } else{ st.textContent='❌ '+d.error; st.className='export-status err'; }
    }).catch(e=>{ st.textContent='❌ '+e; st.className='export-status err'; });
}

/* ── MERGE ────────────────────────────────────────────────────────────────── */
let _mergeSelected=new Set();

function openMerge(){
  _mergeSelected=new Set();
  document.getElementById('mergeModal').classList.add('on');
  document.getElementById('mergeResult').classList.remove('on');
  loadMergeFiles();
}

function loadMergeFiles(){
  const list=document.getElementById('mergeFileList');
  list.innerHTML='<div style="font-family:var(--mono);font-size:11px;color:var(--text3);padding:8px 0">Načítám…</div>';
  fetch('/list_outputs?s='+encodeURIComponent(getSessionId())).then(r=>r.json()).then(files=>{
    const entries=Object.entries(files).sort((a,b)=>b[1].mtime-a[1].mtime);
    list.innerHTML=entries.map(([stem,info])=>`
      <div class="merge-file-row${_mergeSelected.has(stem)?' selected':''}" data-stem="${esc(stem)}" onclick="toggleMergeFile('${esc(stem)}')">
        <span class="merge-file-name">${esc(stem)}</span>
        <span class="merge-file-cnt">${info.count} hesel</span>
      </div>`).join('')||'<div style="font-family:var(--mono);font-size:11px;color:var(--text3)">Žádné soubory</div>';
    updateMergeSelected();
  }).catch(()=>{ list.innerHTML='<div style="color:var(--danger);font-family:var(--mono);font-size:11px">Chyba načítání</div>'; });
}

function toggleMergeFile(stem){
  if(_mergeSelected.has(stem)) _mergeSelected.delete(stem);
  else _mergeSelected.add(stem);
  document.querySelectorAll('.merge-file-row').forEach(r=>{
    r.classList.toggle('selected', _mergeSelected.has(r.dataset.stem));
  });
  updateMergeSelected();
}

function updateMergeSelected(){
  const cnt=_mergeSelected.size;
  document.getElementById('mergeCnt').textContent=`${cnt} soubor${cnt===1?'':cnt<5?'y':'ů'} vybráno`;
  document.getElementById('mergeSelected').innerHTML=[..._mergeSelected].map(s=>
    `<div style="padding:3px 0;border-bottom:1px solid var(--border)">${esc(s)}</div>`
  ).join('')||'Klikni na soubory vlevo…';
}

function doMerge(){
  if(!_mergeSelected.size){clog('⚠ Vyber alespoň jeden soubor','warn');return;}
  const output=document.getElementById('mergeOutput').value.trim()||'merged';
  const res=document.getElementById('mergeResult');
  res.classList.remove('on'); res.textContent='⏳ Slučuji…';
  fetch('/merge_files',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({sources:[..._mergeSelected],output})})
    .then(r=>r.json()).then(d=>{
      if(d.ok){
        res.innerHTML=`✅ Sloučeno: <b>${d.count}</b> hesel (${d.dupes} duplikátů odstraněno)<br>→ <code>${output}.json</code>`;
        res.classList.add('on');
        clog(`✅ Merge → ${output}.json: ${d.count} hesel, ${d.dupes} duplikátů odstraněno`,'ok');
        loadMergeFiles();
      } else{ res.textContent='❌ '+d.error; res.classList.add('on'); }
    }).catch(e=>{ res.textContent='❌ '+e; res.classList.add('on'); });
}

/* ── CONTEXT PREVIEW (cat hover in review) ───────────────────────────────── */
let _previewTimer=null;
const _previewEl=()=>document.getElementById('catPreview');

function catPreviewShow(catName, x, y){
  clearTimeout(_previewTimer);
  _previewTimer=setTimeout(()=>{
    const p=_previewEl();
    document.getElementById('cpTitle').textContent=catName;
    document.getElementById('cpMeta').textContent='Načítám hesla…';
    document.getElementById('cpItems').innerHTML='';
    p.style.left=(x+16)+'px';
    p.style.top=(y-10)+'px';
    p.classList.add('on');

    const output=window._runParams?.output||document.getElementById('output').value||'';
    fetch(`/get_article_preview?output=${encodeURIComponent(output)}&cat=${encodeURIComponent(catName)}&limit=6`)
      .then(r=>r.json())
      .then(d=>{
        document.getElementById('cpMeta').textContent=`${d.total||0} hesel v kategorii`;
        document.getElementById('cpItems').innerHTML=(d.items||[]).map(item=>
          `<div class="cat-preview-item">· ${esc(item.title)}</div>`
        ).join('')+(d.total>6?`<div style="color:var(--text3)">…a ${d.total-6} dalších</div>`:'');
      }).catch(()=>{
        document.getElementById('cpMeta').textContent='Náhled nedostupný';
      });
  }, 400);
}

function catPreviewHide(){
  clearTimeout(_previewTimer);
  _previewEl().classList.remove('on');
}

let _cpData=null;   // {cols_above, cols_below, total}
let _cpSelected=new Set();
let _cpJobs=[];

function openColPicker(output, jobs){
  _cpJobs=jobs||[output];
  document.getElementById('colpickModal').classList.add('on');
  document.getElementById('cpBody').innerHTML='<div style="padding:20px;text-align:center;font-family:var(--mono);font-size:12px;color:var(--text3)">Načítám sloupce…</div>';
  document.getElementById('cpSub').textContent='Načítám…';

  const thresh=parseInt(document.getElementById('cpThresh').value||20)/100;
  fetch(`/xlsx_columns?output=${encodeURIComponent(output)}&thresh=${thresh}`)
    .then(r=>r.json())
    .then(d=>{
      _cpData=d;
      // Defaultně vybrat vše co je nad prahem
      _cpSelected=new Set(d.cols_above.map(c=>c.key));
      renderColPicker();
    })
    .catch(e=>{ document.getElementById('cpBody').innerHTML=`<div style="padding:20px;color:var(--danger);font-family:var(--mono);font-size:12px">❌ ${esc(e.message)}</div>`; });
}

function cpUpdateThresh(){
  if(!_cpData) return;
  const thresh=parseInt(document.getElementById('cpThresh').value||20)/100;
  const output=_cpJobs[0]||document.getElementById('output').value||'wiki_data';
  fetch(`/xlsx_columns?output=${encodeURIComponent(output)}&thresh=${thresh}`)
    .then(r=>r.json())
    .then(d=>{ _cpData=d; renderColPicker(); });
}

function renderColPicker(){
  if(!_cpData) return;
  const {cols_above, cols_below, total} = _cpData;
  const thresh=parseInt(document.getElementById('cpThresh').value||20);
  document.getElementById('cpSub').textContent=
    `${total} hesel · ${cols_above.length} sloupců nad ${thresh}% · ${cols_below.length} pod prahem`;

  const makeRow=(c)=>{
    const on=_cpSelected.has(c.key);
    return `<div class="colpick-row${on?' on':''}" data-key="${encodeURIComponent(c.key)}">
      <div class="cp-chk">${on?'✓':''}</div>
      <div class="cp-name">${esc(c.key)}</div>
      <div class="cp-bar-wrap"><div class="cp-bar" style="width:${Math.min(100,c.pct)}%"></div></div>
      <div class="cp-pct">${c.pct}%</div>
    </div>`;
  };

  document.getElementById('cpBody').innerHTML=
    (cols_above.length?`<div class="colpick-section-hdr">✅ Nad prahem ${thresh}% (${cols_above.length})</div>${cols_above.map(makeRow).join('')}`:'') +
    (cols_below.length?`<div class="colpick-section-hdr">⬇ Pod prahem (${cols_below.length})</div>${cols_below.map(makeRow).join('')}`:'');

  // Event delegation
  document.getElementById('cpBody').querySelectorAll('.colpick-row').forEach(row=>{
    row.onclick=()=>{
      const key=decodeURIComponent(row.dataset.key);
      if(_cpSelected.has(key)) _cpSelected.delete(key);
      else _cpSelected.add(key);
      row.classList.toggle('on',_cpSelected.has(key));
      row.querySelector('.cp-chk').textContent=_cpSelected.has(key)?'✓':'';
      cpUpdateCount();
    };
  });
  cpUpdateCount();
}

function cpUpdateCount(){
  document.getElementById('cpCount').textContent=
    `${_cpSelected.size} sloupců vybráno`;
}

function cpSelectAll(val){
  if(!_cpData) return;
  const all=[..._cpData.cols_above,..._cpData.cols_below];
  if(val) all.forEach(c=>_cpSelected.add(c.key));
  else _cpSelected.clear();
  renderColPicker();
}

function cpSelectAboveThresh(){
  if(!_cpData) return;
  _cpSelected=new Set(_cpData.cols_above.map(c=>c.key));
  renderColPicker();
}

function cpConfirmExport(){
  document.getElementById('colpickModal').classList.remove('on');
  const ibCols=[..._cpSelected].join(',');
  const jobs=_cpJobs;

  document.getElementById('btnXlsx').disabled=true;
  document.getElementById('xlsxSt').textContent=`⏳ Exportuji…`;
  document.getElementById('xlsxSt').className='export-status';

  const doNext=(i)=>{
    if(i>=jobs.length){
      document.getElementById('btnXlsx').disabled=false;
      document.getElementById('xlsxSt').textContent=`✅ Staženo ${jobs.length} soubor${jobs.length===1?'':'ů'}`;
      document.getElementById('xlsxSt').className='export-status ok';
      return;
    }
    if(jobs.length>1) document.getElementById('xlsxSt').textContent=`⏳ ${i+1}/${jobs.length}: ${jobs[i]}…`;
    _dlXlsxSingle(jobs[i], ibCols, ()=>setTimeout(()=>doNext(i+1),300));
  };
  doNext(0);
}

/* ── DATA TABLE ───────────────────────────────────────────────────────────── */
function loadData(output){
  fetch('/results?file='+encodeURIComponent(output))
    .then(r=>r.json())
    .then(d=>{
      allData=d;
      _bulkSelected.clear(); _undoStack=[]; _redoStack=[]; updateUndoButtons();
      document.getElementById('dataCnt').textContent='('+d.length+')';
      document.getElementById('dataEmpty').classList.add('hidden');
      document.querySelector('.stats-bar')?.classList.add('visible');
      renderTbl(); buildViz(d); buildSummary(d);
      buildTagCloud(d); buildValidation(d); buildMap(d);
      addNotif('Data načtena','Soubor '+output+': '+d.length+' hesel','ok');
    })
    .catch(e=>clog('⚠ Nelze načíst data: '+e,'warn'));
}

function loadDataMultiple(outputs){
  Promise.all(outputs.map(o=>
    fetch('/results?file='+encodeURIComponent(o))
      .then(r=>r.json())
      .then(d=>d.map(rec=>({...rec, _source:o})))
      .catch(()=>[])
  )).then(arrays=>{
    allData=arrays.flat();
    _bulkSelected.clear(); _undoStack=[]; _redoStack=[]; updateUndoButtons();
    document.getElementById('dataCnt').textContent='('+allData.length+')';
    document.getElementById('dataEmpty').classList.add('hidden');
    renderTbl(); buildViz(allData); buildSummary(allData);
    buildTagCloud(allData); buildValidation(allData); buildMap(allData);
    if(outputs.length>1)
      clog(`📋 Tabulka: ${allData.length} hesel z ${outputs.length} souborů`,'ok');
  });
}

function renderTbl(){
  const q=document.getElementById('dtQ').value.toLowerCase().trim();
  let rows=allData.map(r=>({...r,
    _introLen:(r.intro||'').length,
    _qualityScore: typeof r._quality==='object'?(r._quality?.score||0):0
  }));

  if(q) rows=rows.filter(r=>ftMatch(r,q));

  if(_activeFilters.has('infobox'))  rows=rows.filter(r=>r.infobox&&Object.keys(r.infobox).length>0);
  if(_activeFilters.has('coords'))   rows=rows.filter(r=>r.coordinates);
  if(_activeFilters.has('fulltext')) rows=rows.filter(r=>r.full_text||r.sections);
  if(_activeFilters.has('error'))    rows=rows.filter(r=>r.error);
  if(_activeFilters.has('images'))   rows=rows.filter(r=>r.images&&r.images.length>0);
  if(_activeFilters.has('tables'))   rows=rows.filter(r=>r.tables&&r.tables.length>0);

  // Validační issue filtry
  if(_validIssueFilter){
    if(_validIssueFilter==='no_infobox')     rows=rows.filter(r=>!r.infobox||!Object.keys(r.infobox).length);
    else if(_validIssueFilter==='no_categories') rows=rows.filter(r=>!(r.categories||[]).length);
    else if(_validIssueFilter==='anomalies')  rows=rows.filter(r=>r._anomalies&&r._anomalies.length>0);
    else if(_validIssueFilter==='anom_error') rows=rows.filter(r=>r._anomalies&&r._anomalies.some(a=>a.severity==='error'));
    else if(_validIssueFilter==='anom_warn')  rows=rows.filter(r=>r._anomalies&&r._anomalies.some(a=>a.severity==='warn'));
    else if(_validIssueFilter==='stubs')     rows=rows.filter(r=>r.is_stub);
    else if(_validIssueFilter==='low_quality') rows=rows.filter(r=>typeof r._quality==='object'&&(r._quality?.score||0)<40);
    else if(_validIssueFilter==='no_intro')  rows=rows.filter(r=>!(r.intro||'').trim());
    else if(_validIssueFilter==='short_intro') rows=rows.filter(r=>{const l=(r.intro||'').length;return l>0&&l<100;});
  }

  if(sk) rows=[...rows].sort((a,b)=>{
    if(sk==='_introLen'||sk==='_quality'){
      const av=sk==='_quality'?a._qualityScore:a._introLen;
      const bv=sk==='_quality'?b._qualityScore:b._introLen;
      return sasc?av-bv:bv-av;
    }
    let av=a[sk]||'',bv=b[sk]||'';
    if(Array.isArray(av))av=av.join(' '); if(Array.isArray(bv))bv=bv.join(' ');
    return sasc?String(av).localeCompare(String(bv)):String(bv).localeCompare(String(av));
  });

  document.getElementById('dtCnt').textContent='…';
  document.getElementById('noRows').classList.add('hidden');

  // Pokud je Worker dostupný, filtruj + řaď na pozadí
  if(_filterWorker){
    const payload={
      rows:allData.map(r=>({...r,
        _introLen:(r.intro||'').length,
        _qualityScore:typeof r._quality==='object'?(r._quality?.score||0):0
      })),
      q:document.getElementById('dtQ').value.toLowerCase().trim(),
      ftMode:_ftMode,
      activeFilters:[..._activeFilters],
      validIssueFilter:_validIssueFilter||null,
      activeTagFilter:_activeTagFilter||null,
      sk, sasc
    };
    _filterWorker.postMessage(payload);
    return;
  }

  // Fallback — hlavní vlákno (Worker nedostupný)
  if(_activeTagFilter) rows=rows.filter(r=>(r._tags||[]).includes(_activeTagFilter));
  document.getElementById('dtCnt').textContent=rows.length+' / '+allData.length+(_ftMode?' [FT]':'');
  document.getElementById('noRows').classList.toggle('hidden',rows.length>0);
  _vtRows=rows;
  _renderVirtualRows();
}

/* ── VIRTUAL SCROLL ENGINE ────────────────────────────────────────────────── */
const ROW_HEIGHT=40;   // odhadovaná výška řádku v px
const VSCROLL_BUF=12;  // buffer řádků mimo viewport
let _vtRows=[];        // aktuálně filtrované+seřazené řádky

function _buildRowHtml(r, i, qRaw){
  const cats=Array.isArray(r.categories)?r.categories.slice(0,3).join(', '):(r.categories||'');
  const ibn=Object.keys(r.infobox||{}).length;
  const ibd=ibn?encodeURIComponent(JSON.stringify(r.infobox)):'';
  const errStyle=r.error?'color:var(--danger)':'';
  const introLen=r._introLen;
  const introColor=introLen<100?'var(--danger)':introLen<300?'var(--warn)':'var(--text3)';
  const snippet=qRaw&&!_ftMode?ftSnippet(r,qRaw):'';
  const stub=r.is_stub?'<span style="font-family:var(--mono);font-size:9px;color:var(--warn);margin-left:4px">stub</span>':'';
  const anomBadge=(r._anomalies&&r._anomalies.length)
    ?`<span style="font-family:var(--mono);font-size:9px;margin-left:4px;color:${
        r._anomalies.some(a=>a.severity==='error')?'var(--danger)':'var(--warn)'
      }" title="${r._anomalies.map(a=>a.message).join(', ')}">⚑${r._anomalies.length}</span>`
    :'';
  const bulkCell=_bulkMode?`<td class="td-sel"><input type="checkbox" class="row-checkbox" ${_bulkSelected.has(r.url)?'checked':''} onchange="toggleBulkRow('${esc(r.url)}',this)"></td>`:'';
  const focusedClass=_tblFocusIdx===i?' tr-focused':'';
  const imgStrip=(r.images&&r.images.length)
    ? `<div class="img-strip">${r.images.slice(0,5).map(img=>
        `<img class="img-thumb" src="${esc(img.thumb||img.url)}" title="${esc(img.caption||'')}" loading="lazy" onerror="this.style.display='none'">`
      ).join('')}</div>`
    : '';
  const tblBadge=(r.tables&&r.tables.length)
    ? `<span class="tbl-badge" onclick="showTables(event,${i})" title="Zobrazit ${r.tables.length} tabulku/tabulky">📊 ${r.tables.length}</span>`
    : '';
  return `<tr data-idx="${i}" class="${focusedClass}" onclick="tblRowClick(${i})">
    ${bulkCell}
    <td class="td-num col-c0" style="${errStyle}">${i+1}</td>
    <td class="td-title col-c1">
      <a href="${esc(r.url)}" target="_blank">${esc(r.title||'—')}</a>${stub}${anomBadge}${tblBadge}
      ${snippet}
      ${r.error?`<span style="color:var(--danger);font-size:10px;display:block">${esc(r.error)}</span>`:''}
      ${imgStrip}
    </td>
    <td class="td-text col-c2">${esc((r.intro||'').slice(0,100))}</td>
    <td class="td-cats td-text col-c3">${esc(cats)}</td>
    <td class="col-c4" style="font-family:var(--mono);font-size:10px;color:${introColor};text-align:right;padding-right:16px">${introLen}</td>
    <td class="col-c5" style="text-align:center">${qualityBadge(r)}</td>
    <td class="td-ib-count col-c6">${ibn?`<span class="ib-link" onclick="showIb(event,'${ibd}')">${ibn} pol.</span>`:'—'}</td>
    <td class="col-c7"><button class="copy-btn" onclick="copyUrl('${esc(r.url)}',this)" title="Kopírovat URL">⎘</button></td>
  </tr>`;
}

function _renderVirtualRows(){
  if(!_vtRows) return;
  const wrap=document.querySelector('.dt-wrap');
  if(!wrap) return;
  const scrollTop=wrap.scrollTop;
  const viewH=wrap.clientHeight||500;
  const cols=_bulkMode?9:8;

  const startIdx=Math.max(0, Math.floor(scrollTop/ROW_HEIGHT)-VSCROLL_BUF);
  const endIdx  =Math.min(_vtRows.length, Math.ceil((scrollTop+viewH)/ROW_HEIGHT)+VSCROLL_BUF);
  const topPad  =startIdx*ROW_HEIGHT;
  const botPad  =Math.max(0,(_vtRows.length-endIdx)*ROW_HEIGHT);
  const qRaw    =document.getElementById('dtQ').value.trim();

  const dtQ=document.getElementById('dtQ');
  const hasFocus=document.activeElement===dtQ;
  const selStart=hasFocus?dtQ.selectionStart:null;
  const selEnd  =hasFocus?dtQ.selectionEnd:null;

  document.getElementById('dtBody').innerHTML=
    `<tr class="vscroll-spacer" style="height:${topPad}px"><td colspan="${cols}"></td></tr>`+
    _vtRows.slice(startIdx,endIdx).map((r,i)=>_buildRowHtml(r,startIdx+i,qRaw)).join('')+
    `<tr class="vscroll-spacer" style="height:${botPad}px"><td colspan="${cols}"></td></tr>`;

  if(hasFocus){ dtQ.focus(); try{dtQ.setSelectionRange(selStart,selEnd);}catch(e){} }
}
function filterTbl(){
  _validIssueFilter=null;
  _activeTagFilter=null;
  _activeTag=null;
  renderTbl();
}
function sortBy(k){
  if(sk===k)sasc=!sasc;else{sk=k;sasc=true;}
  ['title','intro','categories'].forEach(x=>{
    const el=document.getElementById('s'+x.charAt(0).toUpperCase()+x.slice(1));
    if(el) el.textContent=x===k?(sasc?'↑':'↓'):'↕';
  });
  renderTbl();
}
function showIb(e,enc){
  e.stopPropagation();
  const data=JSON.parse(decodeURIComponent(enc));
  const pop=document.getElementById('ibPopup');
  pop.innerHTML=Object.entries(data).map(([k,v])=>`<div class="ib-row"><span class="ib-k">${esc(k)}</span><span class="ib-v">${esc(v)}</span></div>`).join('');
  pop.style.cssText=`display:block;left:${Math.min(e.clientX+8,innerWidth-320)}px;top:${Math.min(e.clientY+8,innerHeight-380)}px`;
}
document.addEventListener('click',()=>document.getElementById('ibPopup').style.display='none');

/* ── COUNT-UP ANIMATION ───────────────────────────────────────────────────── */
function countUp(el, target, duration=700){
  const start=Date.now();
  const from=parseInt(el.textContent)||0;
  const tick=()=>{
    const p=Math.min((Date.now()-start)/duration,1);
    const ease=1-Math.pow(1-p,3); // cubic ease-out
    el.textContent=Math.round(from+ease*(target-from));
    if(p<1) requestAnimationFrame(tick);
  };
  requestAnimationFrame(tick);
}

/* ── SUMMARY CARDS ────────────────────────────────────────────────────────── */
function buildSummary(data){
  document.getElementById('overviewEmpty').classList.add('hidden');
  const cards=document.getElementById('summaryCards');
  cards.classList.remove('hidden');
  const cats={};
  data.forEach(r=>(r.categories||[]).forEach(c=>{cats[c]=(cats[c]||0)+1;}));
  const topCats=Object.entries(cats).sort((a,b)=>b[1]-a[1]).slice(0,5);
  cards.innerHTML=buildStats(data)+`
    <div style="margin-top:16px;font-family:var(--mono);font-size:10px;color:var(--text3);text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px">Top kategorie</div>
    ${topCats.map(([c,n])=>`
      <div style="display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid var(--border);font-family:var(--mono);font-size:11px">
        <span style="color:var(--text2)">${esc(c)}</span>
        <span style="color:var(--accent)">${n}</span>
      </div>`).join('')||'<div style="color:var(--text3);font-family:var(--mono);font-size:11px">Žádné kategorie</div>'}`;
  // Animace čísel v stat kartičkách
  setTimeout(()=>{
    cards.querySelectorAll('.sg-val').forEach(el=>{
      const target=parseInt(el.textContent)||0;
      if(target>0){ el.textContent='0'; countUp(el,target); }
    });
  },60);
}

/* ── VIZ ──────────────────────────────────────────────────────────────────── */
let mapInst=null;
function buildViz(data){
  const area=document.getElementById('vizArea');
  area.innerHTML='';

  // Načíst barvy z CSS proměnných (funguje v dark i light modu)
  const cs=getComputedStyle(document.documentElement);
  const chartGrid=cs.getPropertyValue('--chart-grid').trim()||'#2a2f3d';
  const chartTick=cs.getPropertyValue('--chart-tick').trim()||'#8892a4';

  // Bar chart — kategorie
  const cats={};
  data.forEach(r=>(r.categories||[]).forEach(c=>{cats[c]=(cats[c]||0)+1;}));
  const topCats=Object.entries(cats).sort((a,b)=>b[1]-a[1]).slice(0,12);
  if(topCats.length){
    const card=document.createElement('div');
    card.className='viz-card'; card.style.flex='1'; card.style.minWidth='380px';
    card.innerHTML='<div class="viz-card-title">Top kategorie<span class="sec-caret" style="font-size:10px;margin-left:auto">▾</span></div><div class="viz-card-body"><canvas id="catChart" height="200"></canvas></div>';
    area.appendChild(card);
    setTimeout(()=>{
      new Chart(document.getElementById('catChart').getContext('2d'),{
        type:'bar',
        data:{
          labels:topCats.map(([c])=>c.length>24?c.slice(0,22)+'…':c),
          datasets:[{data:topCats.map(([,n])=>n),backgroundColor:'rgba(110,231,183,.65)',borderColor:'rgba(110,231,183,1)',borderWidth:1,borderRadius:3}]
        },
        options:{indexAxis:'y',plugins:{legend:{display:false}},
          scales:{x:{ticks:{color:chartTick,font:{family:'IBM Plex Mono',size:10}},grid:{color:chartGrid}},
                  y:{ticks:{color:chartTick,font:{family:'IBM Plex Mono',size:10}},grid:{color:chartGrid}}}}
      });
    },100);
  }

  // Map
  const withCoords=data.filter(r=>r.coordinates);
  if(withCoords.length){
    const card=document.createElement('div');
    card.className='viz-card'; card.style.flex='1 1 100%';
    card.innerHTML=`<div class="viz-card-title">Mapa · ${withCoords.length} hesel se souřadnicemi<span class="sec-caret" style="font-size:10px;margin-left:auto">▾</span></div><div class="viz-card-body"><div id="mapWrap"></div></div>`;
    area.appendChild(card);
    setTimeout(()=>{
      if(mapInst){mapInst.remove();mapInst=null;}
      mapInst=L.map('mapWrap').setView([50,15],4);
      L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png',{
        attribution:'&copy; OpenStreetMap &amp; CartoDB',maxZoom:18
      }).addTo(mapInst);
      const icon=L.divIcon({className:'',html:'<div style="width:8px;height:8px;border-radius:50%;background:#6ee7b7;border:1.5px solid #0f1117;box-shadow:0 0 4px rgba(110,231,183,.5)"></div>',iconSize:[8,8]});
      withCoords.forEach(r=>{
        const raw=r.coordinates||'';
        let lat=null,lng=null;
        const dms=raw.match(/([\d.]+)°.*?([NS]).*?([\d.]+)°.*?([EW])/i);
        if(dms){lat=parseFloat(dms[1])*(dms[2].toUpperCase()==='S'?-1:1);lng=parseFloat(dms[3])*(dms[4].toUpperCase()==='W'?-1:1);}
        if(!lat){const dec=raw.match(/([-\d.]+)[,\s]+([-\d.]+)/);if(dec){lat=parseFloat(dec[1]);lng=parseFloat(dec[2]);}}
        if(lat&&lng&&Math.abs(lat)<=90&&Math.abs(lng)<=180)
          L.marker([lat,lng],{icon}).addTo(mapInst).bindPopup(`<b>${r.title}</b><br><a href="${r.url}" target="_blank" style="color:#60a5fa">→ Wiki</a>`);
      });
    },200);
  }

  if(!area.children.length)
    area.innerHTML='<div class="overview-empty" style="width:100%"><div class="overview-icon">📊</div><div>Nedostatek dat</div></div>';
}

function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function escAttr(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');}


/* ── STEPPER ──────────────────────────────────────────────────────────────── */
function setStep(n){
  // n: 0=idle, 1=phase1 active, 2=review, 3=phase2 active, 4=done
  for(let i=0;i<4;i++){
    const s=document.getElementById('step'+i);
    const d=document.getElementById('sd'+i);
    const l=document.getElementById('sl'+i);
    if(s) s.className='step'+(i<n?' done':i===n?' active':'');
    if(l) l.className='step-line'+(i<n?' done':'');
  }
}

/* ── DRAG & DROP URL ──────────────────────────────────────────────────────── */
function urlDragOver(e){
  e.preventDefault();
  document.getElementById('url').classList.add('drag-over');
}
function urlDragLeave(e){
  document.getElementById('url').classList.remove('drag-over');
}
function urlDrop(e){
  e.preventDefault();
  document.getElementById('url').classList.remove('drag-over');
  const text=e.dataTransfer.getData('text/plain')||e.dataTransfer.getData('text/uri-list');
  if(text && (text.startsWith('http')||text.startsWith('https'))){
    let url=text.trim().split('\n')[0];
    try{ url=decodeURIComponent(url); }catch(ex){}
    document.getElementById('url').value=url;
    onUrlChange();
  }
}

/* ── TEST URL ─────────────────────────────────────────────────────────────── */
async function testUrl(){
  const url=document.getElementById('url').value.trim();
  const btn=document.getElementById('btnTest');
  const res=document.getElementById('testResult');
  if(!url){ res.className='test-result on err'; res.textContent='Zadej URL'; return; }
  btn.disabled=true; btn.textContent='⏳';
  res.className='test-result on'; res.textContent='Načítám…';
  try{
    const r=await fetch('/test_url?url='+encodeURIComponent(url));
    const d=await r.json();
    if(d.ok){
      res.className='test-result on ok';
      res.innerHTML=`✅ <b>${d.type}</b> · ${d.cats>0?d.cats+' podkategorií · ':''}${d.pages} hesel na první stránce<br><span style="color:var(--text3)">${d.title}</span>`;
      if(d.suggested_output) document.getElementById('output').value=d.suggested_output;
    } else {
      res.className='test-result on err';
      res.textContent='❌ '+d.error;
    }
  }catch(e){
    res.className='test-result on err';
    res.textContent='❌ Server neodpovídá';
  }
  btn.disabled=false; btn.textContent='🔍 Otestovat';
}

/* ── BLACKLIST ────────────────────────────────────────────────────────────── */
function getBlacklist(){
  const raw=document.getElementById('blacklist')?.value||'';
  return raw.split(',').map(s=>s.trim().toLowerCase()).filter(Boolean);
}

function updateBlacklistCount(){
  const bl=getBlacklist();
  const el=document.getElementById('blCount');
  if(!el) return;
  if(bl.length===0){ el.classList.remove('on'); return; }
  // Count how many cats would be auto-unchecked (only if pending data available)
  if(_reviewPending){
    const cats=Object.keys(_reviewPending.cat_articles||{});
    const n=cats.filter(c=>bl.some(w=>blMatch(c,w))).length;
    el.classList.toggle('on', n>0);
    el.textContent=`⚠ ${n} kategorií bude automaticky odškrtnuto`;
  } else {
    el.classList.toggle('on', bl.length>0);
    el.textContent=`${bl.length} výraz${bl.length===1?'':'ů'}`;
  }
}

// Globální set kategorií odškrtnutých blacklistem
let _blacklistedCats=new Set();

function blMatch(catName, pattern){
  // Podporuje wildcards: hydro* , *natý , *chemi*
  const c=catName.toLowerCase();
  const p=pattern.toLowerCase();
  if(!p) return false;
  const hasStart=p.startsWith('*');
  const hasEnd=p.endsWith('*');
  if(hasStart && hasEnd){
    // *text* — contains
    return c.includes(p.slice(1,-1));
  } else if(hasStart){
    // *text — ends with
    return c.endsWith(p.slice(1));
  } else if(hasEnd){
    // text* — starts with
    return c.startsWith(p.slice(0,-1));
  } else {
    // žádný wildcard — contains (původní chování)
    return c.includes(p);
  }
}

function applyBlacklist(){
  const bl=getBlacklist();
  _blacklistedCats=new Set();
  if(!bl.length||!_cols.length) return;
  let count=0;
  Object.keys(_cols[0].cats).forEach(cat=>{
    if(bl.some(w=>blMatch(cat,w))){
      if(_cols[0].cats[cat]) count++;
      _cols[0].cats[cat]=false;
      _blacklistedCats.add(cat);
    }
  });
  if(count>0){
    clog(`🚫 Blacklist: automaticky odškrtnuto ${count} kategorií`,'warn');
    renderReview();
  }
}

/* ── REVEAL ANIMATION ─────────────────────────────────────────────────────── */
function showReviewReveal(catCount, heselCount, removed){
  document.getElementById('reviewReveal').classList.add('on');
  document.getElementById('reviewColumns').style.display='none';
  document.getElementById('reviewFooter').style.display='none';

  // Animate number count-up
  const target=catCount;
  const el=document.getElementById('rvNumber');
  let current=0;
  const step=Math.max(1,Math.floor(target/30));
  const interval=setInterval(()=>{
    current=Math.min(current+step, target);
    el.textContent=current;
    if(current>=target) clearInterval(interval);
  },30);

  document.getElementById('rvSub').textContent=
    `${heselCount} hesel${removed>0?' · −'+removed+' duplikátů':''}`;
  document.getElementById('rvHesel').textContent=`(${heselCount} hesel)`;
}

function showReviewColumns(){
  document.getElementById('reviewReveal').classList.remove('on');
  document.getElementById('reviewColumns').style.display='';
  document.getElementById('reviewFooter').style.display='';
  renderReview();
}

/* ── FILTER CHIPS ─────────────────────────────────────────────────────────── */
const _activeFilters=new Set();
function toggleFilter(f){
  if(_activeFilters.has(f)) _activeFilters.delete(f);
  else _activeFilters.add(f);
  _validIssueFilter=null;  // zrušit validační filtr
  _activeTagFilter=null; _activeTag=null;
  document.getElementById('rf-'+f).classList.toggle('on',_activeFilters.has(f));
  buildTagCloud(allData);
  renderTbl();
}

/* ── COPY URL ─────────────────────────────────────────────────────────────── */
function copyUrl(url, btn){
  navigator.clipboard.writeText(url).then(()=>{
    btn.classList.add('copied'); btn.textContent='✓';
    setTimeout(()=>{ btn.classList.remove('copied'); btn.textContent='⎘'; },1500);
  });
}

/* ── RENAME FILE ──────────────────────────────────────────────────────────── */
let _currentFile='';
function toggleRename(){
  const sec=document.getElementById('renameSection');
  const inp=document.getElementById('renameInput');
  const showing=sec.style.display!=='none';
  sec.style.display=showing?'none':'inline-flex';
  if(!showing){ inp.value=_currentFile; inp.focus(); inp.select(); }
}
function renameFile(){
  const newName=document.getElementById('renameInput').value.trim();
  if(!newName||newName===_currentFile) return;
  fetch('/rename_file',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({old:_currentFile,new:newName})})
    .then(r=>r.json()).then(d=>{
      if(d.ok){
        _currentFile=newName;
        document.getElementById('output').value=newName;
        document.getElementById('renameSection').style.display='none';
        clog(`✅ Přejmenováno → ${newName}`,'ok');
      } else clog('❌ '+d.error,'err');
    });
}

/* ── STATS ────────────────────────────────────────────────────────────────── */
function buildStats(data){
  // Overview cards
  const withIb=data.filter(r=>r.infobox&&Object.keys(r.infobox).length>0).length;
  const withCoords=data.filter(r=>r.coordinates).length;
  const withText=data.filter(r=>r.full_text||r.sections).length;
  const withImages=data.filter(r=>r.images&&r.images.length>0).length;
  const withTables=data.filter(r=>r.tables&&r.tables.length>0).length;
  const withErr=data.filter(r=>r.error).length;
  const avgIntroLen=data.length?Math.round(data.reduce((s,r)=>(r.intro||'').length+s,0)/data.length):0;

  // Intro length histogram
  const buckets=[
    {label:'0–100',   min:0,   max:100},
    {label:'100–300', min:100, max:300},
    {label:'300–600', min:300, max:600},
    {label:'600+',    min:600, max:Infinity},
  ];
  const maxCount=Math.max(...buckets.map(b=>data.filter(r=>{const l=(r.intro||'').length;return l>=b.min&&l<b.max;}).length));

  const statsHtml=`
    <div class="stats-grid">
      <div class="sg-card"><div class="sg-val">${data.length}</div><div class="sg-lbl">Hesel</div></div>
      <div class="sg-card"><div class="sg-val" style="color:var(--accent2)">${withIb}</div><div class="sg-lbl">S infoboxem</div></div>
      <div class="sg-card"><div class="sg-val" style="color:var(--warn)">${withCoords}</div><div class="sg-lbl">Se souřadnicemi</div></div>
      <div class="sg-card"><div class="sg-val" style="color:var(--text2)">${withText}</div><div class="sg-lbl">S textem</div></div>
      ${withErr?`<div class="sg-card"><div class="sg-val" style="color:var(--danger)">${withErr}</div><div class="sg-lbl">Chyby</div></div>`:''}
      ${withImages?`<div class="sg-card"><div class="sg-val" style="color:var(--accent2)">${withImages}</div><div class="sg-lbl">S obrázky</div></div>`:''}
      ${withTables?`<div class="sg-card"><div class="sg-val" style="color:var(--accent2)">${withTables}</div><div class="sg-lbl">S tabulkami</div></div>`:''}
      <div class="sg-card"><div class="sg-val" style="color:var(--text2)">${avgIntroLen}</div><div class="sg-lbl">Průměr perex</div></div>
    </div>
    <div style="font-family:var(--mono);font-size:10px;color:var(--text3);margin-bottom:6px;text-transform:uppercase;letter-spacing:.08em">Délka perexu (znaky)</div>
    <div class="intro-histogram">
      ${buckets.map(b=>{
        const count=data.filter(r=>{const l=(r.intro||'').length;return l>=b.min&&l<b.max;}).length;
        const pct=maxCount?Math.round(count/maxCount*100):0;
        return `<div class="hist-bar-wrap">
          <span class="hist-label">${b.label}</span>
          <div class="hist-bar" style="width:${Math.max(pct,2)}%"></div>
          <span class="hist-count">${count}</span>
        </div>`;
      }).join('')}
    </div>`;

  return statsHtml;
}

/* ── KEYBOARD NAVIGATION IN REVIEW ───────────────────────────────────────── */
// Ukládáme JMÉNO kategorie, ne DOM referenci — ta se zneplatní po každém renderReview()
let _focusedCatName=null;

function _reviewFocusRowByName(name){
  const rows=[...document.querySelectorAll('.cat-row')];
  rows.forEach(r=>{ r.style.outline=''; r.classList.remove('focused'); });
  if(!name){ _focusedCatName=null; return; }
  const row=rows.find(r=>{
    const el=r.querySelector('[data-name]');
    return el && decodeURIComponent(el.dataset.name)===name;
  });
  if(row){
    row.classList.add('focused');
    row.scrollIntoView({block:'nearest'});
  } else {
    _focusedCatName=null;  // Řádek zmizel (přesun do jiného sloupce)
  }
}

function initReviewKeyboard(){
  document.addEventListener('keydown',e=>{
    const modal=document.getElementById('reviewModal');
    if(!modal.classList.contains('on')) return;
    if(document.activeElement?.tagName==='INPUT') return;

    // Ctrl+Z undo review, Ctrl+Y redo review
    if((e.ctrlKey||e.metaKey) && e.key==='z' && !e.shiftKey){
      e.preventDefault(); e.stopPropagation(); undoReview(); return;
    }
    if((e.ctrlKey||e.metaKey) && (e.key==='y'||(e.key==='z'&&e.shiftKey))){
      e.preventDefault(); e.stopPropagation(); redoReview(); return;
    }

    if(e.key==='ArrowDown'||e.key==='ArrowUp'){
      e.preventDefault();
      const rows=[...document.querySelectorAll('.cat-row')];
      if(!rows.length) return;
      let idx=-1;
      if(_focusedCatName){
        idx=rows.findIndex(r=>{
          const el=r.querySelector('[data-name]');
          return el && decodeURIComponent(el.dataset.name)===_focusedCatName;
        });
      }
      if(e.key==='ArrowDown') idx=Math.min(idx+1,rows.length-1);
      else idx=Math.max(idx-1,0);
      const row=rows[idx];
      if(!row) return;
      const nameEl=row.querySelector('[data-name]');
      _focusedCatName=nameEl?decodeURIComponent(nameEl.dataset.name):null;
      rows.forEach(r=>{ r.style.outline=''; r.classList.remove('focused'); });
      row.classList.add('focused');
      row.scrollIntoView({block:'nearest'});
    }

    if(e.key===' '&&_focusedCatName){
      e.preventDefault();
      const row=document.querySelector(`.cat-row [data-name="${encodeURIComponent(_focusedCatName)}"]`)?.closest('.cat-row');
      if(row){
        // Zjistit stav PŘED togglem
        const isChecked=row.classList.contains('checked');
        // Při odškrtání (checked→unchecked) přesunout focus na další řádek
        if(isChecked){
          const rows=[...document.querySelectorAll('.cat-row')];
          const idx=rows.findIndex(r=>{
            const el=r.querySelector('[data-name]');
            return el && decodeURIComponent(el.dataset.name)===_focusedCatName;
          });
          const next=rows[idx+1];
          if(next){
            const nameEl=next.querySelector('[data-name]');
            if(nameEl) _focusedCatName=decodeURIComponent(nameEl.dataset.name);
          }
        }
        // Kliknout toggle
        const chk=row.querySelector('.cat-check-cell');
        if(chk) chk.click();
        // Po rerenderování obnovit focus (na nové pozici)
        setTimeout(()=>_reviewFocusRowByName(_focusedCatName),0);
      }
    }

    if(e.key==='ArrowRight'&&_focusedCatName){
      e.preventDefault();
      const row=document.querySelector(`.cat-row [data-name="${encodeURIComponent(_focusedCatName)}"]`)?.closest('.cat-row');
      if(row){
        const arr=row.querySelector('.cat-arr:last-child');
        if(arr&&!arr.classList.contains('disabled')){
          arr.click();
          setTimeout(()=>_reviewFocusRowByName(_focusedCatName),0);
        }
      }
    }

    if(e.key==='ArrowLeft'&&_focusedCatName){
      e.preventDefault();
      const row=document.querySelector(`.cat-row [data-name="${encodeURIComponent(_focusedCatName)}"]`)?.closest('.cat-row');
      if(row){
        const arr=row.querySelector('.cat-arr:first-child');
        if(arr&&!arr.classList.contains('disabled')){
          arr.click();
          setTimeout(()=>_reviewFocusRowByName(_focusedCatName),0);
        }
      }
    }
  });
}

/* ── INFO MODAL ───────────────────────────────────────────────────────────── */
function switchInfoTab(n){
  [0,1].forEach(i=>{
    document.getElementById('infoTab'+i).classList.toggle('active', i===n);
    document.getElementById('infoBody'+i).style.display=i===n?'block':'none';
  });
}

/* ── ONBOARDING ───────────────────────────────────────────────────────────── */
function initOnboarding(){
  if(!localStorage.getItem('wikiOnboardDone')){
    document.getElementById('onboardModal').classList.add('on');
  }
}
function closeOnboard(){
  document.getElementById('onboardModal').classList.remove('on');
  localStorage.setItem('wikiOnboardDone','1');
}

/* ── NOTIFICATION CENTER ──────────────────────────────────────────────────── */
let _notifs=[];
const MAX_NOTIFS=50;

function addNotif(title, body, type='info'){
  const n={title, body, type, time:new Date().toLocaleTimeString('cs-CZ'), ts:Date.now()};
  _notifs.unshift(n);
  if(_notifs.length>MAX_NOTIFS) _notifs=_notifs.slice(0,MAX_NOTIFS);
  renderNotifs();
  const badge=document.getElementById('notifBadge');
  if(badge){ badge.textContent=_notifs.length; badge.classList.add('on'); }
}

function renderNotifs(){
  const list=document.getElementById('notifList');
  if(!list) return;
  if(!_notifs.length){ list.innerHTML='<div class="notif-empty">Zatím žádné notifikace</div>'; return; }
  const icons={info:'ℹ',ok:'✅',warn:'⚠',err:'❌'};
  list.innerHTML=_notifs.map(n=>`
    <div class="notif-item">
      <span class="notif-item-time">${n.time}</span>
      <div class="notif-item-title">${icons[n.type]||'·'} ${esc(n.title)}</div>
      <div class="notif-item-body">${esc(n.body||'')}</div>
    </div>`).join('');
}

function clearNotifs(){
  _notifs=[];
  renderNotifs();
  const badge=document.getElementById('notifBadge');
  if(badge){ badge.textContent=''; badge.classList.remove('on'); }
}

function toggleNotifPanel(){
  const p=document.getElementById('notifPanel');
  p.classList.toggle('on');
  if(p.classList.contains('on')){
    // Mark as seen — clear badge
    const badge=document.getElementById('notifBadge');
    if(badge){ badge.textContent=''; badge.classList.remove('on'); }
  }
}

// Close notif panel on click outside
document.addEventListener('click', e=>{
  const wrap=document.querySelector('.notif-bell-wrap');
  if(wrap&&!wrap.contains(e.target)){
    document.getElementById('notifPanel')?.classList.remove('on');
  }
});

/* ── MAP ──────────────────────────────────────────────────────────────────── */
let _map=null, _mapMarkers=[];

function buildMap(data){
  const withCoords=data.filter(r=>r.coordinates_norm?.lat!=null);
  const mapTabBtn=document.getElementById('tabMap');

  if(!withCoords.length){
    // Skrýt tab mapy — žádná GPS data
    if(mapTabBtn) mapTabBtn.style.display='none';
    // Přepnout z mapy jinam pokud je aktivní
    if(document.getElementById('panelMap')?.classList.contains('on')) tab('overview');
    return;
  }
  // Ukázat tab — jsou GPS data
  if(mapTabBtn) mapTabBtn.style.display='';

  const el=document.getElementById('mapContainer');
  if(!el) return;

  // Inicializace mapy — vždy uvnitř setTimeout aby měl container rozměry
  const initMap=()=>{
    if(!_map){
      _map=L.map('mapContainer').setView([50.08, 14.42], 5);
      L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{
        attribution:'© OpenStreetMap', maxZoom:18
      }).addTo(_map);
    } else {
      _mapMarkers.forEach(m=>m.remove());
      _mapMarkers=[];
      _map.invalidateSize();
    }

    if(!withCoords.length){
      const stats=document.getElementById('mapStats');
      if(stats){ stats.style.display='block'; stats.textContent='Žádná hesla se souřadnicemi'; }
      return;
    }

    withCoords.forEach(r=>{
      const {lat,lng}=r.coordinates_norm;
      if(!lat||!lng||Math.abs(lat)>90||Math.abs(lng)>180) return;
      const q=r._quality?.score||0;
      const color=q>=70?'#6ee7b7':q>=40?'#fbbf24':'#f87171';
      const icon=L.divIcon({
        className:'',
        html:`<div style="width:10px;height:10px;border-radius:50%;background:${color};border:2px solid rgba(0,0,0,.3)"></div>`,
        iconSize:[10,10], iconAnchor:[5,5]
      });
      const marker=L.marker([lat,lng],{icon})
        .addTo(_map)
        .bindPopup(`<b>${r.title||''}</b><br><small>${(r.intro||'').slice(0,120)}…</small><br><a href="${r.url}" target="_blank">Wikipedia →</a>`);
      _mapMarkers.push(marker);
    });

    if(_mapMarkers.length>0){
      const group=L.featureGroup(_mapMarkers);
      _map.fitBounds(group.getBounds().pad(0.1));
    }

    const stats=document.getElementById('mapStats');
    if(stats){ stats.style.display='block'; stats.textContent=`📍 ${withCoords.length} hesel se souřadnicemi`; }
  };

  // Pokud je map tab aktivní, hned inicializuj; jinak při přepnutí
  if(document.getElementById('panelMap')?.classList.contains('on')){
    setTimeout(initMap, 50);
  } else {
    // Uložit callback pro inicializaci až se tab otevře
    window._mapInitPending=initMap;
  }
}

/* ── TAG CLOUD ────────────────────────────────────────────────────────────── */
let _activeTag=null;
const TAG_COLORS=['#6ee7b7','#60a5fa','#fbbf24','#f87171','#a78bfa','#fb923c','#34d399','#818cf8','#f472b6','#38bdf8'];

function buildTagCloud(data){
  const tagCounts={};
  data.forEach(r=>{
    (r._tags||[]).forEach(t=>{ tagCounts[t]=(tagCounts[t]||0)+1; });
  });
  const sorted=Object.entries(tagCounts).sort((a,b)=>b[1]-a[1]);
  if(!sorted.length){
    document.getElementById('tagCloud').innerHTML=
      '<div style="font-family:var(--mono);font-size:11px;color:var(--text3)">Žádné tagy — spusť scraper s přepínačem --tags nebo zapni Auto-tagy v nastavení</div>';
    return;
  }
  const max=sorted[0][1];
  document.getElementById('tagCloud').innerHTML=sorted.map(([tag,cnt],i)=>{
    const size=Math.round(10+cnt/max*14);
    const color=TAG_COLORS[i%TAG_COLORS.length];
    return `<span class="tag-chip${_activeTag===tag?' active':''}"
      style="font-size:${size}px;background:${color}22;color:${color};border-color:${color}44"
      onclick="filterByTag('${esc(tag)}')">${esc(tag)} <small>${cnt}</small></span>`;
  }).join('');
}

let _activeTagFilter=null;  // filtruje podle r._tags přímo, ne fulltext

function filterByTag(tag){
  _activeTagFilter=_activeTagFilter===tag?null:tag;
  _activeTag=_activeTagFilter;  // pro highlight v tag cloudu
  _validIssueFilter=null;  // zrušit validační filtr
  buildTagCloud(allData);
  if(_activeTagFilter){
    tab('data');
    document.getElementById('dtQ').value='';
    renderTbl();
  } else {
    renderTbl();
  }
}

/* ── LIVE CHART ───────────────────────────────────────────────────────────── */
let _speedChart=null;
let _speedData=[];
let _lastDone=0, _lastTime=Date.now();

function initLiveChart(){
  const canvas=document.getElementById('liveSpeedChart');
  if(!canvas||!window.Chart) return;
  _speedData=[];
  if(_speedChart){ _speedChart.destroy(); _speedChart=null; }
  const cs=getComputedStyle(document.documentElement);
  const chartTick=cs.getPropertyValue('--chart-tick').trim()||'#4a5568';
  const chartGrid=cs.getPropertyValue('--chart-grid').trim()||'#2a2f3d';
  _speedChart=new Chart(canvas,{
    type:'line',
    data:{
      labels:[],
      datasets:[{
        data:[], borderColor:'#6ee7b7', borderWidth:1.5,
        fill:true, backgroundColor:'rgba(110,231,183,.1)',
        pointRadius:0, tension:0.3
      }]
    },
    options:{
      animation:false, responsive:false,
      plugins:{legend:{display:false}},
      scales:{
        x:{display:false},
        y:{display:true, min:0, ticks:{font:{size:9},color:chartTick},
           grid:{color:chartGrid}}
      }
    }
  });
  document.getElementById('liveChartWrap').classList.add('on');
}

function updateLiveChart(done){
  if(!_speedChart) return;
  const now=Date.now();
  const elapsed=(now-_lastTime)/1000;
  if(elapsed>=2){
    const rate=Math.round((_done_count_global-_lastDone)/elapsed*60);
    _lastDone=_done_count_global;
    _lastTime=now;
    _speedData.push(rate);
    if(_speedData.length>20) _speedData.shift();
    _speedChart.data.labels=_speedData.map((_,i)=>i);
    _speedChart.data.datasets[0].data=[..._speedData];
    _speedChart.update('none');
  }
}
let _done_count_global=0;

/* ── GLOBAL SEARCH ────────────────────────────────────────────────────────── */
let _globalIndex=[];  // [{title, intro, url, source}]
let _globalSearchTimer=null;

function buildGlobalIndex(){
  fetch('/list_outputs?s='+encodeURIComponent(getSessionId())).then(r=>r.json()).then(files=>{
    const names=Object.keys(files);
    _globalIndex=[];
    const loadNext=(i)=>{
      if(i>=names.length) return;
      fetch('/results?file='+encodeURIComponent(names[i]))
        .then(r=>r.json())
        .then(data=>{
          data.forEach(r=>{
            _globalIndex.push({title:r.title||'',intro:(r.intro||'').slice(0,200),url:r.url||'',source:names[i]});
          });
          loadNext(i+1);
        }).catch(()=>loadNext(i+1));
    };
    loadNext(0);
  });
}

function globalSearch(q){
  clearTimeout(_globalSearchTimer);
  const res=document.getElementById('globalResults');
  if(!q.trim()){ res.classList.remove('on'); return; }
  _globalSearchTimer=setTimeout(()=>{
    const ql=q.toLowerCase();
    const matches=_globalIndex.filter(r=>
      r.title.toLowerCase().includes(ql)||r.intro.toLowerCase().includes(ql)
    ).slice(0,20);

    if(!matches.length){ res.innerHTML='<div class="notif-empty">Nic nenalezeno</div>'; res.classList.add('on'); return; }

    // Group by source
    const bySource={};
    matches.forEach(m=>{ if(!bySource[m.source]) bySource[m.source]=[]; bySource[m.source].push(m); });

    res.innerHTML=Object.entries(bySource).map(([src,items])=>
      `<div class="gs-section">📄 ${esc(src)}</div>`+
      items.map(m=>`
        <div class="gs-item" data-url="${esc(m.url)}" data-source="${esc(m.source)}">
          <div class="gs-item-title">${esc(m.title)}</div>
          <div class="gs-item-snip">${esc(m.intro.slice(0,100))}…</div>
        </div>`).join('')
    ).join('');
    res.classList.add('on');
    // Event delegation — safe against apostrophes/quotes in names
    res.querySelectorAll('.gs-item').forEach(el=>{
      el.addEventListener('click',()=>gsJumpTo(el.dataset.url, el.dataset.source));
    });
  }, 200);
}

function gsJumpTo(url, source){
  document.getElementById('globalResults').classList.remove('on');
  document.getElementById('globalSearchInput').value='';
  // Načíst soubor a přejít na data tab
  loadData(source);
  tab('data');
  // Nastavit search na titul
  setTimeout(()=>{
    const item=_globalIndex.find(i=>i.url===url);
    if(item){ document.getElementById('dtQ').value=item.title; renderTbl(); }
  },300);
}

/* ── BULK SELECT ──────────────────────────────────────────────────────────── */
let _bulkSelected=new Set();
let _bulkMode=false;

function toggleBulkMode(){
  _bulkMode=!_bulkMode;
  document.getElementById('bulkToolbar').classList.toggle('on',_bulkMode);
  _bulkSelected.clear();
  renderTbl();
}

function toggleBulkRow(url, cb){
  if(cb.checked) _bulkSelected.add(url);
  else _bulkSelected.delete(url);
  document.getElementById('bulkCount').textContent=_bulkSelected.size+' vybráno';
}

function bulkExport(){
  if(!_bulkSelected.size){ clog('⚠ Nic nevybráno','warn'); return; }
  const selected=allData.filter(r=>_bulkSelected.has(r.url));
  const blob=new Blob([JSON.stringify(selected,null,2)],{type:'application/json'});
  const a=document.createElement('a'); a.href=URL.createObjectURL(blob);
  a.download='bulk_export.json'; a.click();
  clog(`✅ Exportováno ${selected.length} hesel`,'ok');
}

function bulkDelete(){
  if(!_bulkSelected.size){ clog('⚠ Nic nevybráno','warn'); return; }
  if(!confirm(`Odstranit ${_bulkSelected.size} hesel z tabulky? (JSON soubor se nemění)`)) return;
  allData=allData.filter(r=>!_bulkSelected.has(r.url));
  _bulkSelected.clear();
  document.getElementById('bulkCount').textContent='0 vybráno';
  renderTbl();
  clog(`🗑 Odstraněno z tabulky (původní soubor nezměněn)`,'warn');
}

/* ── UNDO/REDO ────────────────────────────────────────────────────────────── */
let _undoStack=[], _redoStack=[];
const UNDO_MAX=7;

function pushUndo(desc, snapshot){
  _undoStack.push({desc, snapshot:JSON.parse(JSON.stringify(snapshot))});
  if(_undoStack.length>UNDO_MAX) _undoStack.shift();
  _redoStack=[];
  updateUndoButtons();
}

function undo(){
  if(!_undoStack.length) return;
  const state=_undoStack.pop();
  _redoStack.push({desc:state.desc, snapshot:JSON.parse(JSON.stringify(allData))});
  allData=state.snapshot;
  renderTbl(); buildTagCloud(allData);
  clog(`↩ Undo: ${state.desc}`,'dim');
  updateUndoButtons();
}

function redo(){
  if(!_redoStack.length) return;
  const state=_redoStack.pop();
  _undoStack.push({desc:state.desc, snapshot:JSON.parse(JSON.stringify(allData))});
  allData=state.snapshot;
  renderTbl(); buildTagCloud(allData);
  clog(`↪ Redo: ${state.desc}`,'dim');
  updateUndoButtons();
}

function updateUndoButtons(){
  const u=document.getElementById('btnUndo'), r=document.getElementById('btnRedo');
  if(u) u.disabled=!_undoStack.length;
  if(r) r.disabled=!_redoStack.length;
}

/* ── VALIDATION ───────────────────────────────────────────────────────────── */
function buildValidation(data){
  const total=data.length;
  if(!total){ document.getElementById('validateContent').innerHTML='<div class="overview-empty"><div class="overview-icon">✅</div><div>Načti data nejdříve</div></div>'; return; }

  // Anomálie ze sanitizace
  const withAnomalies = data.filter(r=>r._anomalies&&r._anomalies.length);
  const errorRecs = withAnomalies.filter(r=>r._anomalies.some(a=>a.severity==='error'));
  const warnRecs  = withAnomalies.filter(r=>r._anomalies.some(a=>a.severity==='warn'));

  const issues=[
    {key:'no_intro',       label:'Bez perexu',             color:'var(--danger)'},
    {key:'short_intro',    label:'Krátký perex (<100 zn)', color:'var(--warn)'},
    {key:'no_infobox',     label:'Bez infoboxu',           color:'var(--warn)'},
    {key:'no_categories',  label:'Bez kategorií',          color:'var(--text3)'},
    {key:'errors',         label:'Chyby stažení',          color:'var(--danger)'},
    {key:'stubs',          label:'Stub články',            color:'var(--warn)'},
    {key:'low_quality',    label:'Nízké quality (<40)',    color:'var(--danger)'},
    {key:'duplicate_titles',label:'Duplicitní tituly',     color:'var(--danger)'},
  ];

  // Compute locally
  const counts={
    no_intro:        data.filter(r=>!(r.intro||'').trim()).length,
    short_intro:     data.filter(r=>{const l=(r.intro||'').length;return l>0&&l<100;}).length,
    no_infobox:      data.filter(r=>!r.infobox||!Object.keys(r.infobox).length).length,
    no_categories:   data.filter(r=>!(r.categories||[]).length).length,
    errors:          data.filter(r=>r.error).length,
    stubs:           data.filter(r=>r.is_stub).length,
    low_quality:     data.filter(r=>typeof r._quality==='object'&&(r._quality?.score||0)<40).length,
    duplicate_titles:Object.values(Object.fromEntries(data.map(r=>[r.title||'',0])
      .map(([t])=>[t,data.filter(r=>(r.title||'')===t).length])))
      .filter(n=>n>1).length,
  };

  const maxCount=Math.max(...Object.values(counts), 1);
  const avgQ=data.length?Math.round(data.reduce((s,r)=>s+(r._quality?.score||0),0)/data.length):0;

  document.getElementById('validateContent').innerHTML=`
    <div style="display:flex;gap:12px;margin-bottom:16px;flex-wrap:wrap">
      <div class="sg-card"><div class="sg-val">${total}</div><div class="sg-lbl">Celkem hesel</div></div>
      <div class="sg-card"><div class="sg-val" style="color:${avgQ>=70?'var(--accent)':avgQ>=40?'var(--warn)':'var(--danger)'}">${avgQ}</div><div class="sg-lbl">Průměrné Q skóre</div></div>
      <div class="sg-card"><div class="sg-val" style="color:var(--accent)">${data.filter(r=>(r._quality?.score||0)>=70).length}</div><div class="sg-lbl">Vysoká kvalita</div></div>
    </div>
    <div style="font-family:var(--mono);font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:var(--text3);margin-bottom:10px">Problémy</div>
    ${issues.map(iss=>{
      const cnt=counts[iss.key]||0;
      const pct=total?Math.round(cnt/total*100):0;
      return `<div class="valid-issue">
        <div class="valid-count" style="color:${cnt>0?iss.color:'var(--text3)'}">${cnt}</div>
        <div class="valid-label">${iss.label}</div>
        <div class="valid-bar-wrap"><div class="valid-bar" style="width:${Math.max(pct,cnt>0?2:0)}%;background:${iss.color}"></div></div>
        <div style="font-family:var(--mono);font-size:10px;color:var(--text3);min-width:32px;text-align:right">${pct}%</div>
        ${cnt>0?`<button class="btn-sm" style="margin-left:8px" onclick="filterValidIssue('${iss.key}')" title="Filtrovat tabulku — zobrazit jen tato hesla">Filtr</button>`:''}
      </div>`;
    }).join('')}

    ${withAnomalies.length ? `<div style="margin-top:16px">
      <div style="font-family:var(--mono);font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:var(--text3);margin-bottom:8px">🔍 Sanitizace dat</div>
      ${errorRecs.length ? `<div class="val-row" style="cursor:pointer" onclick="filterValidIssue('anom_error')">
        <span style="color:var(--danger);font-family:var(--mono);font-size:11px">● Chyby: ${errorRecs.length} hesel</span>
        <button class="btn-sm" style="margin-left:auto">Filtr</button></div>` : ''}
      ${warnRecs.length ? `<div class="val-row" style="cursor:pointer;margin-top:4px" onclick="filterValidIssue('anom_warn')">
        <span style="color:var(--warn);font-family:var(--mono);font-size:11px">▲ Varování: ${warnRecs.length} hesel</span>
        <button class="btn-sm" style="margin-left:auto">Filtr</button></div>` : ''}
      <div style="font-family:var(--mono);font-size:10px;color:var(--text3);margin-top:6px">
        ${Object.entries(withAnomalies.reduce((a,r)=>{(r._anomalies||[]).forEach(x=>{a[x.rule]=(a[x.rule]||0)+1;});return a;},{}))
          .sort((a,b)=>b[1]-a[1]).slice(0,6)
          .map(([k,n])=>`<span style="margin-right:8px">${k}: ${n}</span>`).join('')}
      </div>
    </div>` : ''}
  `;
}

function filterValidIssue(key){
  tab('data');
  // Reset předchozích filtrů
  _activeFilters.clear();
  document.querySelectorAll('.rf-chip').forEach(c=>c.classList.remove('on'));
  document.getElementById('dtQ').value='';

  if(key==='anomalies'||key==='anom_error'||key==='anom_warn'){
    _validIssueFilter=key;
    tab('data');
    _activeFilters.clear();
    document.querySelectorAll('.rf-chip').forEach(c=>c.classList.remove('on'));
    renderTbl();
    return;
  }
  if(key==='errors'){
    toggleFilter('error');
  } else if(key==='no_infobox'){
    // Invertovat chip infobox — zobrazit hesla BEZ infoboxu
    // Chip 'infobox' normálně filtruje "s infoboxem", takže musíme custom filtr
    _validIssueFilter=key;
    renderTbl();
  } else if(key==='no_categories'){
    _validIssueFilter=key;
    renderTbl();
  } else if(key==='stubs'){
    _validIssueFilter=key;
    renderTbl();
  } else if(key==='low_quality'){
    _validIssueFilter=key;
    renderTbl();
  } else if(key==='no_intro'||key==='short_intro'){
    _validIssueFilter=key;
    renderTbl();
  } else {
    renderTbl();
  }
}
let _validIssueFilter=null;

/* ── WORKERS & API SETTINGS ───────────────────────────────────────────────── */
function getWorkers(){ return parseInt(document.getElementById('workers')?.value||'1'); }
function getApiMode(){ return document.getElementById('apiMode')?.checked||false; }
function getWikidata(){ return document.getElementById('wikidata')?.checked||false; }
function getAutoTags(){ return document.getElementById('autoTags')?.checked||false; }
function getIncremental(){ return document.getElementById('incremental')?.checked||false; }

/* ── REVIEW STATE PERSISTENCE ─────────────────────────────────────────────── */
const REVIEW_STATE_KEY='wikiReviewState';

function saveReviewState(){
  if(!_reviewPending) return;
  const state={
    output: window._runParams?.output||'',
    pending: _reviewPending,
    cols: _cols,
    colSearch: _colSearch,
    savedAt: new Date().toISOString()
  };
  try{ localStorage.setItem(REVIEW_STATE_KEY, JSON.stringify(state)); }catch(e){}
}

function clearReviewState(){
  localStorage.removeItem(REVIEW_STATE_KEY);
}

function checkSavedReview(){
  const raw=localStorage.getItem(REVIEW_STATE_KEY);
  if(!raw) return;
  try{
    const state=JSON.parse(raw);
    const output=state.output||'—';
    const saved=new Date(state.savedAt).toLocaleString('cs-CZ');
    const total=Object.values(state.pending?.cat_articles||{}).reduce((s,v)=>s+v.length,0);
    const banner=document.getElementById('savedReviewBanner');
    if(banner){
      banner.style.display='flex';
      document.getElementById('srvText').textContent=
        `Uložený review: "${output}" · ${total} hesel · ${saved}`;
    }
  }catch(e){}
}

function resumeSavedReview(){
  const raw=localStorage.getItem(REVIEW_STATE_KEY);
  if(!raw) return;
  try{
    const state=JSON.parse(raw);
    _reviewPending=state.pending;
    _cols=state.cols;
    _colSearch=state.colSearch||state.cols.map(()=>'');
    window._runParams=window._runParams||{};
    window._runParams.output=state.output;
    const b=document.getElementById('savedReviewBanner');
    if(b) b.style.display='none';
    document.getElementById('reviewModal').classList.add('on');
    document.getElementById('reviewReveal').classList.remove('on');
    document.getElementById('reviewColumns').style.display='';
    document.getElementById('reviewFooter').style.display='';
    document.getElementById('reviewSub').textContent=
      `${Object.keys(_reviewPending.cat_articles||{}).length} kategorií — obnovený stav`;
    renderReview();
    setStep(2);
  }catch(e){ clog('❌ Nelze obnovit uložený review: '+e,'err'); }
}

/* ── PROJECTS (TABS) ──────────────────────────────────────────────────────── */
const PROJECTS_KEY='wikiProjects';
let _projects=[];
let _activeProject=null;

function loadProjects(){
  try{ return JSON.parse(localStorage.getItem(PROJECTS_KEY)||'[]'); }
  catch(e){ return []; }
}
function saveProjects(){
  try{ localStorage.setItem(PROJECTS_KEY, JSON.stringify(_projects)); }catch(e){}
}

function readFormToProject(){
  return {
    url:    document.getElementById('url').value,
    output: document.getElementById('output').value,
    fields: getFields(),
    depth:  document.getElementById('depth').value,
    limit:  document.getElementById('limit').value,
    delay:  document.getElementById('delay').value,
    format: document.getElementById('format').value,
  };
}

function writeProjectToForm(p){
  if(!p) return;
  document.getElementById('url').value=p.url||'';
  document.getElementById('output').value=p.output||'wiki_data';
  document.getElementById('depth').value=p.depth||5;
  document.getElementById('limit').value=p.limit||0;
  document.getElementById('delay').value=p.delay||0.5;
  document.getElementById('format').value=p.format||'both';
  if(p.fields){
    const fl=p.fields.split(',');
    FIELDS.forEach(f=>{
      const on=fl.includes(f.id);
      fState[f.id]=on;
      const el=document.getElementById('ft_'+f.id);
      const fb=document.getElementById('fb_'+f.id);
      if(el) el.classList.toggle('on',on);
      if(fb) fb.textContent=on?'✓':'';
    });
  }
  checkCP();
}

function switchProject(id){
  // NEPŘERUŠOVAT běžící scraping — jen uložit stav a přepnout UI
  // evtSrc zůstane otevřený, job poběží dál na pozadí

  // Uložit aktuální stav do aktivního projektu
  if(_activeProject){
    const idx=_projects.findIndex(p=>p.id===_activeProject);
    if(idx>=0) _projects[idx]={..._projects[idx],...readFormToProject()};
  }

  _activeProject=id;
  const p=_projects.find(p=>p.id===id);
  if(p) writeProjectToForm(p);
  saveProjects();
  renderProjectTabs();

  // Reset UI stavu pro nový projekt (job na pozadí poběží dál)
  allData=[];
  _completedJobs=[];
  _activeTagFilter=null;
  _validIssueFilter=null;
  _activeTag=null;
  _bulkSelected=new Set();
  _undoStack=[]; _redoStack=[];
  // NERESETUJEME _phase2Running — job může běžet na pozadí
  const _hadRunning=_phase2Running;
  setStep(0); setStatus('idle');
  setPhase(1,'Čeká na spuštění…','');
  document.getElementById('btnRun').classList.remove('hidden');
  document.getElementById('btnStop').classList.add('hidden');
  document.getElementById('rlChip').classList.remove('on');

  // Stats
  ['stFound','stDone','stEta','stCats','stTime'].forEach(i=>{
    const el=document.getElementById(i); if(el) el.textContent='—';
  });

  // Progress bars
  const pb1=document.getElementById('pb1');
  if(pb1){ pb1.style.width='0%'; pb1.classList.remove('indeterminate'); }
  const pb2=document.getElementById('pb2'); if(pb2) pb2.style.width='0%';
  const pp1=document.getElementById('pp1'); if(pp1) pp1.textContent='—';
  const pp2=document.getElementById('pp2'); if(pp2) pp2.textContent='—';

  // Boxes
  document.getElementById('exportBox').classList.remove('on');
  document.getElementById('resumeBox').classList.remove('on');

  // Overview tab
  const sc=document.getElementById('summaryCards'); if(sc){ sc.classList.add('hidden'); sc.innerHTML=''; }
  const oe=document.getElementById('overviewEmpty'); if(oe) oe.classList.remove('hidden');

  // Data tab
  const de=document.getElementById('dataEmpty'); if(de) de.classList.remove('hidden');
  const db=document.getElementById('dtBody'); if(db) db.innerHTML='';
  document.getElementById('dtCnt').textContent='';
  document.getElementById('dataCnt').textContent='';
  const dtQ=document.getElementById('dtQ'); if(dtQ) dtQ.value='';
  _activeFilters.clear();
  document.querySelectorAll('.rf-chip').forEach(c=>c.classList.remove('on'));

  // Viz tab
  const viz=document.getElementById('vizArea');
  if(viz) viz.innerHTML='<div class="overview-empty" style="width:100%"><div class="overview-icon">📊</div><div>Vizualizace se zobrazí po dokončení</div></div>';

  // Přejít na přehled
  tab('overview');

  // Scroll sidebar na začátek
  const sidebar=document.querySelector('.sidebar-scroll');
  if(sidebar) sidebar.scrollTop=0;

  // Zkontrolovat checkpoint pro nový projekt
  checkCP();

  clog(`── Projekt: ${p?.name||id} ──`,'dim');
}

function closeProject(id){
  if(_projects.length<=1) return;
  _projects=_projects.filter(p=>p.id!==id);
  saveProjects();
  if(_activeProject===id) switchProject(_projects[0].id);
  else renderProjectTabs();
}

function addProject(){
  if(_activeProject){
    const idx=_projects.findIndex(p=>p.id===_activeProject);
    if(idx>=0) _projects[idx]={..._projects[idx],...readFormToProject()};
  }
  const id='proj_'+Date.now();
  _projects.push({id, name:'Projekt '+(_projects.length+1),
    url:'', output:'wiki_data',
    fields:'title,intro,infobox,categories,coordinates',
    depth:5, limit:0, delay:0.5, format:'both'});
  saveProjects();
  switchProject(id);
}

function renameProject(id, el){
  const old=el.textContent;
  el.contentEditable=true; el.focus();
  const r=document.createRange(); r.selectNodeContents(el);
  window.getSelection().removeAllRanges(); window.getSelection().addRange(r);
  el.onblur=()=>{
    el.contentEditable=false;
    const n=el.textContent.trim()||old; el.textContent=n;
    const p=_projects.find(p=>p.id===id);
    if(p){ p.name=n; saveProjects(); }
  };
  el.onkeydown=e=>{ if(e.key==='Enter'){e.preventDefault();el.blur();} };
}

function renderProjectTabs(){
  const bar=document.getElementById('projectTabs');
  if(!bar) return;
  bar.innerHTML=_projects.map(p=>`
    <div class="proj-tab${p.id===_activeProject?' active':''}" onclick="switchProject('${p.id}')">
      <span class="proj-name" ondblclick="renameProject('${p.id}',this)" title="Dvojklik pro přejmenování">${esc(p.name)}</span>
      ${_projects.length>1?`<span class="proj-close" onclick="event.stopPropagation();closeProject('${p.id}')" title="Zavřít projekt">×</span>`:''}
    </div>`).join('')+
    `<button class="proj-add" onclick="addProject()" title="Nový projekt">+</button>`;
}

function initProjects(){
  _projects=loadProjects();
  if(!_projects.length){
    const p={id:'proj_default',name:'Projekt 1',url:'',output:'wiki_data',
             fields:getFields(),depth:5,limit:0,delay:0.5,format:'both'};
    _projects=[p]; _activeProject=p.id;
  } else {
    _activeProject=_projects[0].id;
    writeProjectToForm(_projects[0]);
  }
  saveProjects(); renderProjectTabs();
}

function toggleTheme(){
  const light=document.documentElement.classList.toggle('light');
  document.getElementById('themeBtn').textContent=light?'🌑':'🌙';
  localStorage.setItem('wikiTheme', light?'light':'dark');
  // Aktualizovat theme-color meta tag
  const metaTc=document.getElementById('metaThemeColor');
  if(metaTc) metaTc.content=light?'#f7f6f3':'#0f1117';
  // Přebuildnout charty aby použily nové CSS barvy
  if(allData.length) buildViz(allData);
}
function initTheme(){
  const saved=localStorage.getItem('wikiTheme');
  if(saved==='light'){
    document.documentElement.classList.add('light');
    const metaTc=document.getElementById('metaThemeColor');
    if(metaTc) metaTc.content='#f7f6f3';
    document.getElementById('themeBtn').textContent='🌑';
  }
}

/* ── HISTORY ──────────────────────────────────────────────────────────────── */
const HISTORY_KEY='wikiScraperHistory';
const HISTORY_MAX=8;

function loadHistory(){
  try{ return JSON.parse(localStorage.getItem(HISTORY_KEY)||'[]'); }
  catch(e){ return []; }
}

function saveToHistory(entry){
  // entry = {url, name, output, hesel, duration_s, date}
  let h=loadHistory();
  // Deduplikace — stejná URL + output = přepsat
  h=h.filter(e=>!(e.url===entry.url && e.output===entry.output));
  h.unshift(entry);
  if(h.length>HISTORY_MAX) h=h.slice(0,HISTORY_MAX);
  localStorage.setItem(HISTORY_KEY, JSON.stringify(h));
  renderHistory();
}

function fmtDuration(s){
  if(s<60) return Math.round(s)+'s';
  if(s<3600) return Math.floor(s/60)+'m '+Math.round(s%60)+'s';
  return Math.floor(s/3600)+'h '+Math.floor((s%3600)/60)+'m';
}

const HISTORY_SHOW_DEFAULT = 3;
let _historyExpanded = false;

function renderHistory(){
  const h=loadHistory();
  const sec=document.getElementById('historySection');
  const list=document.getElementById('historyList');
  if(!h.length){ sec.style.display='none'; return; }
  sec.style.display='block';
  const visible=_historyExpanded ? h : h.slice(0,HISTORY_SHOW_DEFAULT);
  list.innerHTML=visible.map((e,i)=>`
    <div class="ex-btn" data-hi="${i}" style="cursor:default">
      <span class="ex-icon" style="cursor:default">🕐</span>
      <span class="ex-info" data-load="${i}" style="flex:1;min-width:0;cursor:pointer" title="Kliknutím načíst URL a název souboru">
        <span class="ex-name">${esc(e.name||e.output)}</span>
        <span class="ex-url">${e.hesel} hesel · ${fmtDuration(e.duration_s)} · ${e.date}</span>
      </span>
      <button class="hist-del" data-del="${i}" title="Smazat ze seznamu"
        style="background:transparent;border:none;color:var(--text3);cursor:pointer;padding:2px 6px;font-size:14px;flex-shrink:0;line-height:1;border-radius:3px;transition:color .15s"
        onmouseover="this.style.color='var(--danger)'" onmouseout="this.style.color=''">×</button>
    </div>`).join('') +
    (h.length > HISTORY_SHOW_DEFAULT
      ? `<span class="history-more-btn" onclick="toggleHistoryExpand()">${
          _historyExpanded
            ? '▴ Skrýt'
            : '▾ Zobrazit vše ('+h.length+')'
        }</span>`
      : '');
  // Načíst URL kliknutím na info část
  list.querySelectorAll('[data-load]').forEach(info=>{
    info.addEventListener('click',()=>{
      const entry=h[+info.dataset.load];
      if(!entry) return;
      document.getElementById('url').value=entry.url;
      document.getElementById('output').value=entry.output;
      checkCP();
      // Feedback
      const row=info.closest('.ex-btn');
      if(row){ row.style.borderColor='var(--accent2)'; setTimeout(()=>row.style.borderColor='',600); }
    });
  });
  // Smazat položku
  list.querySelectorAll('[data-del]').forEach(btn=>{
    btn.addEventListener('click',ev=>{
      ev.stopPropagation();
      const idx=+btn.dataset.del;
      const hist=loadHistory();
      hist.splice(idx,1);
      localStorage.setItem(HISTORY_KEY, JSON.stringify(hist));
      renderHistory();
    });
  });
}

function toggleHistoryExpand(){
  _historyExpanded=!_historyExpanded;
  renderHistory();
}

/* ── DEDUP ────────────────────────────────────────────────────────────────── */
function dedupCatArticles(catArticles){
  // Každé URL jen jednou — u první kategorie kde se vyskytlo
  const seen=new Set();
  const result={};
  let removed=0;
  for(const [cat, urls] of Object.entries(catArticles)){
    const uniq=urls.filter(u=>{ if(seen.has(u)){removed++;return false;} seen.add(u);return true; });
    if(uniq.length) result[cat]=uniq;
  }
  return {result, removed};
}

/* ── SIZE ESTIMATE ────────────────────────────────────────────────────────── */
// Průměrné velikosti hesla v bytech podle vybraných polí
const FIELD_SIZES={title:60, intro:400, infobox:300, categories:80, coordinates:30, full_text:3000, sections:2500, links:200, images:600, tables:800};

function estimateSize(hesel, fields){
  const fieldList=fields.split(',');
  const bytesPerArticle=fieldList.reduce((s,f)=>s+(FIELD_SIZES[f]||50),0)+50; // +50 JSON overhead
  const jsonBytes=hesel*bytesPerArticle;
  const csvBytes=hesel*(bytesPerArticle*0.7); // CSV bývá menší
  const xlsxBytes=hesel*(bytesPerArticle*0.5)+10000; // XLSX s hlavičkou
  return {json:jsonBytes, csv:csvBytes, xlsx:xlsxBytes};
}

function fmtBytes(b){
  if(b<1024) return b+'B';
  if(b<1024*1024) return (b/1024).toFixed(0)+'KB';
  return (b/1024/1024).toFixed(1)+'MB';
}

function updateSizeEstimate(hesel){
  const fields=getFields();
  const est=estimateSize(hesel, fields);
  const el=document.getElementById('sizeEstimate');
  if(el) el.textContent=`~ ${fmtBytes(est.json)} JSON · ${fmtBytes(est.csv)} CSV · ${fmtBytes(est.xlsx)} XLSX`;
}

function updateReviewSizeEst(){
  const cats=_reviewPending?.cat_articles||{};
  let total=0;
  _cols.forEach(col=>Object.entries(col.cats).filter(([,on])=>on).forEach(([c])=>{ total+=(cats[c]?.length||0); }));
  const fields=getFields();
  const est=estimateSize(total, fields);
  const el=document.getElementById('reviewSizeEst');
  if(el) el.textContent=`~ ${fmtBytes(est.json)} JSON · ${fmtBytes(est.xlsx)} XLSX`;
}

/* ── NOTIFICATIONS ────────────────────────────────────────────────────────── */
let _notifPermission='default';
function initNotifications(){
  if('Notification' in window){
    _notifPermission=Notification.permission;
    // requestPermission() NELZE volat automaticky pri nacteni stranky —
    // prohlizec to dovoli jen z user gestu (klik). Zadame v go() pri prvnim spusteni.
  }
}

function requestNotifPermission(){
  if(!('Notification' in window)) return;
  if(Notification.permission==='default'){
    Notification.requestPermission().then(p=>{ _notifPermission=p; });
  }
}

function sendNotification(title, body){
  addNotif(title, body, 'ok');
  if(_notifPermission==='granted'){
    const n=new Notification(title,{body, icon:'data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32"><text y="28" font-size="28">📚</text></svg>'});
    setTimeout(()=>n.close(), 6000);
  }
}

/* ── KEYBOARD SHORTCUTS ───────────────────────────────────────────────────── */
function initKeyboard(){
  document.addEventListener('keydown', e=>{
    const modal=document.getElementById('reviewModal');
    const modalOpen=modal.classList.contains('on');

    // Esc — zavřít modal
    if(e.key==='Escape' && modalOpen){ closeReview(); return; }
    if(e.key==='Escape'){ document.getElementById('discardModal').style.display='none'; }

    // Ctrl+Enter — potvrdit modal
    if((e.ctrlKey||e.metaKey) && e.key==='Enter' && modalOpen){
      e.preventDefault();
      startPhase2(); return;
    }

    // Ctrl+Z undo, Ctrl+Y redo
    if((e.ctrlKey||e.metaKey) && e.key==='z' && !e.shiftKey){
      e.preventDefault(); undo(); return;
    }
    if((e.ctrlKey||e.metaKey) && (e.key==='y'||(e.key==='z'&&e.shiftKey))){
      e.preventDefault(); redo(); return;
    }

    // Ignore if typing in input
    if(['INPUT','TEXTAREA','SELECT'].includes(document.activeElement?.tagName)) return;

    // Space — spustit/zastavit
    if(e.key===' ' && !modalOpen){
      e.preventDefault();
      const runBtn=document.getElementById('btnRun');
      const stopBtn=document.getElementById('btnStop');
      if(!runBtn.classList.contains('hidden')) go(false);
      else if(!stopBtn.classList.contains('hidden')) stop();
      return;
    }

    // Ctrl+F — focus search v tabulce dat
    if((e.ctrlKey||e.metaKey) && e.key==='f'){
      const dtQ=document.getElementById('dtQ');
      if(dtQ && document.getElementById('panelData').classList.contains('on')){
        e.preventDefault();
        dtQ.focus();
      }
      return;
    }
  });
}

/* ── SIDEBAR SPLIT PANE (drag resize) ────────────────────────────────────── */
function toggleSidebar(){
  const body=document.getElementById('bodyGrid');
  const btn=document.getElementById('sidebarToggleBtn');
  const collapsed=body.classList.toggle('sidebar-collapsed');
  btn.textContent=collapsed?'▶':'◀';
  btn.title=collapsed?'Zobrazit panel':'Skrýt panel';
  // Smazat inline left styl aby ho přebrala CSS třída
  btn.style.left='';
  if(_map) setTimeout(()=>_map.invalidateSize(),320);
}

function initSidebarResize(){
  const btn=document.getElementById('sidebarToggleBtn');
  const body=document.getElementById('bodyGrid');
  if(!btn||!body) return;  // guard — element neexistuje
  let dragging=false, startX=0, startW=0;

  btn.addEventListener('mousedown',e=>{
    if(body.classList.contains('sidebar-collapsed')){ toggleSidebar(); return; }
    dragging=true; startX=e.clientX;
    startW=parseInt(getComputedStyle(document.documentElement)
      .getPropertyValue('--sidebar-w'))||320;
    // Vypnout transition při tažení pro plynulost
    body.style.transition='none';
    document.body.style.cursor='col-resize';
    document.body.style.userSelect='none';
    e.preventDefault();
  });
  document.addEventListener('mousemove',e=>{
    if(!dragging) return;
    const w=Math.max(180, Math.min(600, startW+(e.clientX-startX)));
    // Aktualizovat jen CSS proměnnou — button pozici řeší CSS
    document.documentElement.style.setProperty('--sidebar-w', w+'px');
  });
  document.addEventListener('mouseup',()=>{
    if(!dragging) return;
    dragging=false;
    body.style.transition='';
    document.body.style.cursor='';
    document.body.style.userSelect='';
    if(_map) _map.invalidateSize();
  });
}

/* ── COLUMN HIDER ─────────────────────────────────────────────────────────── */
const COL_DEFS=[
  {id:'c0', label:'#'},
  {id:'c1', label:'Název'},
  {id:'c2', label:'Perex'},
  {id:'c3', label:'Kategorie'},
  {id:'c4', label:'Délka'},
  {id:'c5', label:'Q skóre'},
  {id:'c6', label:'Infobox'},
  {id:'c7', label:'Kopír.'},
];
const _colVisible={c0:true,c1:true,c2:true,c3:true,c4:true,c5:true,c6:true,c7:true};

function toggleColHiderPanel(){
  const panel=document.getElementById('colHiderPanel');
  if(!panel.classList.contains('on')){
    panel.innerHTML=COL_DEFS.map(c=>`
      <label class="col-hider-row">
        <input type="checkbox" ${_colVisible[c.id]?'checked':''}
          onchange="setColVisible('${c.id}',this.checked)">
        ${c.label}
      </label>`).join('');
    panel.classList.add('on');
  } else {
    panel.classList.remove('on');
  }
}

function setColVisible(id, visible){
  _colVisible[id]=visible;
  // Přidat/odebrat CSS rule
  let styleEl=document.getElementById('colHiderStyle');
  if(!styleEl){
    styleEl=document.createElement('style');
    styleEl.id='colHiderStyle';
    document.head.appendChild(styleEl);
  }
  const hidden=Object.entries(_colVisible).filter(([,v])=>!v).map(([k])=>`.col-${k}`);
  styleEl.textContent=hidden.length ? `${hidden.join(',')}{display:none}` : '';
}

// Zavřít panel kliknutím mimo
document.addEventListener('click',e=>{
  const wrap=document.querySelector('.col-hider-wrap');
  if(wrap && !wrap.contains(e.target))
    document.getElementById('colHiderPanel')?.classList.remove('on');
});

/* ── TABLE KEYBOARD NAVIGATION ────────────────────────────────────────────── */
let _tblFocusIdx=null;

function tblRowClick(i){
  _tblFocusIdx=i;
  _renderVirtualRows();
}

function initTblKeyboard(){
  const wrap=document.querySelector('.dt-wrap');
  if(!wrap) return;
  wrap.setAttribute('tabindex','0');
  wrap.addEventListener('keydown',e=>{
    if(!_vtRows.length) return;
    if(e.key==='ArrowDown'){
      e.preventDefault();
      _tblFocusIdx=(_tblFocusIdx===null)?0:Math.min(_tblFocusIdx+1,_vtRows.length-1);
      _scrollTblFocused();
    } else if(e.key==='ArrowUp'){
      e.preventDefault();
      _tblFocusIdx=(_tblFocusIdx===null)?0:Math.max(_tblFocusIdx-1,0);
      _scrollTblFocused();
    } else if(e.key==='Enter' && _tblFocusIdx!==null){
      e.preventDefault();
      const r=_vtRows[_tblFocusIdx];
      if(r?.url) window.open(r.url,'_blank');
    }
  });
}

function _scrollTblFocused(){
  if(_tblFocusIdx===null) return;
  _renderVirtualRows();
  // Zajistit viditelnost — scroll pokud mimo viewport
  const wrap=document.querySelector('.dt-wrap');
  if(!wrap) return;
  const rowTop=_tblFocusIdx*ROW_HEIGHT;
  const rowBot=rowTop+ROW_HEIGHT;
  if(rowTop < wrap.scrollTop) wrap.scrollTop=rowTop-4;
  else if(rowBot > wrap.scrollTop+wrap.clientHeight) wrap.scrollTop=rowBot-wrap.clientHeight+4;
}

/* ── REVIEW UNDO/REDO ─────────────────────────────────────────────────────── */
let _reviewUndoStack=[], _reviewRedoStack=[];
const REVIEW_UNDO_MAX=30;

function _snapshotReview(){
  return JSON.parse(JSON.stringify({cols:_cols, colSearch:_colSearch}));
}

function pushReviewUndo(desc){
  _reviewUndoStack.push({desc, snap:_snapshotReview()});
  if(_reviewUndoStack.length>REVIEW_UNDO_MAX) _reviewUndoStack.shift();
  _reviewRedoStack=[];
  _updateReviewUndoUI();
}

function undoReview(){
  if(!_reviewUndoStack.length) return;
  _reviewRedoStack.push({desc:_reviewUndoStack[_reviewUndoStack.length-1].desc, snap:_snapshotReview()});
  const state=_reviewUndoStack.pop();
  _applyReviewSnap(state.snap);
  _updateReviewUndoUI();
}

function redoReview(){
  if(!_reviewRedoStack.length) return;
  _reviewUndoStack.push({desc:_reviewRedoStack[_reviewRedoStack.length-1].desc, snap:_snapshotReview()});
  const state=_reviewRedoStack.pop();
  _applyReviewSnap(state.snap);
  _updateReviewUndoUI();
}

function _applyReviewSnap(snap){
  _cols=snap.cols; _colSearch=snap.colSearch;
  renderReview(); updateSummary();
}

function _updateReviewUndoUI(){
  const u=document.getElementById('btnReviewUndo');
  const r=document.getElementById('btnReviewRedo');
  const lbl=document.getElementById('reviewUndoLabel');
  const bar=document.getElementById('reviewUndoBar');
  if(u) u.disabled=!_reviewUndoStack.length;
  if(r) r.disabled=!_reviewRedoStack.length;
  const total=_reviewUndoStack.length+_reviewRedoStack.length;
  if(bar) bar.classList.toggle('on', total>0);
  if(lbl){
    const last=_reviewUndoStack[_reviewUndoStack.length-1];
    lbl.textContent=last?'↩ '+last.desc:'';
  }
}

/* ── REVIEW COLUMN SPLIT PANE ────────────────────────────────────────────── */
function addColResizers(container){
  const cols=container.querySelectorAll('.review-col');
  cols.forEach((col,i)=>{
    if(i===cols.length-1) return;
    const resizer=document.createElement('div');
    resizer.className='col-resizer';
    col.after(resizer);
    let dragging=false, startX=0, startW=0, nextW=0;
    const nextCol=col.nextElementSibling?.nextElementSibling; // přeskočit resizer
    resizer.addEventListener('mousedown',e=>{
      dragging=true; startX=e.clientX;
      startW=col.getBoundingClientRect().width;
      nextW=nextCol?nextCol.getBoundingClientRect().width:0;
      resizer.classList.add('active');
      document.body.style.cursor='col-resize';
      document.body.style.userSelect='none';
      e.preventDefault();
    });
    document.addEventListener('mousemove',e=>{
      if(!dragging) return;
      const diff=e.clientX-startX;
      const newW=Math.max(160, startW+diff);
      col.style.flex='none'; col.style.width=newW+'px';
      if(nextCol){ nextCol.style.flex='none'; nextCol.style.width=Math.max(160,nextW-diff)+'px'; }
    });
    document.addEventListener('mouseup',()=>{
      if(!dragging) return;
      dragging=false; resizer.classList.remove('active');
      document.body.style.cursor=''; document.body.style.userSelect='';
    });
  });
}

/* ── TITLE PROGRESS ───────────────────────────────────────────────────────── */
function setTitleProgress(pct){
  if(pct===null){ document.title='Wiki Scraper'; return; }
  document.title=`[${pct}%] WikiScraper`;
}

/* ── WEB WORKER PRO FILTROVÁNÍ ────────────────────────────────────────────── */
// Worker kód jako inline Blob — žádný extra soubor
const _workerCode=`
self.onmessage=function(e){
  const {rows, q, ftMode, activeFilters, validIssueFilter, activeTagFilter, sk, sasc}=e.data;

  let res=rows;

  // Fulltextový filtr
  if(q){
    const ql=q.toLowerCase();
    if(!ftMode){
      res=res.filter(r=>{
        const t=(r.title||'').toLowerCase();
        const intro=(r.intro||'').toLowerCase();
        const cats=((r.categories||[]).join(' ')).toLowerCase();
        const ib=Object.values(r.infobox||{}).join(' ').toLowerCase();
        return t.includes(ql)||intro.includes(ql)||cats.includes(ql)||ib.includes(ql);
      });
    } else {
      // AND/OR/NOT
      const orParts=q.split(/\\bOR\\b/i).map(s=>s.trim()).filter(Boolean);
      res=res.filter(r=>{
        const text=[r.title||'',r.intro||'',(r.categories||[]).join(' '),
          Object.values(r.infobox||{}).join(' ')].join(' ').toLowerCase();
        return orParts.some(op=>{
          const ands=op.split(/\\bAND\\b/i).map(s=>s.trim()).filter(Boolean);
          return ands.every(ap=>{
            const nots=ap.split(/\\bNOT\\b/i).map(s=>s.trim()).filter(Boolean);
            const must=nots[0]?.replace(/^"|"$/g,'').toLowerCase()||'';
            return text.includes(must)&&nots.slice(1).every(n=>!text.includes(n.replace(/^"|"$/g,'').toLowerCase()));
          });
        });
      });
    }
  }

  // Filtry
  if(activeFilters.includes('infobox'))  res=res.filter(r=>r.infobox&&Object.keys(r.infobox).length>0);
  if(activeFilters.includes('coords'))   res=res.filter(r=>r.coordinates);
  if(activeFilters.includes('fulltext')) res=res.filter(r=>r.full_text||r.sections);
  if(activeFilters.includes('error'))    res=res.filter(r=>r.error);
  if(activeFilters.includes('images'))   res=res.filter(r=>r.images&&r.images.length>0);
  if(activeFilters.includes('tables'))   res=res.filter(r=>r.tables&&r.tables.length>0);

  // Tag filtr — přesná shoda v _tags poli
  if(activeTagFilter) res=res.filter(r=>(r._tags||[]).includes(activeTagFilter));

  if(validIssueFilter){
    if(validIssueFilter==='no_infobox')     res=res.filter(r=>!r.infobox||!Object.keys(r.infobox).length);
    else if(validIssueFilter==='no_categories') res=res.filter(r=>!(r.categories||[]).length);
    else if(validIssueFilter==='anomalies')  res=res.filter(r=>r._anomalies&&r._anomalies.length>0);
    else if(validIssueFilter==='anom_error') res=res.filter(r=>r._anomalies&&r._anomalies.some(a=>a.severity==='error'));
    else if(validIssueFilter==='anom_warn')  res=res.filter(r=>r._anomalies&&r._anomalies.some(a=>a.severity==='warn'));
    else if(validIssueFilter==='stubs')     res=res.filter(r=>r.is_stub);
    else if(validIssueFilter==='low_quality') res=res.filter(r=>typeof r._quality==='object'&&(r._quality?.score||0)<40);
    else if(validIssueFilter==='no_intro')  res=res.filter(r=>!(r.intro||'').trim());
    else if(validIssueFilter==='short_intro') res=res.filter(r=>{const l=(r.intro||'').length;return l>0&&l<100;});
  }

  // Řazení
  if(sk){
    res=[...res].sort((a,b)=>{
      if(sk==='_introLen'||sk==='_quality'){
        const av=sk==='_quality'?(typeof a._quality==='object'?a._quality?.score||0:0):a._introLen||0;
        const bv=sk==='_quality'?(typeof b._quality==='object'?b._quality?.score||0:0):b._introLen||0;
        return sasc?av-bv:bv-av;
      }
      let av=a[sk]||'',bv=b[sk]||'';
      if(Array.isArray(av))av=av.join(' ');if(Array.isArray(bv))bv=bv.join(' ');
      return sasc?String(av).localeCompare(String(bv)):String(bv).localeCompare(String(av));
    });
  }

  self.postMessage(res);
};`;

let _filterWorker=null;
let _workerPending=null;

function initWorker(){
  try{
    const blob=new Blob([_workerCode],{type:'application/javascript'});
    _filterWorker=new Worker(URL.createObjectURL(blob));
    _filterWorker.onmessage=e=>{
      _vtRows=e.data;
      document.getElementById('dtCnt').textContent=
        _vtRows.length+' / '+allData.length+(_ftMode?' [FT]':'');
      document.getElementById('noRows').classList.toggle('hidden',_vtRows.length>0);
      _renderVirtualRows();
      _workerPending=null;
    };
  }catch(err){
    console.warn('Web Worker nedostupný, fallback na hlavní vlákno',err);
    _filterWorker=null;
  }
}

/* ── RIPPLE ───────────────────────────────────────────────────────────────── */
function initRipple(){
  document.addEventListener('click',e=>{
    const btn=e.target.closest('.btn,.btn-sm');
    if(!btn) return;
    const circle=document.createElement('span');
    circle.className='ripple-circle';
    const rect=btn.getBoundingClientRect();
    const size=Math.max(rect.width,rect.height)*2;
    circle.style.cssText=`width:${size}px;height:${size}px;left:${e.clientX-rect.left-size/2}px;top:${e.clientY-rect.top-size/2}px`;
    btn.appendChild(circle);
    circle.addEventListener('animationend',()=>circle.remove());
  });
}


/* ── SMART TOOLTIPS ──────────────────────────────────────────────────────── */
function initTooltips(){
  const tip = document.getElementById('globalTooltip');
  let _tipTimer = null;
  let _cursorX = 0, _cursorY = 0;

  document.addEventListener('mouseover', e => {
    const el = e.target.closest('[data-tip]');
    if(!el) return;
    clearTimeout(_tipTimer);
    _tipTimer = setTimeout(() => {
      const text = el.getAttribute('data-tip');
      if(!text) return;
      tip.textContent = text;
      // Zobrazit mimo viewport aby se změřily rozměry
      tip.style.left = '-9999px';
      tip.style.top  = '-9999px';
      tip.classList.add('visible');
      // Změřit skutečné rozměry a pak správně umístit
      const tw = tip.offsetWidth;
      const th = tip.offsetHeight;
      placeTip(tw, th);
    }, 1500);
  });

  document.addEventListener('mousemove', e => {
    _cursorX = e.clientX;
    _cursorY = e.clientY;
    if(tip.classList.contains('visible')){
      const tw = tip.offsetWidth;
      const th = tip.offsetHeight;
      placeTip(tw, th);
    }
  });

  document.addEventListener('mouseout', e => {
    const el = e.target.closest('[data-tip]');
    if(el && !el.contains(e.relatedTarget)){
      clearTimeout(_tipTimer);
      tip.classList.remove('visible');
    }
  });

  // Skrýt při scrollu nebo kliku
  document.addEventListener('scroll', () => { clearTimeout(_tipTimer); tip.classList.remove('visible'); }, true);
  document.addEventListener('mousedown', () => { clearTimeout(_tipTimer); tip.classList.remove('visible'); });

  function placeTip(tw, th){
    const pad = 12;
    const vw = window.innerWidth;
    const vh = window.innerHeight;
    let x = _cursorX + pad;
    let y = _cursorY - th - pad;   // defaultně nad kurzorem

    // Nepřetékat nahoře → dát pod kurzor
    if(y < 8) y = _cursorY + pad;
    // Nepřetékat dole
    if(y + th > vh - 8) y = vh - th - 8;
    // Nepřetékat vpravo → dát vlevo od kurzoru
    if(x + tw > vw - 8) x = _cursorX - tw - pad;
    // Nepřetékat vlevo
    if(x < 8) x = 8;

    tip.style.left = Math.round(x) + 'px';
    tip.style.top  = Math.round(y) + 'px';
  }
}

/* ── COLLAPSIBLE SECTIONS ─────────────────────────────────────────────────── */
const COLLAPSED_KEY = 'wikiCollapsed';

function loadCollapsed(){
  try { return JSON.parse(localStorage.getItem(COLLAPSED_KEY)||'{}'); } catch(e){ return {}; }
}
function saveCollapsed(state){
  try { localStorage.setItem(COLLAPSED_KEY, JSON.stringify(state)); } catch(e){}
}

function toggleSec(el){
  // el = kliknuto na sec-header nebo export-title.sec-header
  const section = el.closest('.collapsible-section');
  if(!section) return;
  const id = section.dataset.sec;
  const state = loadCollapsed();
  const isCollapsed = section.classList.toggle('collapsed');
  if(id){
    state[id] = isCollapsed;
    saveCollapsed(state);
  }
}

function initCollapsible(){
  const state = loadCollapsed();

  // Sidebar sekce — sec-header
  document.querySelectorAll('.sec-header').forEach(header => {
    header.addEventListener('click', e => {
      // Neklapsat pokud kliknuto na shuffle btn nebo jiny child button/span
      if(e.target.classList.contains('ex-shuffle')) return;
      toggleSec(header);
    });
  });

  // Obnovit stav z localStorage
  document.querySelectorAll('.collapsible-section[data-sec]').forEach(sec => {
    const id = sec.dataset.sec;
    if(state[id]) sec.classList.add('collapsed');
  });

  // Viz karty — title klikatelný
  document.addEventListener('click', e => {
    const title = e.target.closest('.viz-card-title');
    if(!title) return;
    const card = title.closest('.viz-card');
    if(!card) return;
    card.classList.toggle('collapsed');
  });
}


/* ── RESIZE HANDLES ───────────────────────────────────────────────────────── */
const RESIZE_KEY = 'wikiResize';

function loadResizeState(){
  try{ return JSON.parse(localStorage.getItem(RESIZE_KEY)||'{}'); }catch(e){ return {}; }
}
function saveResizeState(state){
  try{ localStorage.setItem(RESIZE_KEY, JSON.stringify(state)); }catch(e){}
}

function initResizeHandles(){
  const state = loadResizeState();

  // ── Sidebar šířka ─────────────────────────────────────────────────────────
  const sideHandle = document.getElementById('sidebarResizeHandle');
  const bodyGrid   = document.getElementById('bodyGrid');
  const MIN_SIDE   = 200, MAX_SIDE = 520;

  if(state.sidebarW){
    bodyGrid.style.gridTemplateColumns = state.sidebarW + 'px 1fr';
  }

  sideHandle.addEventListener('mousedown', e => {
    e.preventDefault();
    const startX   = e.clientX;
    const startW   = bodyGrid.querySelector('.sidebar').getBoundingClientRect().width;
    document.body.classList.add('resizing');
    sideHandle.classList.add('dragging');

    function onMove(e){
      const w = Math.min(MAX_SIDE, Math.max(MIN_SIDE, startW + e.clientX - startX));
      bodyGrid.style.gridTemplateColumns = w + 'px 1fr';
    }
    function onUp(){
      document.body.classList.remove('resizing');
      sideHandle.classList.remove('dragging');
      const w = parseInt(bodyGrid.style.gridTemplateColumns);
      if(w){ const st=loadResizeState(); st.sidebarW=w; saveResizeState(st); }
      document.removeEventListener('mousemove', onMove);
      document.removeEventListener('mouseup', onUp);
    }
    document.addEventListener('mousemove', onMove);
    document.addEventListener('mouseup', onUp);
  });

  // ── Konzole výška ─────────────────────────────────────────────────────────
  const conHandle = document.getElementById('consoleResizeHandle');
  const conBody   = document.getElementById('conBody');
  const MIN_CON   = 60;
  const MAX_CON   = ()=>Math.round(window.innerHeight * 0.88);

  if(state.consoleH){
    conBody.classList.add('resized');
    conBody.style.setProperty('--console-h', state.consoleH+'px');
    conBody.style.maxHeight = state.consoleH + 'px';
  }

  conHandle.addEventListener('mousedown', e => {
    e.preventDefault();
    if(!conOpen){ toggleConsole(); }  // otevrit pokud je zavrena
    const startY  = e.clientY;
    const startH  = conBody.getBoundingClientRect().height || 140;
    document.body.classList.add('resizing-y');
    conHandle.classList.add('dragging');

    function onMove(e){
      const h = Math.min(MAX_CON(), Math.max(MIN_CON, startH + e.clientY - startY));
      conBody.style.setProperty('--console-h', h+'px');
      conBody.classList.add('resized');
      conBody.style.maxHeight = h + 'px';
      conBody.style.transition = 'none';
    }
    function onUp(){
      document.body.classList.remove('resizing-y');
      conHandle.classList.remove('dragging');
      conBody.style.transition = '';
      const h = parseInt(conBody.style.maxHeight);
      if(h){ const st=loadResizeState(); st.consoleH=h; saveResizeState(st);
        // Přepnout .open max-height přes CSS proměnnou
        document.documentElement.style.setProperty('--console-saved-h', h+'px');
      }
      document.removeEventListener('mousemove', onMove);
      document.removeEventListener('mouseup', onUp);
    }
    document.addEventListener('mousemove', onMove);
    document.addEventListener('mouseup', onUp);
  });

  // ── Sloupce tabulky ───────────────────────────────────────────────────────
  const COL_MIN = 30;

  // Nacist ulozene sirky
  if(state.colWidths){
    Object.entries(state.colWidths).forEach(([col, w]) => {
      document.querySelectorAll('.col-c'+col).forEach(el => {
        el.style.width = w + 'px';
      });
    });
  }

  document.querySelectorAll('.th-resize').forEach(handle => {
    handle.addEventListener('mousedown', e => {
      e.preventDefault();
      e.stopPropagation();   // nekliknout sortBy
      const col    = handle.dataset.col;
      const th     = handle.closest('th');
      const startX = e.clientX;
      const startW = th.getBoundingClientRect().width;
      document.body.classList.add('resizing');
      handle.classList.add('dragging');

      function onMove(e){
        const w = Math.max(COL_MIN, startW + e.clientX - startX);
        document.querySelectorAll('.col-c'+col).forEach(el => el.style.width = w+'px');
      }
      function onUp(){
        document.body.classList.remove('resizing');
        handle.classList.remove('dragging');
        const w = th.getBoundingClientRect().width;
        const st = loadResizeState();
        if(!st.colWidths) st.colWidths = {};
        st.colWidths[col] = Math.round(w);
        saveResizeState(st);
        document.removeEventListener('mousemove', onMove);
        document.removeEventListener('mouseup', onUp);
      }
      document.addEventListener('mousemove', onMove);
      document.addEventListener('mouseup', onUp);
    });
  });

  // Dvojklik na th-resize = reset sloupce na výchozí
  document.querySelectorAll('.th-resize').forEach(handle => {
    handle.addEventListener('dblclick', e => {
      e.stopPropagation();
      const col = handle.dataset.col;
      const defaults = {1:'26%',2:'28%',3:'18%',4:'52px',5:'42px',6:'72px'};
      document.querySelectorAll('.col-c'+col).forEach(el => el.style.width = defaults[col]||'');
      const st = loadResizeState();
      if(st.colWidths) delete st.colWidths[col];
      saveResizeState(st);
    });
  });
}


/* ── SESSION PREFIX ───────────────────────────────────────────────────────── */
// Kazdy prohlizec dostane unikatni prefix — zabraní kolizím na sdíleném serveru
const SESSION_KEY = 'wikiSessionId';
function getSessionId(){
  let id = sessionStorage.getItem(SESSION_KEY);
  if(!id){
    // Kratke UUID — 8 hex znaku, dostatecne unikatni pro session izolaci
    id = Math.random().toString(36).slice(2,10) + Math.random().toString(36).slice(2,6);
    sessionStorage.setItem(SESSION_KEY, id);
  }
  return id;
}
// Prefixuje nazev souboru session ID — wiki_data -> s7x3k2_wiki_data
function sessionOutput(name){
  return getSessionId() + '_' + (name||'wiki_data');
}

/* ── INIT ─────────────────────────────────────────────────────────────────── */
renderFields();
renderEx();
checkCP();
renderHistory();
initTheme();
initNotifications();
initKeyboard();
initReviewKeyboard();
initProjects();
checkSavedReview();
initOnboarding();
buildGlobalIndex();
initCollapsible();
initResizeHandles();
initTooltips();
initMobile();
document.getElementById("sessionBadge").textContent=getSessionId();
initSidebarResize();
initWorker();
initRipple();
setStep(0);
document.getElementById('btnRun').classList.remove('hidden');
document.getElementById('shuffleBtn').onclick=renderEx;

// Virtual scroll + tbl keyboard — listener na .dt-wrap
(()=>{
  const wrap=document.querySelector('.dt-wrap');
  if(wrap){
    wrap.addEventListener('scroll',()=>{
      if(_vtRows.length>0) _renderVirtualRows();
    },{passive:true});
    initTblKeyboard();
  }
})();
</script>
</body>
</html>"""


@app.route("/download_multiple")
def download_multiple():
    outputs_raw = request.args.get("outputs","")
    outputs = [o.strip() for o in outputs_raw.split(",") if o.strip()]
    if not outputs:
        return "Žádné soubory", 400
    files = []
    for output in outputs:
        for f in OUTPUT_DIR.iterdir():
            if f.stem == output and f.suffix in (".json", ".csv"):
                files.append(f)
    if not files:
        return "Žádné soubory nenalezeny", 404
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        [zf.write(f, f.name) for f in files]
    buf.seek(0)
    return send_file(buf, mimetype="application/zip",
                     as_attachment=True, download_name="wiki_results_all.zip")



@app.route("/export_sqlite", methods=["POST"])
def export_sqlite():
    d = request.get_json(silent=True) or {}
    log_audit("export_sqlite", str(d.get("outputs","")), d.get("s",""))
    output = request.get_json().get("output","wiki_data")
    json_path = OUTPUT_DIR / f"{output}.json"
    if not json_path.exists():
        return jsonify({"error": f"{output}.json nenalezen"}), 404
    try:
        import sys as _sys
        scraper_dir = Path(__file__).parent
        if str(scraper_dir) not in _sys.path:
            _sys.path.insert(0, str(scraper_dir))
        import wiki_scraper as ws
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        db_path = OUTPUT_DIR / f"{output}.db"
        ws.save_sqlite(data, str(db_path))
        return jsonify({"ok": True, "path": str(db_path), "count": len(data)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/export_readme", methods=["POST"])
def export_readme():
    d = request.get_json()
    output = d.get("output","wiki_data")
    source_url = d.get("url","")
    json_path = OUTPUT_DIR / f"{output}.json"
    if not json_path.exists():
        return jsonify({"error": f"{output}.json nenalezen"}), 404
    try:
        import sys as _sys
        scraper_dir = Path(__file__).parent
        if str(scraper_dir) not in _sys.path:
            _sys.path.insert(0, str(scraper_dir))
        import wiki_scraper as ws
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        fields = set(k for r in data for k in r.keys()
                     if k not in ("_quality","is_stub","redirected_from","error","url","title"))
        readme_path = OUTPUT_DIR / f"{output}.README.md"
        ws.save_readme(data, str(readme_path), source_url,
                       output, fields)
        return jsonify({"ok": True, "path": str(readme_path)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/list_outputs")
def list_outputs():
    sp = request.args.get("s","")
    files = {}
    for f in OUTPUT_DIR.iterdir():
        if f.suffix == ".json" and not f.stem.endswith((".checkpoint",".pending",".selected_urls")):
            # Pokud je session prefix zadan, vracime jen soubory teto session
            if sp and not f.stem.startswith(sp + "_"):
                continue
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    files[f.stem] = {"count": len(data), "size": f.stat().st_size,
                                     "mtime": f.stat().st_mtime}
            except Exception:
                pass
    return jsonify(files)


@app.route("/merge_files", methods=["POST"])
def merge_files():
    d = request.get_json()
    sources = d.get("sources", [])
    output  = d.get("output", "merged")
    if not sources:
        return jsonify({"error": "Žádné soubory k sloučení"}), 400
    all_records = []
    seen_urls   = set()
    dupes = 0
    for stem in sources:
        p = OUTPUT_DIR / f"{stem}.json"
        if not p.exists(): continue
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            for r in data:
                url = r.get("url","")
                if url and url in seen_urls:
                    dupes += 1
                    continue
                if url: seen_urls.add(url)
                all_records.append(r)
        except Exception:
            pass
    out_path = OUTPUT_DIR / f"{output}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True, "count": len(all_records), "dupes": dupes, "output": output})


@app.route("/get_article_preview")
def get_article_preview():
    """Vrátí title+intro pro hesla dané kategorie z pending souboru."""
    output   = request.args.get("output","")
    cat_name = request.args.get("cat","")
    limit    = int(request.args.get("limit","5"))
    pending  = OUTPUT_DIR / f"{output}.pending.json"
    if not pending.exists():
        return jsonify({"items":[]})
    try:
        data = json.loads(pending.read_text(encoding="utf-8"))
        urls = (data.get("cat_articles") or {}).get(cat_name, [])[:limit]
        # Vrátit jen URL+název (bez stahování obsahu)
        items = [{"url": u, "title": u.split("/wiki/")[-1].replace("_"," ")} for u in urls]
        return jsonify({"items": items, "total": len(
            (data.get("cat_articles") or {}).get(cat_name, []))})
    except Exception as e:
        return jsonify({"items":[], "error": str(e)})


@app.route("/download_file")
def download_file():
    filename = request.args.get("file","")
    if not filename or "/" in filename or "\\" in filename:
        return "Neplatný název", 400
    p = OUTPUT_DIR / filename
    if not p.exists(): return "Soubor nenalezen", 404
    return send_file(p, as_attachment=True, download_name=filename)



@app.route("/favicon.ico")
def favicon():
    ico = Path(__file__).parent / "favicon.ico"
    if ico.exists():
        return send_file(ico, mimetype="image/x-icon")
    return "", 204  # no content — žádná chyba v konzoli




@app.route("/export_txt_zip", methods=["POST"])
def export_txt_zip():
    """Vygeneruje ZIP s jedním .txt souborem na heslo. Infobox pred textem."""
    data = request.get_json()
    outputs = data.get("outputs", [])
    if not outputs:
        return jsonify({"error": "Zadne soubory"}), 400

    all_records = []
    for stem in outputs:
        p = OUTPUT_DIR / f"{stem}.json"
        if not p.exists():
            continue
        try:
            records = json.loads(p.read_text(encoding="utf-8"))
            for r in records:
                r["_source"] = stem
            all_records.extend(records)
        except Exception:
            pass

    if not all_records:
        return jsonify({"error": "Zadna data k exportu"}), 404

    def make_txt(r):
        lines = []
        title = r.get("title") or "Bez nazvu"
        lines.append("=" * 60)
        lines.append(f"NAZEV: {title}")
        lines.append(f"URL:   {r.get('url', '')}")
        if r.get("redirected_from"):
            lines.append(f"REDIRECT: {r['redirected_from']}")
        cats = r.get("categories") or []
        if cats:
            lines.append(f"KATEGORIE: {', '.join(cats)}")
        coords = r.get("coordinates") or ""
        if coords:
            lines.append(f"SOURADNICE: {coords}")
        if r.get("is_stub"):
            lines.append("[STUB CLANEK]")
        q = r.get("_quality")
        if isinstance(q, dict):
            lines.append(f"KVALITA: {q.get('score', '?')}/100")
        lines.append("=" * 60)

        # Infobox pred textem
        infobox = r.get("infobox") or {}
        if infobox:
            lines.append("")
            lines.append("--- INFOBOX ---")
            for k, v in infobox.items():
                lines.append(f"  {k}: {v}")

        # Perex
        intro = r.get("intro") or ""
        if intro:
            lines.append("")
            lines.append("--- PEREX ---")
            lines.append(intro)

        # Sekce (pokud existuji)
        sections = r.get("sections") or {}
        if sections:
            lines.append("")
            lines.append("--- TEXT ---")
            for sec_name, sec_text in sections.items():
                lines.append(f"\n[{sec_name}]")
                lines.append(sec_text)

        # full_text jako fallback pokud neni sections
        elif r.get("full_text"):
            lines.append("")
            lines.append("--- TEXT ---")
            lines.append(r["full_text"])

        # Wikidata
        wd = r.get("wikidata") or {}
        if wd:
            lines.append("")
            lines.append("--- WIKIDATA ---")
            for k, v in wd.items():
                lines.append(f"  {k}: {v}")

        # Chyba
        if r.get("error"):
            lines.append("")
            lines.append(f"[CHYBA STAZENI: {r['error']}]")

        lines.append("")
        return "\n".join(lines)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        used_names = {}
        for r in all_records:
            title = r.get("title") or "bez_nazvu"
            # Sanitize filename — odebrat znaky ktere Windows/Mac nepovoluje
            import re as _re
            safe = _re.sub(r'[<>:"/\\|?*]', '_', title)
            safe = safe.strip(". ")[:80] or "bez_nazvu"
            # Deduplikace nazvu souboru
            if safe in used_names:
                used_names[safe] += 1
                safe = f"{safe}_{used_names[safe]}"
            else:
                used_names[safe] = 0
            txt = make_txt(r)
            zf.writestr(f"{safe}.txt", txt.encode("utf-8"))

    buf.seek(0)
    name = (outputs[0] if len(outputs) == 1 else "wiki_export") + "_texty.zip"
    return send_file(buf, mimetype="application/zip",
                     as_attachment=True, download_name=name)


# ─── LOGIN / LOGOUT ───────────────────────────────────────────────────────────

LOGIN_HTML = """<!DOCTYPE html>
<html lang="cs">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>WikiScraper — přihlášení</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      min-height: 100vh; display: flex; align-items: center; justify-content: center;
      background: #0f1117; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .card {
      background: #1a1d27; border: 1px solid #2a2d3a;
      border-radius: 12px; padding: 40px 36px; width: 360px;
      box-shadow: 0 16px 48px rgba(0,0,0,.5);
    }
    .logo { text-align: center; font-size: 32px; margin-bottom: 8px; }
    h1 { text-align: center; font-size: 18px; font-weight: 600; color: #e2e8f0;
         margin-bottom: 4px; }
    .sub { text-align: center; font-size: 12px; color: #64748b; margin-bottom: 28px; }
    label { display: block; font-size: 12px; color: #94a3b8;
            margin-bottom: 6px; margin-top: 16px; }
    input {
      width: 100%; padding: 10px 14px; border-radius: 8px;
      border: 1px solid #2a2d3a; background: #0f1117;
      color: #e2e8f0; font-size: 14px; outline: none;
      transition: border-color .15s;
    }
    input:focus { border-color: #6ee7b7; }
    .error {
      background: rgba(248,113,113,.1); border: 1px solid rgba(248,113,113,.3);
      color: #f87171; border-radius: 8px; padding: 10px 14px;
      font-size: 12px; margin-bottom: 16px; display: {error_display};
    }
    button {
      width: 100%; margin-top: 24px; padding: 11px;
      background: #6ee7b7; color: #0f1117; border: none;
      border-radius: 8px; font-size: 14px; font-weight: 600;
      cursor: pointer; transition: background .15s;
    }
    button:hover { background: #5dd4a4; }
    .footer { text-align: center; margin-top: 20px; font-size: 11px; color: #334155; }
  </style>
</head>
<body>
  <div class="card">
    <div class="logo">📚</div>
    <h1>WikiScraper</h1>
    <p class="sub">Přihlaste se pro přístup</p>
    <details style="margin-bottom:4px">
      <summary style="font-size:11px;color:#64748b;cursor:pointer;padding:6px 2px;user-select:none;list-style:none;display:flex;align-items:center;gap:6px">
        <span>&#9432;</span> Jak to funguje?
      </summary>
      <div style="background:rgba(110,231,183,.07);border:1px solid rgba(110,231,183,.15);border-radius:8px;padding:12px 14px;margin-top:6px;font-size:11px;color:#94a3b8;line-height:1.8">
        Přihlášení odděluje tvoje stahování od ostatních — aby se nikomu nepletly výsledky.<br><br>
        Přihlášení vydrží <strong style="color:#cbd5e1">30 dní</strong>, pak tě aplikace odhlásí. <strong style="color:#cbd5e1">Tvůj účet tím nezanikne</strong> — jen tě příště poprosí o heslo znovu.<br><br>
        Heslo se nikdy neukládá jako text — aplikace ho okamžitě zašifruje.
      </div>
    </details>
    <div class="error" style="display:{error_display}">{error_msg}</div>
    <form method="POST" action="login">
      <input type="hidden" name="next" value="{next_url}">
      <label for="username">Uživatelské jméno</label>
      <input id="username" name="username" type="text" autocomplete="username"
             autofocus placeholder="Jméno" required>
      <label for="password">Heslo</label>
      <input id="password" name="password" type="password"
             autocomplete="current-password" placeholder="••••••••" required>
      <button type="submit">Přihlásit se →</button>
    </form>
    <p class="footer">WikiScraper v8</p>
  </div>
</body>
</html>"""


@app.route("/login", methods=["GET", "POST"])
def login_page():
    next_url = request.args.get("next") or request.form.get("next") or "/"
    # Pokud je already logged in, přesměrovat
    if session.get("user"):
        return redirect(next_url)

    error_msg = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if _check_password(username, password):
            session["user"] = username
            session.permanent = True
            # Admin = prakticky trvalá session (10 let), ostatní 30 dní
            if _is_admin(username):
                session.modified = True
                app.permanent_session_lifetime = datetime.timedelta(days=3650)
            else:
                app.permanent_session_lifetime = datetime.timedelta(days=30)
            log_audit("login_ok", username)
            app_log.info(f"Login OK: {username} from {request.remote_addr}")
            # Po přihlášení: pokud next_url je absolutní URL, vzít jen cestu
            import urllib.parse as _up
            if next_url.startswith("http"):
                _parsed = _up.urlparse(next_url)
                next_url = _parsed.path or "/"
            return redirect(next_url if next_url.startswith("/") else "/")
        else:
            error_msg = "Nesprávné jméno nebo heslo"
            log_audit("login_fail", username)
            app_log.warning(f"Login FAIL: {username} from {request.remote_addr}")

    html = LOGIN_HTML.replace(
        "{error_display}", "block" if error_msg else "none"
    ).replace(
        "{error_msg}", error_msg
    ).replace(
        "{next_url}", next_url if next_url.startswith("/") else "/"
    )
    return Response(html, mimetype="text/html")


@app.route("/logout")
def logout():
    user = session.pop("user", "")
    log_audit("logout", user)
    return _redirect("/login")

# ─── AUTH GUARD ───────────────────────────────────────────────────────────────
PUBLIC_ROUTES = frozenset(["/login", "/logout", "/favicon.ico", "/api/openapi.yaml"])

@app.before_request
def require_login():
    """Každý request projde sem — chráníme vše kromě PUBLIC_ROUTES."""
    if request.path in PUBLIC_ROUTES: return
    if request.path.startswith("/static"): return
    if not session.get("user"):
        return redirect(url_for("login_page", next=request.url))

@app.route("/")
def index(): return HTML


@app.route("/run")
def run_scraper():
    url    = request.args.get("url","")
    depth  = request.args.get("depth","5")
    limit  = request.args.get("limit","0")
    delay  = request.args.get("delay","0.5")
    fmt    = request.args.get("format","both")
    output = request.args.get("output","wiki_data")
    resume = request.args.get("resume","0")=="1"
    fields = request.args.get("fields","title,intro,infobox,categories,coordinates")
    if not url:
        return Response("data:"+json.dumps({"type":"error","msg":"Chybí URL"})+"\n\n",mimetype="text/event-stream")
    scraper = Path(__file__).parent / "wiki_scraper.py"
    if not scraper.exists():
        return Response("data:"+json.dumps({"type":"error","msg":"wiki_scraper.py nenalezen!"})+"\n\n",mimetype="text/event-stream")
    out_path = OUTPUT_DIR / output
    cmd = [sys.executable, str(scraper), url,
           "--depth",depth,"--limit",limit,"--delay",delay,
           "--format",fmt,"--output",str(out_path),
           "--fields",fields,"--verbose"]
    if resume: cmd.append("--resume")

    s = request.args.get("s", "")
    job = get_job(s)

    def generate():
        job["running"] = True
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, bufsize=1, encoding="utf-8", errors="replace")
            job["proc"] = proc
            for line in proc.stdout:
                line = line.rstrip()
                if not line: continue
                try:
                    if line.startswith("PHASE:"):
                        pass
                    elif line.startswith("PHASE1:CAT:"):
                        _,_,cats,subcats,articles,name = line.split(":",5)
                        yield f"data:{json.dumps({'type':'phase1_cat','cats':int(cats),'subcats':int(subcats),'articles':int(articles),'name':name})}\n\n"
                    elif line.startswith("PHASE1:LIST:"):
                        _,_,count,name = line.split(":",3)
                        yield f"data:{json.dumps({'type':'phase1_list','count':int(count),'name':name})}\n\n"
                    elif line.startswith("PHASE1_DONE:"):
                        p = line.split(":")
                        yield f"data:{json.dumps({'type':'phase1_done','total':int(p[1]),'done':int(p[2]),'eta':float(p[3])})}\n\n"
                    elif line.startswith("PROGRESS:"):
                        p = line.split(":",4)
                        eta_v = float(p[3]) if p[3]!="-1" else -1
                        yield f"data:{json.dumps({'type':'progress','i':int(p[1]),'total':int(p[2]),'eta':eta_v,'name':p[4]})}\n\n"
                    elif line.startswith("RATELIMIT:"):
                        _,old,new_d,wait = line.split(":")
                        yield f"data:{json.dumps({'type':'ratelimit','old_delay':old,'new_delay':new_d,'wait':wait})}\n\n"
                    elif line.startswith("DELAYOK:"):
                        yield f"data:{json.dumps({'type':'delay_ok','delay':line.split(':')[1]})}\n\n"
                    else:
                        yield f"data:{json.dumps({'type':'raw','msg':line})}\n\n"
                except Exception:
                    yield f"data:{json.dumps({'type':'raw','msg':line})}\n\n"
            proc.wait()
            job["running"] = False
            has_files = any(OUTPUT_DIR.glob(output+".*"))
            yield f"data:{json.dumps({'type':'done','has_files':has_files})}\n\n"
        except Exception as e:
            job["running"] = False
            yield f"data:{json.dumps({'type':'error','msg':str(e)})}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


@app.route("/stop", methods=["POST"])
def stop():
    data = request.get_json(silent=True) or {}
    log_audit("stop", "", data.get("s",""))
    sp = (request.get_json(silent=True) or {}).get("s","")
    job = get_job(sp)
    proc = job.get("proc")
    if proc and proc.poll() is None: proc.terminate()
    job["running"] = False
    return jsonify({"ok":True})


@app.route("/results")
def results():
    file = request.args.get("file","wiki_data")
    path = OUTPUT_DIR / f"{file}.json"
    if not path.exists(): return jsonify([])
    with open(path, encoding="utf-8") as f: return jsonify(json.load(f))


@app.route("/check_cp")
def check_cp():
    output = request.args.get("output","wiki_data")
    # Fáze 2 checkpoint
    cp2 = OUTPUT_DIR / f"{output}.checkpoint.json"
    if cp2.exists():
        try:
            with open(cp2, encoding="utf-8") as f: d=json.load(f)
            return jsonify({"exists":True,"phase":2,
                           "done":len(d.get("done_urls",[])),
                           "total":len(d.get("article_urls",[])),
                           "saved_at":d.get("saved_at","—")})
        except: pass
    # Fáze 1 checkpoint
    cp1 = OUTPUT_DIR / f"{output}.phase1_cp.json"
    if cp1.exists():
        try:
            with open(cp1, encoding="utf-8") as f: d=json.load(f)
            s=d.get("stats",{})
            return jsonify({"exists":True,"phase":1,
                           "cats":s.get("cats",0),
                           "articles":s.get("articles",0),
                           "saved_at":d.get("saved_at","—")})
        except: pass
    return jsonify({"exists":False})


@app.route("/download")
def download():
    log_audit("download_zip", request.args.get("output",""), request.args.get("s",""))
    output = request.args.get("output","")
    if output:
        files=[f for f in OUTPUT_DIR.iterdir() if f.stem==output and f.suffix in (".json",".csv")]
    else:
        files=[f for f in OUTPUT_DIR.iterdir() if f.suffix in (".json",".csv")]
    if not files: return "Žádné soubory",404
    buf=io.BytesIO()
    with zipfile.ZipFile(buf,"w",zipfile.ZIP_DEFLATED) as zf:
        [zf.write(f,f.name) for f in files]
    buf.seek(0)
    return send_file(buf,mimetype="application/zip",as_attachment=True,
                     download_name=(output+"_results.zip" if output else "wiki_results.zip"))


@app.route("/xlsx_columns")
def xlsx_columns():
    """Vrátí infobox klíče seřazené dle četnosti, s prahem 20%."""
    output  = request.args.get("output", "wiki_data")
    thresh  = float(request.args.get("thresh", "0.20"))
    path    = OUTPUT_DIR / f"{output}.json"
    if not path.exists():
        return jsonify({"error": "Soubor nenalezen"}), 404
    with open(path, encoding="utf-8") as f:
        records = json.load(f)
    if not records:
        return jsonify({"cols": [], "total": 0})

    total = len(records)
    counts: dict = {}
    for r in records:
        for k in r.get("infobox", {}).keys():
            counts[k] = counts.get(k, 0) + 1

    # Seřadit dle četnosti, vrátit s % výskytu
    cols_all = sorted(counts.items(), key=lambda x: -x[1])
    cols_filtered = [
        {"key": k, "count": n, "pct": round(n / total * 100, 1)}
        for k, n in cols_all
    ]
    cols_above = [c for c in cols_filtered if c["count"] / total >= thresh]
    return jsonify({
        "total": total,
        "thresh": thresh,
        "cols_above": cols_above,   # splňují práh
        "cols_below": [c for c in cols_filtered if c["count"] / total < thresh],
    })


@app.route("/export_xlsx")
def export_xlsx():
    output      = request.args.get("output","wiki_data")
    selected_ib = request.args.get("ib_cols","")   # čárkou oddělené infobox klíče, prázdné = auto
    thresh      = float(request.args.get("thresh","0.20"))

    json_path = OUTPUT_DIR / f"{output}.json"
    if not json_path.exists():
        return jsonify({"error":f"{output}.json nenalezen v wiki_output/"}),404
    with open(json_path, encoding="utf-8") as f:
        records = json.load(f)
    if not records:
        return jsonify({"error":"Žádná data k exportu"}),400
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active; ws.title = "Data"

        base_cols = ["title","url","intro","categories","coordinates","links"]
        if any("full_text" in r for r in records): base_cols.append("full_text")
        if any("sections"  in r for r in records): base_cols.append("sections_json")

        # Infobox klíče — buď z parametru nebo automaticky s 20% prahem
        if selected_ib:
            ib_keys = [k.strip() for k in selected_ib.split(",") if k.strip()]
        else:
            total = len(records)
            counts: dict = {}
            for r in records:
                for k in r.get("infobox", {}).keys():
                    counts[k] = counts.get(k, 0) + 1
            ib_keys = [k for k, n in sorted(counts.items(), key=lambda x: -x[1])
                       if n / total >= thresh]

        all_cols = base_cols + ib_keys

        col_labels = {"title":"Název","url":"URL","intro":"Perex","categories":"Kategorie",
                      "coordinates":"Souřadnice","links":"Ext. odkazy","full_text":"Plný text","sections_json":"Sekce (JSON)"}
        HDR_FILL = PatternFill("solid", start_color="171B24")
        HDR_FONT = Font(bold=True, color="6EE7B7", name="Arial", size=10)
        BODY     = Font(name="Arial", size=10)
        LINK     = Font(name="Arial", size=10, color="60A5FA", underline="single")
        ALT      = PatternFill("solid", start_color="1E2330")

        for ci,col in enumerate(all_cols,1):
            c=ws.cell(1,ci,col_labels.get(col,col)); c.font=HDR_FONT; c.fill=HDR_FILL
            c.alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[1].height=20; ws.freeze_panes="A2"

        for ri,rec in enumerate(records,2):
            alt=ri%2==0
            for ci,col in enumerate(all_cols,1):
                if   col=="title":        val=rec.get("title","")
                elif col=="url":          val=rec.get("url","")
                elif col=="intro":        val=rec.get("intro","")
                elif col=="categories":   val=" | ".join(rec.get("categories",[]) if isinstance(rec.get("categories"),list) else [])
                elif col=="coordinates":  val=rec.get("coordinates","")
                elif col=="links":        val="; ".join(l.get("url","") for l in rec.get("links",[])[:5])
                elif col=="full_text":    val=(rec.get("full_text","") or "")[:5000]
                elif col=="sections_json":val=json.dumps(rec.get("sections",{}),ensure_ascii=False)[:3000]
                else:                     val=rec.get("infobox",{}).get(col,"")
                c=ws.cell(ri,ci,val); c.font=LINK if col=="url" else BODY
                c.alignment=Alignment(wrap_text=False,vertical="top")
                if alt: c.fill=ALT

        widths={"title":35,"url":45,"intro":55,"categories":35,"coordinates":20,"links":50,"full_text":60,"sections_json":50}
        for ci,col in enumerate(all_cols,1):
            ws.column_dimensions[get_column_letter(ci)].width=widths.get(col,22)
        ws.auto_filter.ref=f"A1:{get_column_letter(len(all_cols))}1"

        # Přehled list
        ws2=wb.create_sheet("Přehled"); ws2["A1"]="Wiki Scraper — shrnutí"
        ws2["A1"].font=Font(bold=True,name="Arial",size=13,color="6EE7B7")
        for ri,(lbl,val) in enumerate([
            ("Celkem hesel",len(records)),
            ("S infoboxem",sum(1 for r in records if r.get("infobox"))),
            ("Se souřadnicemi",sum(1 for r in records if r.get("coordinates"))),
            ("S plným textem",sum(1 for r in records if r.get("full_text"))),
        ],3):
            ws2.cell(ri,1,lbl).font=Font(name="Arial",size=10,bold=True)
            ws2.cell(ri,2,val).font=Font(name="Arial",size=10,color="4FFFB0")
        ws2.column_dimensions["A"].width=25; ws2.column_dimensions["B"].width=12

        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"{output}.xlsx")
    except ImportError:
        return jsonify({"error":"openpyxl není nainstalován — spusť: pip install openpyxl"}),500
    except Exception as e:
        return jsonify({"error":str(e)}),500



@app.route("/test_url")
def test_url():
    url = request.args.get("url","")
    if not url:
        return jsonify({"ok":False,"error":"Chybí URL"})
    try:
        import requests as req
        r = req.get(url, timeout=10, headers={"User-Agent":"WikiScraper/1.0"})
        r.raise_for_status()
        from bs4 import BeautifulSoup
        from urllib.parse import unquote, urlparse
        soup = BeautifulSoup(r.text, "html.parser")
        h1 = soup.find("h1", id="firstHeading") or soup.find("h1")
        title = h1.get_text(strip=True) if h1 else url

        path = urlparse(url).path.lower()
        is_cat = any(p in path for p in ("kategorie:","category:","catégorie:","categoría:","categoria:"))
        is_cat = is_cat or bool(soup.find("div", id="mw-subcategories") or soup.find("div", id="mw-pages"))

        cats = len(soup.find("div", id="mw-subcategories").find_all("a")) if soup.find("div", id="mw-subcategories") else 0
        pages_div = soup.find("div", id="mw-pages")
        pages = len(pages_div.find_all("a")) if pages_div else 0

        # Suggest output filename from title
        import unicodedata
        raw_slug = unquote(urlparse(url).path.split("/")[-1]).replace(" ","_").lower()
        # Prevest na ASCII — odstranit hacky a carky
        raw_slug = unicodedata.normalize("NFKD", raw_slug)
        raw_slug = raw_slug.encode("ascii", errors="ignore").decode("ascii")
        slug = re.sub(r"[^\w_]","",raw_slug)[:30]

        return jsonify({
            "ok": True,
            "type": "Kategorie" if is_cat else "Seznam hesel",
            "title": title,
            "cats": cats,
            "pages": pages,
            "suggested_output": slug or "wiki_data"
        })
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})


@app.route("/rename_file", methods=["POST"])
def rename_file():
    data = request.get_json()
    old_name = data.get("old","")
    new_name = data.get("new","")
    if not old_name or not new_name:
        return jsonify({"ok":False,"error":"Chybí název"})
    # Sanitize
    new_name = re.sub(r"[^\w_\-]","_",new_name)[:60]
    renamed = []
    errors  = []
    for suffix in [".json",".csv",".checkpoint.json",".pending.json"]:
        src = OUTPUT_DIR / f"{old_name}{suffix}"
        dst = OUTPUT_DIR / f"{new_name}{suffix}"
        if src.exists():
            try:
                src.rename(dst)
                renamed.append(suffix)
            except Exception as e:
                errors.append(str(e))
    if errors:
        return jsonify({"ok":False,"error":", ".join(errors)})
    return jsonify({"ok":True,"renamed":renamed,"new":new_name})




@app.route("/run_phase1")
def run_phase1():
    log_audit("phase1_start", request.args.get("url","")[:100], request.args.get("s",""))
    """Spustí jen fázi 1 — sbírání URL. GUI pak ukáže review."""
    url    = request.args.get("url","")
    depth  = request.args.get("depth","5")
    limit  = request.args.get("limit","0")
    delay  = request.args.get("delay","0.5")
    output = request.args.get("output","wiki_data")
    sp     = request.args.get("s","")
    job    = get_job(sp)

    if not url:
        return Response("data:"+json.dumps({"type":"error","msg":"Chybí URL"})+"\n\n",mimetype="text/event-stream")
    scraper = Path(__file__).parent / "wiki_scraper.py"
    if not scraper.exists():
        return Response("data:"+json.dumps({"type":"error","msg":"wiki_scraper.py nenalezen!"})+"\n\n",mimetype="text/event-stream")

    out_path = OUTPUT_DIR / output
    cmd = [sys.executable, str(scraper), url,
           "--depth",depth,"--limit",limit,"--delay",delay,
           "--output",str(out_path),"--phase1-only","--verbose"]

    def generate():
        job["running"] = True
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, bufsize=1, encoding="utf-8", errors="replace")
            job["proc"] = proc
            for line in proc.stdout:
                line = line.rstrip()
                if not line: continue
                try:
                    if line.startswith("PHASE1:CAT:"):
                        _,_,cats,subcats,articles,name = line.split(":",5)
                        yield f"data:{json.dumps({'type':'phase1_cat','cats':int(cats),'subcats':int(subcats),'articles':int(articles),'name':name})}\n\n"
                        yield f"data:{json.dumps({'type':'raw','msg':f'  📂 {name}  ({articles} hesel, {subcats} podkat)'})}\n\n"
                    elif line.startswith("PHASE1:LIST:"):
                        _,_,count,name = line.split(":",3)
                        yield f"data:{json.dumps({'type':'phase1_list','count':int(count),'name':name})}\n\n"
                        if int(count) % 20 == 0:
                            yield f"data:{json.dumps({'type':'raw','msg':f'  🔗 {count} hesel nalezeno...'})}\n\n"
                    elif line.startswith("PENDING_SAVED:"):
                        parts = line.split(":", 2)  # max 2 splits
                        try:
                            total = int(parts[1])
                            yield f"data:{json.dumps({'type':'pending_saved','total':total})}\n\n"
                            yield f"data:{json.dumps({'type':'raw','msg':f'✅ Fáze 1 hotová — {total} hesel'})}\n\n"
                        except (ValueError, IndexError):
                            pass
                    elif line.startswith("RATELIMIT:"):
                        _,old,new_d,wait = line.split(":")
                        yield f"data:{json.dumps({'type':'ratelimit','old_delay':old,'new_delay':new_d,'wait':wait})}\n\n"
                        yield f"data:{json.dumps({'type':'raw','msg':f'⏱ Rate limit — delay {old}s → {new_d}s, čekám {wait}s'})}\n\n"
                    else:
                        yield f"data:{json.dumps({'type':'raw','msg':line})}\n\n"
                except Exception:
                    yield f"data:{json.dumps({'type':'raw','msg':line})}\n\n"
            proc.wait()
            job["running"] = False
            # Signál že fáze 1 hotová — GUI načte pending
            yield f"data:{json.dumps({'type':'phase1_complete','output':output})}\n\n"
        except Exception as e:
            job["running"] = False
            yield f"data:{json.dumps({'type':'error','msg':str(e)})}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


@app.route("/save_urls", methods=["POST"])
def save_urls():
    data   = request.get_json()
    output = data.get("output","wiki_data")
    urls   = data.get("urls",[])
    path   = OUTPUT_DIR / f"{output}.selected_urls.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(urls, f, ensure_ascii=False)
    return jsonify({"ok":True, "count":len(urls)})


@app.route("/run_phase2_sse")
def run_phase2_sse():
    output      = request.args.get("output","wiki_data")
    fmt         = request.args.get("format","both")
    delay       = request.args.get("delay","0.5")
    fields      = request.args.get("fields","title,intro,infobox,categories,coordinates")
    workers     = request.args.get("workers","1")
    use_api     = request.args.get("api","0") == "1"
    do_tags     = request.args.get("tags","0") == "1"
    do_wikidata = request.args.get("wikidata","0") == "1"
    incremental = request.args.get("incremental","0") == "1"
    sp          = request.args.get("s","")
    job         = get_job(sp)

    url_file = OUTPUT_DIR / f"{output}.selected_urls.json"
    if not url_file.exists():
        return Response("data:"+json.dumps({"type":"error","msg":"selected_urls.json nenalezen"})+"\n\n",
                        mimetype="text/event-stream")

    scraper  = Path(__file__).parent / "wiki_scraper.py"
    out_path = OUTPUT_DIR / output
    cmd = [sys.executable, str(scraper), "placeholder",
           "--url-file", str(url_file),
           "--delay", delay, "--format", fmt,
           "--output", str(out_path), "--fields", fields,
           "--workers", workers, "--verbose"]
    if use_api:     cmd.append("--api")
    if do_tags:     cmd.append("--tags")
    if do_wikidata: cmd.append("--wikidata")
    if incremental: cmd.append("--incremental")

    def generate():
        job["running"] = True
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, bufsize=1, encoding="utf-8", errors="replace")
            job["proc"] = proc
            for line in proc.stdout:
                line = line.rstrip()
                if not line: continue
                try:
                    if line.startswith("PHASE1_DONE:"):
                        p = line.split(":")
                        yield f"data:{json.dumps({'type':'phase1_done','total':int(p[1]),'done':int(p[2]),'eta':float(p[3])})}\n\n"
                    elif line.startswith("PROGRESS:"):
                        p = line.split(":",4)
                        eta_v = float(p[3]) if p[3]!="-1" else -1
                        yield f"data:{json.dumps({'type':'progress','i':int(p[1]),'total':int(p[2]),'eta':eta_v,'name':p[4]})}\n\n"
                    elif line.startswith("RATELIMIT:"):
                        _,old,new_d,wait = line.split(":")
                        yield f"data:{json.dumps({'type':'ratelimit','old_delay':old,'new_delay':new_d,'wait':wait})}\n\n"
                    elif line.startswith("DELAYOK:"):
                        yield f"data:{json.dumps({'type':'delay_ok','delay':line.split(':')[1]})}\n\n"
                    else:
                        yield f"data:{json.dumps({'type':'raw','msg':line})}\n\n"
                except Exception:
                    yield f"data:{json.dumps({'type':'raw','msg':line})}\n\n"
            proc.wait()
            job["running"] = False
            url_file.unlink(missing_ok=True)
            (OUTPUT_DIR / f"{output}.pending.json").unlink(missing_ok=True)
            has_files = any(f for f in OUTPUT_DIR.glob(output+".*") if f.suffix in (".json",".csv"))
            yield f"data:{json.dumps({'type':'done','has_files':has_files})}\n\n"
        except Exception as e:
            job["running"] = False
            yield f"data:{json.dumps({'type':'error','msg':str(e)})}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


@app.route("/get_pending")
def get_pending():
    """Vrátí pending soubor pro review."""
    output = request.args.get("output","wiki_data")
    path = OUTPUT_DIR / f"{output}.pending.json"
    if not path.exists():
        return jsonify({"error":"Pending soubor nenalezen"}), 404
    with open(path, encoding="utf-8") as f:
        return jsonify(json.load(f))


@app.route("/discard_checkpoint", methods=["POST"])
def discard_checkpoint():
    d = request.get_json() or {}
    output = d.get("output","wiki_data")
    for suffix in [".checkpoint.json",".pending.json",".selected_urls.json",".phase1_cp.json"]:
        p = OUTPUT_DIR / f"{output}{suffix}"
        p.unlink(missing_ok=True)
    return jsonify({"ok":True})



@app.route("/admin/users")
def admin_users():
    """Správa uživatelů — přístup jen pro admina (první uživatel)."""
    user = session.get("user")
    if not user: return redirect(url_for("login_page"))
    if not _is_admin(user): return Response("Přístup odepřen — jen pro správce.", status=403)
    msg = request.args.get("msg","")
    err = request.args.get("err","")
    # Varování na výchozí heslo
    warn = ('<div style="background:rgba(251,191,36,.1);border:1px solid rgba(251,191,36,.3);'
            'border-radius:8px;padding:12px;font-size:12px;color:#fbbf24;margin-bottom:16px">'
            '&#9888; Výchozí heslo pro "admin" je stále "admin" — změň ho níže!</div>'
            if _check_password("admin","admin") else "")
    rows = ""
    for u in USERS:
        is_me = (u == user)
        if is_me:
            row_badge = '<span style="font-size:10px;padding:2px 7px;border-radius:10px;background:rgba(110,231,183,.1);color:#6ee7b7">ty</span>'
            row_del = ""
        else:
            row_badge = ""
            row_del = "<form method='POST' action='/admin/users/delete' style='display:inline'><input type='hidden' name='username' value='{n}'><button style='background:transparent;border:1px solid #2a2d3a;color:#64748b;border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer'>Smazat</button></form>".replace("{n}", u)
        rows += "<div style='display:flex;align-items:center;gap:8px;padding:10px 0;border-bottom:1px solid #1e2130'><span style='flex:1;font-family:monospace;font-size:13px'>{n} {b}</span>{d}</div>".replace("{n}", u).replace("{b}", row_badge).replace("{d}", row_del)
    msg_html = f'<div style="background:rgba(110,231,183,.1);border:1px solid rgba(110,231,183,.3);border-radius:8px;padding:10px;font-size:12px;color:#6ee7b7;margin-bottom:12px">{msg}</div>' if msg else ""
    err_html = f'<div style="background:rgba(248,113,113,.1);border:1px solid rgba(248,113,113,.3);border-radius:8px;padding:10px;font-size:12px;color:#f87171;margin-bottom:12px">{err}</div>' if err else ""
    html = f"""<!DOCTYPE html><html lang="cs"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Správa uživatelů</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{min-height:100vh;background:#0f1117;font-family:system-ui,sans-serif;color:#e2e8f0;padding:32px 16px}}
.wrap{{max-width:480px;margin:0 auto}}.card{{background:#1a1d27;border:1px solid #2a2d3a;border-radius:12px;padding:24px;margin-bottom:16px}}
h2{{font-size:13px;font-weight:600;color:#94a3b8;margin-bottom:16px;text-transform:uppercase;letter-spacing:.05em}}
label{{display:block;font-size:12px;color:#94a3b8;margin-bottom:6px;margin-top:14px}}
label:first-of-type{{margin-top:0}}
input{{width:100%;padding:9px 13px;border-radius:8px;border:1px solid #2a2d3a;background:#0f1117;color:#e2e8f0;font-size:13px;outline:none}}
input:focus{{border-color:#6ee7b7}}
.btn{{width:100%;margin-top:14px;padding:10px;background:#6ee7b7;color:#0f1117;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer}}
.btn-sm{{background:transparent;border:1px solid #2a2d3a;color:#94a3b8;border-radius:8px;padding:9px 14px;font-size:12px;cursor:pointer}}
.btn-sm:hover{{border-color:#6ee7b7;color:#6ee7b7}}
a{{color:#64748b;text-decoration:none;font-size:12px}}a:hover{{color:#94a3b8}}</style></head>
<body><div class="wrap">
<a href="/">&larr; Zpět do WikiScraper</a>
<h1 style="font-size:20px;font-weight:700;margin:16px 0 4px">&#x1F465; Správa uživatelů</h1>
<p style="font-size:12px;color:#64748b;margin-bottom:24px">Přidávej a odebírej uživatele. Každý má vlastní přihlášení.</p>
{warn}{msg_html}{err_html}
<div class="card"><h2>Uživatelé</h2>{rows}</div>
<div class="card"><h2>Přidat uživatele</h2>
<form method="POST" action="/admin/users/add">
<label>Jméno</label><input name="username" type="text" placeholder="novak" autocomplete="off" required>
<label>Heslo</label><input name="password" type="password" placeholder="••••••••" required>
<button type="submit" class="btn">+ Přidat uživatele</button></form></div>
<div class="card"><h2>Změnit moje heslo</h2>
<form method="POST" action="/admin/users/change_password" style="display:flex;gap:8px">
<input name="new_password" type="password" placeholder="Nové heslo (min. 6 znaků)" required style="flex:1">
<button type="submit" class="btn-sm">Uložit</button></form></div>
</div></body></html>"""
    return Response(html, mimetype="text/html")


@app.route("/admin/users/add", methods=["POST"])
def admin_add_user():
    if not _is_admin(session.get("user","")): return "Přístup odepřen", 403
    u = request.form.get("username","").strip()
    p = request.form.get("password","")
    if not u or not p: return _redirect("/admin/users?err=Jméno+a+heslo+nesmí+být+prázdné")
    if u in USERS: return redirect(f"/admin/users?err=Uživatel+{u}+už+existuje")
    if len(p) < 6: return _redirect("/admin/users?err=Heslo+musí+mít+aspoň+6+znaků")
    USERS[u] = generate_password_hash(p)
    _save_users(USERS)
    log_audit("user_add", u, session["user"])
    return redirect(f"/admin/users?msg=Uživatel+{u}+přidán")


@app.route("/admin/users/delete", methods=["POST"])
def admin_delete_user():
    if not _is_admin(session.get("user","")): return "Přístup odepřen", 403
    u = request.form.get("username","")
    if u == session["user"]: return _redirect("/admin/users?err=Nemůžeš+smazat+sám+sebe")
    if u not in USERS: return _redirect("/admin/users?err=Uživatel+neexistuje")
    del USERS[u]
    _save_users(USERS)
    log_audit("user_delete", u, session["user"])
    return redirect(f"/admin/users?msg=Uživatel+{u}+smazán")


@app.route("/admin/users/change_password", methods=["POST"])
def admin_change_password():
    user = session.get("user","")
    if not user or user not in USERS: return _redirect("/login")
    pw = request.form.get("new_password","")
    if len(pw) < 6:
        return _redirect("/admin/users?err=Heslo+musí+mít+aspoň+6+znaků")
    USERS[user] = generate_password_hash(pw)
    _save_users(USERS)
    log_audit("password_change", user, user)
    return _redirect("/admin/users?msg=Heslo+úspěšně+změněno")


@app.route("/profile")
def profile():
    """Změna hesla pro běžné uživatele (ne-admini)."""
    user = session.get("user")
    if not user: return redirect(url_for("login_page"))
    msg = request.args.get("msg","")
    err = request.args.get("err","")
    msg_html = f'<div style="background:rgba(110,231,183,.1);border:1px solid rgba(110,231,183,.3);border-radius:8px;padding:10px;font-size:12px;color:#6ee7b7;margin-bottom:12px">{msg}</div>' if msg else ""
    err_html = f'<div style="background:rgba(248,113,113,.1);border:1px solid rgba(248,113,113,.3);border-radius:8px;padding:10px;font-size:12px;color:#f87171;margin-bottom:12px">{err}</div>' if err else ""
    html = f"""<!DOCTYPE html><html lang="cs"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Profil</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{min-height:100vh;background:#0f1117;font-family:system-ui,sans-serif;color:#e2e8f0;padding:32px 16px}}
.wrap{{max-width:400px;margin:0 auto}}.card{{background:#1a1d27;border:1px solid #2a2d3a;border-radius:12px;padding:24px}}
label{{display:block;font-size:12px;color:#94a3b8;margin-bottom:6px;margin-top:14px}}label:first-of-type{{margin-top:0}}
input{{width:100%;padding:9px 13px;border-radius:8px;border:1px solid #2a2d3a;background:#0f1117;color:#e2e8f0;font-size:13px;outline:none}}
input:focus{{border-color:#6ee7b7}}.btn{{display:flex;gap:8px;margin-top:14px}}
.btn button{{flex:1;padding:10px;background:#6ee7b7;color:#0f1117;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer}}
a{{color:#64748b;text-decoration:none;font-size:12px}}a:hover{{color:#94a3b8}}</style></head>
<body><div class="wrap">
<a href="/">&larr; Zpět</a>
<h1 style="font-size:20px;font-weight:700;margin:16px 0 4px">Profil: {user}</h1>
<p style="font-size:12px;color:#64748b;margin-bottom:24px">Změna hesla.</p>
{msg_html}{err_html}
<div class="card">
<form method="POST" action="/admin/users/change_password">
<label>Nové heslo</label><input name="new_password" type="password" placeholder="••••••••" required>
<label>Zopakovat heslo</label><input name="confirm_password" type="password" placeholder="••••••••" required>
<div class="btn"><button type="submit">Uložit heslo</button></div>
</form></div></div></body></html>"""
    return Response(html, mimetype="text/html")


@app.route("/profile/change_password", methods=["POST"])
def profile_change_password():
    user = session.get("user","")
    if not user or user not in USERS: return _redirect("/login")
    pw  = request.form.get("new_password","")
    pw2 = request.form.get("confirm_password","")
    if pw != pw2: return _redirect("/profile?err=Hesla+se+neshodují")
    if len(pw) < 6: return _redirect("/profile?err=Heslo+musí+mít+aspoň+6+znaků")
    USERS[user] = generate_password_hash(pw)
    _save_users(USERS)
    log_audit("password_change", user, user)
    return _redirect("/profile?msg=Heslo+úspěšně+změněno")


@app.route("/export_parquet")
def export_parquet():
    output = request.args.get("output", "")
    s = request.args.get("s", "")
    log_audit("export_parquet", output, s)
    json_path = OUTPUT_DIR / f"{output}.json"
    if not json_path.exists():
        return jsonify({"error": "JSON soubor nenalezen"}), 404
    try:
        import pyarrow as pa, pyarrow.parquet as pq
        import io as _bio
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        rows, all_keys_set = [], set()
        for r in data:
            row = {}
            for k, v in r.items():
                if isinstance(v, (dict, list)):
                    row[k+"_json"] = json.dumps(v, ensure_ascii=False) if v else None
                elif isinstance(v, (str, int, float, bool)) or v is None:
                    row[k] = v
                else:
                    row[k] = str(v)
            all_keys_set.update(row.keys())
            rows.append(row)
        all_keys = sorted(all_keys_set)
        for row in rows:
            for k in all_keys: row.setdefault(k, None)
        table = pa.table({k: [row[k] for row in rows] for k in all_keys})
        buf = _bio.BytesIO()
        pq.write_table(table, buf, compression="snappy")
        buf.seek(0)
        return send_file(buf, mimetype="application/octet-stream",
                         as_attachment=True, download_name=f"{output}.parquet")
    except ImportError:
        return jsonify({"error": "Chybí pyarrow — nainstaluj: pip install pyarrow"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/export_jsonld")
def export_jsonld():
    import io as _bio
    output = request.args.get("output", "")
    s = request.args.get("s", "")
    log_audit("export_jsonld", output, s)
    json_path = OUTPUT_DIR / f"{output}.json"
    if not json_path.exists():
        return jsonify({"error": "JSON soubor nenalezen"}), 404
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        graph = []
        for r in data:
            obj = {"@type": "Thing", "@id": r.get("url",""),
                   "name": r.get("title",""), "description": r.get("intro",""),
                   "url": r.get("url","")}
            if r.get("categories"):
                obj["keywords"] = ", ".join(r["categories"])
            if r.get("coordinates_norm"):
                cn = r["coordinates_norm"]
                obj["geo"] = {"@type":"GeoCoordinates",
                              "latitude":cn.get("lat"),"longitude":cn.get("lng")}
            if r.get("infobox"):
                obj["additionalProperty"] = [
                    {"@type":"PropertyValue","name":k,"value":v}
                    for k,v in list(r["infobox"].items())[:20]
                ]
            graph.append(obj)
        doc = {"@context":"https://schema.org","@graph":graph}
        buf = _bio.BytesIO(json.dumps(doc, ensure_ascii=False, indent=2).encode("utf-8"))
        return send_file(buf, mimetype="application/ld+json",
                         as_attachment=True, download_name=f"{output}.jsonld")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/docs")
def api_docs():
    """Swagger UI pro interaktivní dokumentaci API."""
    return Response("""<!DOCTYPE html>
<html>
<head>
  <title>WikiScraper API Docs</title>
  <meta charset="utf-8">
  <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/swagger-ui/5.11.0/swagger-ui.min.css">
</head>
<body>
<div id="swagger-ui"></div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/swagger-ui/5.11.0/swagger-ui-bundle.min.js"></script>
<script>
SwaggerUIBundle({
  url: "/api/openapi.yaml",
  dom_id: "#swagger-ui",
  presets: [SwaggerUIBundle.presets.apis, SwaggerUIBundle.SwaggerUIStandalonePreset],
  layout: "BaseLayout",
  deepLinking: true,
  displayRequestDuration: true,
});
</script>
</body>
</html>""", mimetype="text/html")


@app.route("/api/openapi.yaml")
def api_openapi():
    """Vrátí OpenAPI specifikaci jako YAML."""
    spec_path = Path(__file__).parent / "openapi.yaml"
    if spec_path.exists():
        return Response(spec_path.read_text(encoding="utf-8"), mimetype="application/yaml")
    return Response("# openapi.yaml not found", mimetype="application/yaml"), 404


if __name__=="__main__":
    print("\n"+"─"*44)
    print("  📚  Wiki Scraper GUI v8")
    print("─"*44)
    print("  http://localhost:7842")
    print("  Ukonči: Ctrl+C")
    print("─"*44+"\n")
    _make_app().run(host="0.0.0.0",port=7842,debug=False,threaded=True)
